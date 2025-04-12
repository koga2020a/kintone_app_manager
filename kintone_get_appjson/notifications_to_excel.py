#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
kintoneアプリの通知設定をExcelに出力するスクリプト
"""

import os
import sys
import yaml
import json
import argparse
import logging
from pathlib import Path
from datetime import datetime
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from dotenv import load_dotenv

# 定数定義
SCRIPT_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = SCRIPT_DIR.parent / "output"

def setup_logging():
    """ロギングの設定"""
    log_dir = SCRIPT_DIR.parent / "logs"
    log_dir.mkdir(exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = log_dir / f"notifications_to_excel_{timestamp}.log"
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler()
        ]
    )
    
    return logging.getLogger("notifications_to_excel")

def load_yaml_file(file_path):
    """YAMLファイルを読み込む"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            return yaml.safe_load(f)
    except Exception as e:
        raise Exception(f"YAMLファイルの読み込みに失敗しました: {file_path} - {e}")

def find_app_directory(base_dir, app_id):
    """アプリIDに対応するディレクトリを探す"""
    return next((d for d in base_dir.iterdir() if d.is_dir() and d.name.startswith(f"{app_id}_")), None)

def find_latest_group_user_excel(output_dir):
    """最新のユーザー・グループExcelファイルを探す"""
    excel_files = list(output_dir.glob("kintone_users_groups_*.xlsx"))
    
    if not excel_files:
        return None
    
    # 日時でソートして最新のファイルを返す
    latest_file = max(excel_files, key=lambda f: f.stat().st_mtime)
    return latest_file

def load_group_yaml_data(yaml_path):
    """
    group_user_list.yaml からグループとユーザー情報を読み込む
    
    Args:
        yaml_path: group_user_list.yaml のファイルパス
        
    Returns:
        dict: グループ情報の辞書
    """
    try:
        if not yaml_path or not yaml_path.exists():
            logging.warning(f"グループユーザーリストファイルが見つかりません: {yaml_path}")
            return {}
            
        with open(yaml_path, 'r', encoding='utf-8') as f:
            group_data = yaml.safe_load(f)
            logging.info(f"group_user_list.yaml から {len(group_data)} 件のグループ情報を読み込みました。")
            return group_data
    except Exception as e:
        logging.warning(f"グループユーザーリストの読み込みに失敗しました: {e}")
        return {}

# もう一つのデータソースからグループ情報を読み込む関数を追加
def load_group_list_yaml(yaml_dir):
    """
    グループコードとグループ名のマッピングを読み込む
    """
    try:
        group_list_path = yaml_dir / "group_list.yaml"
        if not group_list_path.exists():
            logging.warning(f"グループリストファイルが見つかりません: {group_list_path}")
            return {}
            
        with open(group_list_path, 'r', encoding='utf-8') as f:
            group_mapping = yaml.safe_load(f)
            return group_mapping
    except Exception as e:
        logging.warning(f"グループリストの読み込みに失敗しました: {e}")
        return {}

# ユーザー情報を読み込む関数を追加
def load_user_list_yaml(yaml_dir):
    """
    ユーザー情報を読み込む
    """
    try:
        user_list_path = yaml_dir / "user_list.yaml"
        if not user_list_path.exists():
            logging.warning(f"ユーザーリストファイルが見つかりません: {user_list_path}")
            return {}
            
        with open(user_list_path, 'r', encoding='utf-8') as f:
            user_data = yaml.safe_load(f)
            return user_data
    except Exception as e:
        logging.warning(f"ユーザーリストの読み込みに失敗しました: {e}")
        return {}

def find_group_user_list_yaml():
    """group_user_list.yamlファイルを探す"""
    # 検索場所のリスト
    search_paths = [
        Path(__file__).resolve().parent.parent,  # プロジェクトルート
        Path(__file__).resolve().parent,  # notifications_to_excel.pyと同じディレクトリ
        OUTPUT_DIR,  # 出力ディレクトリ
        Path.cwd()  # カレントディレクトリ
    ]
    
    # まず group_user_raw_list.yaml を探す
    for path in search_paths:
        yaml_path = path / "group_user_raw_list.yaml"
        if yaml_path.exists():
            logging.info(f"group_user_raw_list.yaml が見つかりました: {yaml_path}")
            return yaml_path
    
    # 見つからなければ group_user_list.yaml を探す
    for path in search_paths:
        yaml_path = path / "group_user_list.yaml"
        if yaml_path.exists():
            logging.info(f"group_user_list.yaml が見つかりました: {yaml_path}")
            return yaml_path
    
    logging.warning("group_user_raw_list.yaml と group_user_list.yaml のどちらも見つかりませんでした")
    return None

def load_field_values_from_tsv(app_dir, field_code):
    """
    records.tsvファイルからフィールドの値一覧を取得する
    
    Args:
        app_dir: アプリディレクトリ
        field_code: フィールドコード
    
    Returns:
        list: フィールドの値一覧
    """
    try:
        tsv_files = list(app_dir.glob("*_records.tsv"))
        if not tsv_files:
            logging.warning(f"records.tsvファイルが見つかりません: {app_dir}")
            return []
        
        # 最新のTSVファイルを使用
        tsv_file = max(tsv_files, key=lambda f: f.stat().st_mtime)
        logging.info(f"records.tsvファイルを読み込みます: {tsv_file}")
        
        # TSVファイルを読み込む
        df = pd.read_csv(tsv_file, sep='\t', encoding='utf-8')
        
        # フィールドコードがヘッダーに含まれるか確認
        if field_code not in df.columns:
            logging.warning(f"フィールド '{field_code}' がTSVファイルに見つかりません")
            return []
        
        # フィールドの値を取得し、ユニークなもののみ抽出
        values = df[field_code].dropna().unique().tolist()
        
        # 最大100個まで
        return values[:100]
    
    except Exception as e:
        logging.warning(f"TSVファイルの読み込みに失敗しました: {e}")
        return []

def add_field_values_reference(ws, row_idx, field_codes, app_dir, header_font, header_fill, header_alignment, thin_border, form_fields=None, group_yaml_data=None):
    """フィールド値の参考一覧を追加"""
    
    if not field_codes or not app_dir:
        return row_idx
    
    # 重複するフィールドコードを除去
    unique_field_codes = list(set(field_codes))
    
    field_header_fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid")  # フィールド用の背景色（薄い青）
    
    for field_code in unique_field_codes:
        values = load_field_values_from_tsv(app_dir, field_code)
        
        if not values:
            continue
        
        # フィールドタイプを取得
        field_type = None
        if form_fields and field_code in form_fields.get('properties', {}):
            field_info = form_fields['properties'][field_code]
            field_type = field_info.get('type', '')
        
        # 見出し
        row_idx += 2
        cell = ws.cell(row=row_idx, column=1)
        cell.value = f"通知先種別：フィールド  フィールドタイプ：{'グループ選択（GROUP_SELECT）' if field_type == 'GROUP_SELECT' else 'ユーザー選択（USER_SELECT）'}"
        cell.font = Font(bold=True, size=12)
        cell.fill = field_header_fill
        row_idx += 1
        
        # フィールドの見出し
        type_info = f" ({field_type})" if field_type else ""
        # A列とB列を結合
        ws.merge_cells(start_row=row_idx, start_column=1, end_row=row_idx, end_column=2)
        cell = ws.cell(row=row_idx, column=1)
        cell.value = f"フィールド名：{field_code}     ※値は過去データより収集)"
        cell.font = Font(bold=True, size=13)
        cell.fill = field_header_fill
        row_idx += 1
        
        # USER_SELECTまたはGROUP_SELECTの場合はヘッダーを追加
        if field_type == 'GROUP_SELECT':
            headers = ["グループ名", "アカウント名", "メールアドレス", "停止中"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            row_idx += 1
        elif field_type == 'USER_SELECT':
            headers = ["", "アカウント名", "メールアドレス", "停止中"]
            for col_idx, header in enumerate(headers, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = header
                cell.font = header_font
                cell.fill = header_fill
                cell.alignment = header_alignment
                cell.border = thin_border
            row_idx += 1
        
        # 縦に表示するタイプのフィールドかどうか
        is_vertical_display = field_type in ['USER_SELECT', 'GROUP_SELECT', 'ORGANIZATION_SELECT']
        
        # データを表示するための準備
        col_count = 0
        current_row = row_idx
        
        for value in values:
            # JSON風の値かどうかをチェック
            is_json = False
            json_objects = []
            group_codes = []  # GROUP_SELECT用のグループコード
            user_codes = []   # USER_SELECT用のユーザーコード
            
            # 特定のフィールドタイプは常に縦表示
            force_vertical = False
            
            # USER_SELECT, GROUP_SELECTは特別処理
            if is_vertical_display:
                # ユーザー選択やグループ選択フィールドの場合、JSON形式でなくても処理
                try:
                    # JSON形式かチェック
                    if isinstance(value, str) and ((value.startswith('{') and 'code' in value and 'name' in value) or
                                                  (field_type == 'USER_SELECT' and '@' in value)):
                        # 1. 標準的なJSON形式の場合
                        if value.startswith('{'):
                            json_parts = value.replace('}, {', '}|{').split('|')
                            
                            for part in json_parts:
                                part = part.replace("'", '"')
                                obj = json.loads(part)
                                if 'code' in obj and 'name' in obj:
                                    json_objects.append(f"{obj['name']}({obj['code']})")
                                    # GROUP_SELECTの場合、グループコードを保存
                                    if field_type == 'GROUP_SELECT':
                                        group_codes.append(obj['code'])
                                    # USER_SELECTの場合、ユーザーコードを保存
                                    elif field_type == 'USER_SELECT':
                                        user_codes.append(obj['code'])
                        
                        # 2. USER_SELECTで特殊な形式の場合（例：user@example.com）
                        elif field_type == 'USER_SELECT' and '@' in value:
                            # メールアドレスをユーザーコードとして扱う
                            json_objects.append(value)
                            user_codes.append(value)
                        
                        if json_objects:
                            is_json = True
                            force_vertical = True
                except:
                    # 処理失敗の場合は通常値として扱う
                    pass
            else:
                # 通常のJSON形式チェック
                try:
                    if isinstance(value, str) and value.startswith('{') and 'code' in value and 'name' in value:
                        json_parts = value.replace('}, {', '}|{').split('|')
                        
                        for part in json_parts:
                            part = part.replace("'", '"')
                            obj = json.loads(part)
                            if 'code' in obj and 'name' in obj:
                                json_objects.append(f"{obj['name']}({obj['code']})")
                        
                        if json_objects:
                            is_json = True
                except:
                    pass
            
            # GROUP_SELECTの特別処理
            if field_type == 'GROUP_SELECT' and is_json and group_yaml_data:
                # グループごとにユーザーを表示
                for idx, (group_obj, group_code) in enumerate(zip(json_objects, group_codes)):
                    group_info = group_yaml_data.get(group_code, {})
                    members = group_info.get('users', [])
                    
                    if not members:
                        # メンバーがいない場合、グループ情報だけ表示
                        cell_a = ws.cell(row=current_row, column=1)
                        cell_a.value = group_obj
                        cell_a.border = thin_border
                        
                        for col in range(2, 5):
                            ws.cell(row=current_row, column=col).border = thin_border
                        
                        current_row += 1
                    else:
                        # メンバーをソート（アクティブユーザーを上部に、同一ドメインでグループ化）
                        members = sort_group_members(members)
                        
                        # グループの最初の行を記録
                        first_row_of_group = current_row
                        
                        # 各メンバーを行に表示
                        for i, member in enumerate(members):
                            # A列: グループ情報（最初のメンバーの行のみ）
                            cell_a = ws.cell(row=current_row, column=1)
                            if i == 0:
                                if '(' in group_obj and ')' in group_obj:
                                    name, code = group_obj.split('(')
                                    code = code.rstrip(')')
                                    cell_a.value = f"{name}\n({code})"
                                    # セル内で改行が表示されるように設定
                                    cell_a.alignment = Alignment(wrap_text=True)
                                else:
                                    cell_a.value = group_obj
                            cell_a.border = thin_border
                            
                            # B列: アカウント名
                            cell_b = ws.cell(row=current_row, column=2)
                            cell_b.value = member.get('username', '')
                            cell_b.border = thin_border
                            
                            # C列: メールアドレス
                            cell_c = ws.cell(row=current_row, column=3)
                            cell_c.value = member.get('email', '')
                            cell_c.border = thin_border
                            
                            # D列: 停止中かどうか
                            cell_d = ws.cell(row=current_row, column=4)
                            cell_d.value = "停止中" if member.get('isDisabled', False) else ""
                            cell_d.border = thin_border
                            
                            # D列が「停止中」の場合、B, C, D列の背景色を淡いグレーに設定
                            if cell_d.value == "停止中":
                                gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                                for col in range(2, 5):  # B, C, D列
                                    ws.cell(row=current_row, column=col).fill = gray_fill
                            
                            current_row += 1
                        
                        # メンバーが複数いる場合、A列を結合
                        last_row_of_group = current_row - 1
                        if last_row_of_group > first_row_of_group:
                            merge_cells_in_column_a(ws, first_row_of_group, last_row_of_group)
            
            # USER_SELECTの特別処理
            elif field_type == 'USER_SELECT' and is_json:
                # ユーザー情報の取得（group_yaml_dataからユーザー情報を探す）
                all_users = {}
                user_objects = []
                
                if group_yaml_data:
                    for group_data in group_yaml_data.values():
                        for user in group_data.get('users', []):
                            user_code = user.get('code', '')
                            if user_code:
                                all_users[user_code] = user
                
                # user_list.yamlからもユーザー情報を取得
                user_yaml_data = load_user_list_yaml(Path(SCRIPT_DIR).parent)
                if user_yaml_data:
                    for user_code, user_info in user_yaml_data.items():
                        if user_code not in all_users:
                            all_users[user_code] = user_info

                # 選択されたユーザーのリストを作成
                for user_code in user_codes:
                    if user_code in all_users:
                        user_objects.append(all_users[user_code])
                    else:
                        # ユーザー情報が見つからない場合は最低限の情報で作成
                        user_objects.append({
                            'username': user_code,
                            'email': '',
                            'isDisabled': False
                        })
                
                # ユーザーをソート
                user_objects = sort_group_members(user_objects)
                
                # グループの最初の行を記録
                first_row_of_group = current_row

                # ユーザーごとに行を作成
                for i, user_info in enumerate(user_objects):
                    # A列: 空白
                    cell_a = ws.cell(row=current_row, column=1)
                    cell_a.border = thin_border
                    
                    # B列: アカウント名
                    cell_b = ws.cell(row=current_row, column=2)
                    cell_b.value = user_info.get('username', '')
                    cell_b.border = thin_border
                    
                    # C列: メールアドレス
                    cell_c = ws.cell(row=current_row, column=3)
                    cell_c.value = user_info.get('email', '')
                    cell_c.border = thin_border
                    
                    # D列: 停止中かどうか
                    cell_d = ws.cell(row=current_row, column=4)
                    cell_d.value = "停止中" if user_info.get('isDisabled', False) else ""
                    cell_d.border = thin_border
                    
                    # D列が「停止中」の場合、B, C, D列の背景色を淡いグレーに設定
                    if cell_d.value == "停止中":
                        gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                        for col in range(2, 5):  # B, C, D列
                            ws.cell(row=current_row, column=col).fill = gray_fill
                    
                    current_row += 1

                # メンバーが複数いる場合、A列を結合
                last_row_of_group = current_row - 1
                if last_row_of_group > first_row_of_group:
                    merge_cells_in_column_a(ws, first_row_of_group, last_row_of_group + 1)
                                
            # 通常のJSONまたは強制縦表示（既にUSER_SELECTとGROUP_SELECTは処理済み）
            elif force_vertical and field_type != 'USER_SELECT' and field_type != 'GROUP_SELECT':
                # 強制縦表示 - 1行に1つずつ表示
                if not json_objects:
                    # JSON解析に失敗した場合、元の値を使用
                    cell = ws.cell(row=current_row, column=1)
                    cell.value = value
                    cell.border = thin_border
                    current_row += 1
                else:
                    # JSON解析に成功した場合
                    for obj_value in json_objects:
                        cell = ws.cell(row=current_row, column=1)
                        cell.value = obj_value
                        cell.border = thin_border
                        current_row += 1
            
            # 通常のJSON表示（ORGANIZATION_SELECTなど縦表示）
            elif is_vertical_display and is_json and field_type != 'USER_SELECT' and field_type != 'GROUP_SELECT':
                # 縦方向に1行に1つずつ表示
                for obj_value in json_objects:
                    cell = ws.cell(row=current_row, column=1)
                    cell.value = obj_value
                    cell.border = thin_border
                    current_row += 1
            
            # 通常のデータ処理（横表示）
            elif not is_json:
                col = col_count % 5 + 1
                cell = ws.cell(row=current_row, column=col)
                cell.value = value
                cell.border = thin_border
                
                col_count += 1
                if col_count % 5 == 0:
                    current_row += 1
            
            # その他のJSON風データ（横表示）
            else:
                col = col_count % 5 + 1
                cell = ws.cell(row=current_row, column=col)
                cell.value = ", ".join(json_objects)
                cell.border = thin_border
                
                col_count += 1
                if col_count % 5 == 0:
                    current_row += 1
        
        # 次のフィールドのために行を進める
        if not force_vertical and not (is_vertical_display and is_json) and col_count % 5 != 0:
            current_row += 1
        row_idx = current_row + 1
    
    return row_idx

def create_notification_excel(app_id, general_data, record_data, reminder_data, form_fields=None, output_file=None, app_dir=None):
    """通知設定をExcelに出力する"""
    
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = OUTPUT_DIR / f"{app_id}_notifications_{timestamp}.xlsx"
    
    # Excelワークブックを作成
    wb = Workbook()
    
    # デフォルトのSheet1を削除
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # スタイル定義
    header_font = Font(bold=True, size=11)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    
    # 罫線スタイル
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # group_user_list.yamlからグループ情報を読み込む
    group_yaml_path = find_group_user_list_yaml()
    group_yaml_data = load_group_yaml_data(group_yaml_path)
    
    # 収集したグループコードのリスト
    collected_group_codes = []
    
    # 1. 一般通知設定のシート作成
    if general_data:
        create_general_notifications_sheet(wb, general_data, header_font, header_fill, header_alignment, thin_border, group_yaml_data, collected_group_codes, form_fields, app_dir)
    
    # 2. レコード通知設定のシート作成
    if record_data:
        create_record_notifications_sheet(wb, record_data, header_font, header_fill, header_alignment, thin_border, group_yaml_data, collected_group_codes, form_fields, app_dir)
    
    # 3. リマインダー通知設定のシート作成
    if reminder_data:
        create_reminder_notifications_sheet(wb, reminder_data, header_font, header_fill, header_alignment, thin_border, group_yaml_data, collected_group_codes)
    
    # Excelファイルを保存
    wb.save(output_file)
    logging.info(f"通知設定をExcelに出力しました: {output_file}")
    
    return output_file

def create_general_notifications_sheet(wb, data, header_font, header_fill, header_alignment, thin_border, group_yaml_data, collected_group_codes, form_fields=None, app_dir=None):
    """一般通知設定のシートを作成"""
    ws = wb.create_sheet(title="一般通知設定")
    
    # A, B, C列の幅を330pxに設定（約47文字分）
    ws.column_dimensions["A"].width = 47
    ws.column_dimensions["B"].width = 47
    ws.column_dimensions["C"].width = 47
    
    # 通知先種別ごとの背景色を定義
    user_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # 薄い赤
    group_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # 薄い緑
    field_fill = PatternFill(start_color="CCCCFF", end_color="CCCCFF", fill_type="solid")  # 薄い青
    
    # ヘッダー行 - フィールドタイプ列を追加
    headers = ["No.", "通知先種別", "フィールドタイプ", "通知先", "フィールドタイプ", "サブグループ含む", "レコード追加", "レコード編集", "コメント追加", "ステータス変更", "ファイル読込"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # データ行
    notifications = data.get("notifications", [])
    
    # データ行の背景色を交互に設定
    light_blue_fill = PatternFill(start_color="EBF1F5", end_color="EBF1F5", fill_type="solid")
    
    # グループの通知先を収集
    group_codes = []
    
    # ユーザーの通知先を収集
    user_codes = []
    
    # フィールドコードの収集
    field_codes = []
    
    for row_idx, notify in enumerate(notifications, 2):
        entity = notify.get("entity", {})
        entity_type = entity.get("type", "")
        entity_code = entity.get("code", "")
        
        # グループコードを収集
        if entity_type == "GROUP":
            group_codes.append(entity_code)
        
        # ユーザーコードを収集
        if entity_type == "USER":
            user_codes.append(entity_code)
        
        # 通知先タイプを日本語に変換
        type_jp = ""
        field_type = ""
        form_field_type = ""  # フォームフィールドから取得するタイプ用の変数
        
        if entity_type == "USER":
            type_jp = "ユーザー"
        elif entity_type == "GROUP":
            type_jp = "グループ"
        elif entity_type == "ORGANIZATION":
            type_jp = "組織"
        elif entity_type == "FIELD_ENTITY":
            type_jp = "フィールド"
            # フィールドタイプの取得
            if "type" in entity:
                if entity.get("type") == "CREATOR":
                    field_type = "作成者"
                elif entity.get("type") == "MODIFIER":
                    field_type = "更新者"
                elif entity.get("type") == "USER_SELECT":
                    field_type = "ユーザー選択"
                elif entity.get("type") == "GROUP_SELECT":
                    field_type = "グループ選択"
                elif entity.get("type") == "ORGANIZATION_SELECT":
                    field_type = "組織選択"
                else:
                    field_type = entity.get("type", "")
            
            # フォームフィールド情報からタイプを取得
            if form_fields and entity_code in form_fields.get('properties', {}):
                field_info = form_fields['properties'][entity_code]
                form_field_type = field_info.get('type', '')
            
            # FIELD_ENTITYの場合、フィールドコードを収集
            field_codes.append(entity_code)
        else:
            type_jp = entity_type
        
        # データを行に設定
        row_data = [
            row_idx - 1,  # No.
            type_jp,  # 通知先タイプ
            field_type,  # フィールドタイプ - 新しい列
            entity_code,  # 通知先
            form_field_type,  # フォームフィールドから取得したフィールドタイプ - E列
            "●" if notify.get("includeSubs", False) else "",  # サブグループ含む
            "●" if notify.get("recordAdded", False) else "",  # レコード追加
            "●" if notify.get("recordEdited", False) else "",  # レコード編集
            "●" if notify.get("commentAdded", False) else "",  # コメント追加
            "●" if notify.get("statusChanged", False) else "",  # ステータス変更
            "●" if notify.get("fileImported", False) else "",  # ファイル読込
        ]
        
        # 行の背景色を交互に設定
        row_fill = light_blue_fill if row_idx % 2 == 0 else None
        
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = value
            cell.border = thin_border
            
            # 通知先種別に応じた背景色を設定（B列）
            if col_idx == 2 and value:
                if value == "ユーザー":
                    cell.fill = user_fill
                elif value == "グループ":
                    cell.fill = group_fill
                elif value == "フィールド":
                    cell.fill = field_fill
                elif row_fill:
                    cell.fill = row_fill
            elif row_fill and col_idx != 2:  # B列以外
                cell.fill = row_fill
                
            if col_idx >= 6:  # チェックボックス的な列は中央揃え
                cell.alignment = Alignment(horizontal='center')
    
    # コメント通知設定
    row_idx = len(notifications) + 3
    ws.cell(row=row_idx, column=1).value = "コメント投稿者への通知:"
    ws.cell(row=row_idx, column=1).font = Font(bold=True)
    ws.cell(row=row_idx, column=2).value = "はい" if data.get("notifyToCommenter", False) else "いいえ"
    ws.cell(row=row_idx, column=2).alignment = Alignment(horizontal='center')
    
    # セクション背景色
    section_fill = PatternFill(start_color="DCE6F1", end_color="DCE6F1", fill_type="solid")
    for col_idx in range(1, 3):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.fill = section_fill
        cell.border = thin_border
    
    # 列幅の調整
    for col_idx in range(4, len(headers) + 1):  # A, B, C列は既に設定済みなので4列目から設定
        column_letter = get_column_letter(col_idx)
        if col_idx == 4:  # 通知先列
            ws.column_dimensions[column_letter].width = 20
        elif col_idx == 5:  # フィールドタイプ列（新しく追加）
            ws.column_dimensions[column_letter].width = 15
        else:  # その他の列
            ws.column_dimensions[column_letter].width = 15
    
    # グループメンバー情報を追加
    if group_codes:
        row_idx = add_group_members_table(ws, row_idx, group_codes, header_font, header_fill, header_alignment, thin_border, group_yaml_data, collected_group_codes)
    
    # ユーザー情報を追加
    if user_codes:
        user_yaml_data = load_user_list_yaml(Path(SCRIPT_DIR).parent)
        row_idx = add_user_information_table(ws, row_idx, user_codes, header_font, header_fill, header_alignment, thin_border, user_yaml_data)
    
    # フィールド値の参考一覧を追加
    if field_codes and app_dir:
        row_idx = add_field_values_reference(ws, row_idx, field_codes, app_dir, header_font, header_fill, header_alignment, thin_border, form_fields, group_yaml_data)

def create_record_notifications_sheet(wb, data, header_font, header_fill, header_alignment, thin_border, group_yaml_data, collected_group_codes, form_fields=None, app_dir=None):
    """レコード通知設定のシートを作成"""
    ws = wb.create_sheet(title="レコード通知設定")
    
    # ヘッダー設定
    headers = ["No.", "通知タイトル", "通知条件", "通知先種別", "通知先", "フィールドタイプ", "下位組織継承"]
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 列幅の設定 - A, B, C列を330pxに設定（約47文字分）
    ws.column_dimensions['A'].width = 47
    ws.column_dimensions['B'].width = 47
    ws.column_dimensions['C'].width = 47
    ws.column_dimensions['D'].width = 15
    ws.column_dimensions['E'].width = 30
    ws.column_dimensions['F'].width = 15
    ws.column_dimensions['G'].width = 15
    
    # データの書き込み
    row = 2
    current_notification_id = None
    first_row_of_notification = None
    
    # フィールドコードの収集
    field_codes = []
    
    for idx, notification in enumerate(data.get('notifications', []), 1):
        title = notification.get('title', '')
        condition = notification.get('filterCond', '')
        
        # 現在の通知IDを設定
        current_notification_id = idx
        first_row_of_notification = row
        
        # 通知先ごとに行を作成
        for target_idx, target in enumerate(notification.get('targets', []), 0):
            entity = target.get('entity', {})
            entity_type = entity.get('type', '')
            entity_code = entity.get('code', '')
            include_subs = target.get('includeSubs', False)
            
            # グループコードを収集
            if entity_type == "GROUP":
                collected_group_codes.append(entity_code)
            
            # 通知先タイプを日本語に変換
            type_jp = ""
            field_type = ""
            if entity_type == "USER":
                type_jp = "ユーザー"
            elif entity_type == "GROUP":
                type_jp = "グループ"
            elif entity_type == "ORGANIZATION":
                type_jp = "組織"
            elif entity_type == "FIELD_ENTITY":
                type_jp = "フィールド"
                # フォームフィールド情報からタイプを取得
                if form_fields and entity_code in form_fields.get('properties', {}):
                    field_info = form_fields['properties'][entity_code]
                    field_type = field_info.get('type', '')
            
            # データを書き込み
            cells = [
                (row, 1, idx if target_idx == 0 else None),
                (row, 2, title if target_idx == 0 else None),
                (row, 3, condition if target_idx == 0 else None),
                (row, 4, type_jp),
                (row, 5, entity_code),
                (row, 6, field_type),
                (row, 7, "継承する" if include_subs else "継承しない")
            ]
            
            for r, c, value in cells:
                cell = ws.cell(row=r, column=c, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(vertical='center', wrap_text=True)
            
            row += 1
        
        # 同じ通知IDの行が複数ある場合、A列、B列、C列を結合
        if row > first_row_of_notification + 1:
            for col in range(1, 4):  # A, B, C列
                ws.merge_cells(
                    start_row=first_row_of_notification,
                    start_column=col,
                    end_row=row - 1,
                    end_column=col
                )
        
        # フィールドコードの収集
        for target in notification.get('targets', []):
            entity = target.get('entity', {})
            entity_type = entity.get('type', '')
            entity_code = entity.get('code', '')
            
            # FIELD_ENTITYの場合、フィールドコードを収集
            if entity_type == "FIELD_ENTITY":
                field_codes.append(entity_code)
    
    # フィールド値の参考一覧を追加
    if field_codes and app_dir:
        row_idx = row
        row_idx = add_field_values_reference(ws, row_idx, field_codes, app_dir, header_font, header_fill, header_alignment, thin_border, form_fields, group_yaml_data)
    
    return ws

def create_reminder_notifications_sheet(wb, data, header_font, header_fill, header_alignment, thin_border, group_yaml_data, collected_group_codes):
    """リマインダー通知設定のシートを作成"""
    ws = wb.create_sheet(title="リマインダー通知設定")
    
    # A, B, C列の幅を330pxに設定（約47文字分）
    ws.column_dimensions["A"].width = 47
    ws.column_dimensions["B"].width = 47
    ws.column_dimensions["C"].width = 47
    
    # ヘッダー行 - フィールド名列を追加
    headers = ["No.", "リマインダー名", "通知先タイプ", "フィールド名", "通知先", "日時フィールド", "条件", "通知タイミング"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # データ行
    row_idx = 2
    reminders = data.get("reminders", [])
    
    # データ行の背景色を設定
    light_blue_fill = PatternFill(start_color="EBF1F5", end_color="EBF1F5", fill_type="solid")
    light_yellow_fill = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
    
    # グループの通知先を収集
    group_codes = []
    
    for remind_idx, remind in enumerate(reminders, 1):
        name = remind.get("title", "")
        date_field = remind.get("timing", {}).get("field", {}).get("code", "")
        
        # 通知タイミング
        timing_type = remind.get("timing", {}).get("type", "")
        timing_value = remind.get("timing", {}).get("value", "")
        timing_unit = remind.get("timing", {}).get("unit", "")
        
        # 通知タイミングを整形
        timing_jp = ""
        if timing_type == "BEFORE":
            timing_jp = f"{timing_value}{timing_unit}前"
        elif timing_type == "AFTER":
            timing_jp = f"{timing_value}{timing_unit}後"
        else:
            timing_jp = f"{timing_type}: {timing_value} {timing_unit}"
        
        # 条件を整形
        condition_type = remind.get("filterCond", "")
        condition_jp = "全レコード" if not condition_type else f"条件式: {condition_type}"
        
        # リマインダーごとに背景色を交互に変更
        remind_fill = light_blue_fill if remind_idx % 2 == 1 else light_yellow_fill
        
        # 通知先
        recipients = remind.get("recipients", [])
        
        if not recipients:
            # 通知先がない場合は1行だけ出力
            row_data = [
                remind_idx,  # No.
                name,  # リマインダー名
                "",  # 通知先タイプ
                "",  # フィールドタイプ - 新しい列
                "通知先なし",  # 通知先
                date_field,  # 日時フィールド
                condition_jp,  # 条件
                timing_jp,  # 通知タイミング
            ]
            
            for col_idx, value in enumerate(row_data, 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.value = value
                cell.border = thin_border
                cell.fill = remind_fill
            
            row_idx += 1
        else:
            # 通知先ごとに行を作成
            for recip_idx, recipient in enumerate(recipients):
                entity = recipient.get("entity", {})
                entity_type = entity.get("type", "")
                entity_code = entity.get("code", "")
                
                # グループコードを収集
                if entity_type == "GROUP":
                    group_codes.append(entity_code)
                
                # 通知先タイプを日本語に変換
                type_jp = ""
                field_type = ""
                if entity_type == "USER":
                    type_jp = "ユーザー"
                elif entity_type == "GROUP":
                    type_jp = "グループ"
                elif entity_type == "ORGANIZATION":
                    type_jp = "組織"
                elif entity_type == "FIELD_ENTITY":
                    type_jp = "フィールド"
                    # フィールドタイプの取得
                    if "type" in entity:
                        if entity.get("type") == "CREATOR":
                            field_type = "作成者"
                        elif entity.get("type") == "MODIFIER":
                            field_type = "更新者"
                        elif entity.get("type") == "USER_SELECT":
                            field_type = "ユーザー選択"
                        elif entity.get("type") == "GROUP_SELECT":
                            field_type = "グループ選択"
                        elif entity.get("type") == "ORGANIZATION_SELECT":
                            field_type = "組織選択"
                        else:
                            field_type = entity.get("type", "")
                else:
                    type_jp = entity_type
                
                row_data = [
                    remind_idx if recip_idx == 0 else "",  # No.（最初の通知先の行のみ表示）
                    name if recip_idx == 0 else "",  # リマインダー名（最初の通知先の行のみ表示）
                    type_jp,  # 通知先タイプ
                    field_type,  # フィールドタイプ - 新しい列
                    entity_code,  # 通知先
                    date_field if recip_idx == 0 else "",  # 日時フィールド（最初の通知先の行のみ表示）
                    condition_jp if recip_idx == 0 else "",  # 条件（最初の通知先の行のみ表示）
                    timing_jp if recip_idx == 0 else "",  # 通知タイミング（最初の通知先の行のみ表示）
                ]
                
                for col_idx, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.value = value
                    cell.border = thin_border
                    cell.fill = remind_fill
                
                row_idx += 1
    
    # 列幅の調整
    for col_idx in range(4, len(headers) + 1):  # A, B, C列は既に設定済みなので4列目から設定
        column_letter = get_column_letter(col_idx)
        if col_idx == 4:  # フィールドタイプ列
            ws.column_dimensions[column_letter].width = 15
        elif col_idx == 5:  # 通知先列
            ws.column_dimensions[column_letter].width = 20
        elif col_idx == 6:  # 日時フィールド列
            ws.column_dimensions[column_letter].width = 15
        elif col_idx == 7:  # 条件列
            ws.column_dimensions[column_letter].width = 30
        else:  # 通知タイミング列
            ws.column_dimensions[column_letter].width = 15

    # グループメンバー情報を追加
    if group_codes:
        row_idx = add_group_members_table(ws, row_idx, group_codes, header_font, header_fill, header_alignment, thin_border, group_yaml_data, collected_group_codes)

def add_group_members_table(ws, row_idx, group_codes, header_font, header_fill, header_alignment, thin_border, group_yaml_data, collected_group_codes):
    """グループメンバー情報の表を追加"""

    if not group_codes or not group_yaml_data:
        return row_idx
    
    # グループ情報の見出し
    row_idx += 2
    group_header_fill = PatternFill(start_color="CCFFCC", end_color="CCFFCC", fill_type="solid")  # グループ用の背景色（薄い緑）
    
    ws.cell(row=row_idx, column=1).value = "通知先種別：グループ"
    ws.cell(row=row_idx, column=1).font = Font(bold=True, size=12)
    ws.cell(row=row_idx, column=1).fill = group_header_fill
    row_idx += 1
    
    # 重複するグループコードを除去
    unique_group_codes = list(set(group_codes))
    
    for group_code in unique_group_codes:
        # グループが存在しない場合はスキップ
        if group_code not in group_yaml_data:
            logging.warning(f"グループ {group_code} の情報が見つかりません")
            continue
        
        group_info = group_yaml_data[group_code]
        group_name = group_info.get('name', '不明なグループ')
        members = group_info.get('users', [])
        
        # メンバーをソート（アクティブユーザーを上部に、同一ドメインでグループ化）
        members = sort_group_members(members)

        # ヘッダー行
        headers = ["グループ名", "アカウント名", "メールアドレス", "停止中"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.value = header
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = header_alignment
            cell.border = thin_border
        row_idx += 1
        
        # メンバー行
        first_row_of_group = row_idx  # グループの最初の行を記録
        
        for i, user in enumerate(members):
            # A列: グループ情報（最初のメンバーの行のみ）
            cell_a = ws.cell(row=row_idx, column=1)
            if i == 0:
                cell_a.value = f"{group_name} ({group_code})"
            cell_a.border = thin_border
            
            # B列: アカウント名
            cell_b = ws.cell(row=row_idx, column=2)
            cell_b.value = user.get('username', '')
            cell_b.border = thin_border
            
            # C列: メールアドレス
            cell_c = ws.cell(row=row_idx, column=3)
            cell_c.value = user.get('email', '')
            cell_c.border = thin_border
            
            # D列: 停止中かどうか
            cell_d = ws.cell(row=row_idx, column=4)
            cell_d.value = "停止中" if user.get('isDisabled', False) else ""
            cell_d.border = thin_border

            # D列が「停止中」の場合、B, C, D列の背景色を淡いグレーに設定
            if cell_d.value == "停止中":
                gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
                for col in range(2, 5):  # B, C, D列
                    ws.cell(row=row_idx, column=col).fill = gray_fill
            
            row_idx += 1
        
        # メンバーが複数いる場合、A列を結合
        last_row_of_group = row_idx - 1
        
        if last_row_of_group > first_row_of_group:
            merge_cells_in_column_a(ws, first_row_of_group, last_row_of_group)
        
        # グループ間の空白
        row_idx += 1
    
    return row_idx

def merge_cells_in_column_a(ws, start_row, end_row):
    """A列で縦にセルを結合する"""
    if end_row > start_row:
        ws.merge_cells(
            start_row=start_row,
            start_column=1,
            end_row=end_row,
            end_column=1
        )
        
        # 結合したセルの配置を中央揃えに
        merged_cell = ws.cell(row=start_row, column=1)
        merged_cell.alignment = Alignment(vertical='center')

def sort_group_members(members):
    """
    グループメンバーをソートする関数

    1. 優先ドメイン（環境変数 USER_DOMAIN で指定）を含むアクティブなユーザーを上部に配置
    2. その他のアクティブなユーザーを中部に配置
    3. 停止中（isDisabled=True）のユーザーを下部に配置
    4. 各グループ内はユーザー名でソート

    Args:
        members: ユーザー情報のリスト

    Returns:
        list: ソートされたユーザーリスト
    """
    # .env ファイルからドメイン情報を読み込む
    try:
        # SCRIPT_DIRを使用して絶対パスを取得
        env_path = Path(SCRIPT_DIR).parent / '.kintone.env'
        
        # ファイルの存在を確認してからロード
        if env_path.exists():
            logging.info(f".kintone.env ファイルを読み込みます: {env_path}")
            load_dotenv(env_path)
            
            # 環境変数から優先ドメインを取得
            priority_domain = os.getenv('USER_DOMAIN', '').lower()
            logging.info(f"取得した優先ドメイン: {priority_domain}")
        else:
            logging.warning(f".kintone.env ファイルが見つかりません: {env_path}")
            priority_domain = ''
            
    except Exception as e:
        logging.error(f"環境変数の読み込みでエラーが発生しました: {e}")
        priority_domain = ''
    
    # 優先ドメインが空の場合、デフォルト値を設定
    if not priority_domain:
        priority_domain = 'kirin.co.jp'  # 既定値を設定
        logging.info(f"優先ドメインが設定されていないため、デフォルト値'{priority_domain}'を使用します")
    else:
        logging.info(f"優先ドメイン'{priority_domain}'を使用してユーザーをソートします")

    def get_group(user):
        """ユーザーをグループ分けする
           0: アクティブかつ優先ドメイン一致
           1: アクティブで優先ドメイン以外（ドメイン未設定も含む）
           2: 停止中
        """
        is_disabled = user.get('isDisabled', False)
        email = user.get('email', '')
        domain = ''
        if email and '@' in email:
            domain = email.split('@')[1].lower()

        if is_disabled:
            return 2
        elif domain == priority_domain:
            return 0
        else:
            return 1

    def sort_key(user):
        # 各グループ内はユーザー名でソート
        return (get_group(user), user.get('username', ''))
    
    return sorted(members, key=sort_key)

def add_user_information_table(ws, row_idx, user_codes, header_font, header_fill, header_alignment, thin_border, user_yaml_data):
    """ユーザー情報の表を追加"""

    if not user_codes or not user_yaml_data:
        return row_idx
    
    # ユーザー情報の見出し
    row_idx += 2
    user_header_fill = PatternFill(start_color="FFCCCC", end_color="FFCCCC", fill_type="solid")  # ユーザー用の背景色（薄い赤）
    
    ws.cell(row=row_idx, column=1).value = "通知先種別：ユーザー 情報"
    ws.cell(row=row_idx, column=1).font = Font(bold=True, size=12)
    ws.cell(row=row_idx, column=1).fill = user_header_fill
    row_idx += 1
    
    # 重複するユーザーコードを除去
    unique_user_codes = list(set(user_codes))
    
    # ヘッダー行
    headers = ["アカウント名", "メールアドレス", "停止中"]
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=row_idx, column=col_idx)
        cell.value = header
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    row_idx += 1
    
    for user_code in unique_user_codes:
        # ユーザーが存在しない場合はコードのみ表示
        if user_code not in user_yaml_data:
            # A列: アカウント名（コードのみ）
            cell_a = ws.cell(row=row_idx, column=1)
            cell_a.value = user_code
            cell_a.border = thin_border
            
            # B列: メールアドレス（空欄）
            cell_b = ws.cell(row=row_idx, column=2)
            cell_b.border = thin_border
            
            # C列: 停止中（空欄）
            cell_c = ws.cell(row=row_idx, column=3)
            cell_c.border = thin_border
            
            row_idx += 1
            continue
        
        user_info = user_yaml_data[user_code]
        
        # A列: アカウント名
        cell_a = ws.cell(row=row_idx, column=1)
        cell_a.value = user_info.get('username', user_code)
        cell_a.border = thin_border
        
        # B列: メールアドレス
        cell_b = ws.cell(row=row_idx, column=2)
        cell_b.value = user_info.get('email', '')
        cell_b.border = thin_border
        
        # C列: 停止中かどうか
        cell_c = ws.cell(row=row_idx, column=3)
        cell_c.value = "停止中" if user_info.get('isDisabled', False) else ""
        cell_c.border = thin_border

        # C列が「停止中」の場合、行全体の背景色を淡いグレーに設定
        if cell_c.value == "停止中":
            gray_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
            for col in range(1, 4):  # A, B, C列
                ws.cell(row=row_idx, column=col).fill = gray_fill
        
        row_idx += 1
    
    return row_idx

def parse_args():
    import sys
    print("=== デバッグ情報: notifications_to_excel.py ===")
    print("生の引数情報:", sys.argv)
    
    # 環境変数から認証情報を取得
    subdomain = os.environ.get('KINTONE_SUBDOMAIN')
    username = os.environ.get('KINTONE_USERNAME')
    password = os.environ.get('KINTONE_PASSWORD')
    
    parser = argparse.ArgumentParser(description='kintoneの通知設定をExcelに出力する')
    parser.add_argument('app_id', type=int, help='アプリID（必須）')
    parser.add_argument('--subdomain', default=subdomain, help='kintoneのサブドメイン（必須）')
    parser.add_argument('--username', default=username, help='kintoneのユーザー名（必須）')
    parser.add_argument('--password', default=password, help='kintoneのパスワード（必須）')
    parser.add_argument('--api-token', help='kintoneのAPIトークン（オプション）')
    parser.add_argument('--output', required=True, help='出力するExcelファイルのパス（必須）')
    parser.add_argument('--group-master', help='グループマスターファイルのパス（オプション）')
    parser.add_argument('--log-level', choices=['DEBUG', 'INFO', 'WARNING', 'ERROR', 'CRITICAL'], default='INFO', help='ログレベル（デフォルト: INFO）')
    
    args = parser.parse_args()
    print("パース後の引数情報:", vars(args))
    print("================================")
    return args

def main():
    args = parse_args()
    # メイン処理の実装

if __name__ == "__main__":
    main() 