import os
import sys
import yaml
import re
import argparse
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from openpyxl.utils.cell import get_column_letter
import csv
import logging
from collections import Counter



def setup_logging(log_level, silent=False):
    """
    ロギングの設定を行う

    Args:
        log_level (str): ログレベル ('DEBUG', 'INFO', 'WARNING', 'ERROR', 'CRITICAL')
        silent (bool): サイレントモード（ログ出力を抑制）
    """
    if silent:
        logging.getLogger().setLevel(logging.CRITICAL)
        return

    log_format = '%(asctime)s - %(levelname)s - %(message)s'
    logging.basicConfig(
        level=getattr(logging, log_level.upper()),
        format=log_format,
        datefmt='%Y-%m-%d %H:%M:%S'
    )

def get_all_entities(data):
  """
  全てのユーザ/グループを抽出（グループコードを返す）

  Args:
    data (dict): YAMLファイルから読み込んだデータ

  Returns:
    list: グループコードのリスト（ソート済み）
  """
  entities = set()
  for rights_block in data.get('rights', []):
    for entity in rights_block.get('entities', []):
      code = entity['entity']['code']
      if code != 'everyone':  # everyoneを除外
        entities.add(code)
  return sorted(list(entities))

def load_master_status(header_name, base_dir):
  """
  プロセス管理マスタからステータス一覧を読み込む

  Args:
    header_name (str): ヘッダー名（例: '52'）
    base_dir (str): ディレクトリのパス

  Returns:
    set: ステータス名のセット
  """
  master_file = os.path.join(base_dir, f"{header_name}_process_management.yaml")
  try:
    with open(master_file, 'r', encoding='utf-8') as f:
      master_data = yaml.safe_load(f)
      if not master_data.get('enable', True):
        logging.debug("プロセス管理が無効です")
        return set()
      status_list = set(master_data.get('states', {}).keys())
      logging.debug(f"ステータス一覧を読み込みました: {status_list}")
      return status_list
  except Exception as e:
    logging.warning(f"マスタファイル {master_file} の読み込みに失敗しました: {str(e)}")
    return set()

def load_group_list(group_master_path):
  """
  グループマスタからグループコードとグループ名のマッピングを読み込む

  Args:
    group_master_path (str): グループマスタファイルのパス

  Returns:
    dict: グループコードをキー、グループ名を値とする辞書
  """
  try:
    with open(group_master_path, 'r', encoding='utf-8') as f:
      group_data = yaml.safe_load(f)
      # グループデータから name フィールドのみを抽出
      return {code: info['name'] for code, info in group_data.items()}
  except Exception as e:
    logging.warning(f"グループマスタファイル {group_master_path} の読み込みに失敗しました: {str(e)}")
    return {}

def load_form_fields(header_name, base_dir):
  """
  フォームフィールドファイルからFIELD_ENTITYのフィールド名とラベルを読み込む

  Args:
    header_name (str): ヘッダー名（例: '52'）
    base_dir (str): ディレクトリのパス

  Returns:
    dict: フィールドコードをキー、ラベルを値とする辞書
  """
  form_fields_file = os.path.join(base_dir, f"{header_name}_form_fields.yaml")
  try:
    with open(form_fields_file, 'r', encoding='utf-8') as f:
      form_fields_data = yaml.safe_load(f)
      field_entities = {}
      for field_code, field_info in form_fields_data.get('properties', {}).items():
        label = field_info.get('label', field_code)
        # フィールドエンティティの場合は「(フィールド)」を付加
        label += '(フィールド)'
        field_entities[field_code] = label
      return field_entities
  except Exception as e:
    logging.warning(f"フォームフィールドファイル {form_fields_file} の読み込みに失敗しました: {str(e)}")
    return {}

def load_entity_type_map(header_name, base_dir):
  """
  [ヘッダ名]_record_acl.yaml からエンティティコードとタイプをマッピングする

  Args:
    header_name (str): ヘッダー名（例: '52'）
    base_dir (str): ディレクトリのパス

  Returns:
    dict: エンティティコードをキー、タイプ（GROUP、FIELD_ENTITY、USER）を値とする辞書
  """
  record_acl_file = os.path.join(base_dir, f"{header_name}_record_acl.yaml")
  try:
    with open(record_acl_file, 'r', encoding='utf-8') as f:
      record_acl_data = yaml.safe_load(f)
      entity_type_map = {}
      for rights_block in record_acl_data.get('rights', []):
        for entity in rights_block.get('entities', []):
          code = entity['entity']['code']
          entity_type = entity['entity'].get('type', 'GROUP') # デフォルトは GROUP
          entity_type_map[code] = entity_type
      return entity_type_map
  except Exception as e:
    logging.error(f"エラー: record_aclファイル {record_acl_file} の読み込みに失敗しました: {str(e)}")
    sys.exit(1)

def load_group_map(header_name, base_dir, group_master_path, field_entities):
  """
  グループマスタとフォームフィールドからグループマッピングを読み込む

  Args:
    header_name (str): ヘッダー名（例: '52'）
    base_dir (str): ディレクトリのパス
    group_master_path (str): グループマスタファイルのパス
    field_entities (dict): FIELD_ENTITYのフィールドコードとラベルのマッピング

  Returns:
    dict: エンティティコードをキー、グループ名（文字列）またはフィールドラベルを値とする辞書
  """
  group_data = load_group_list(group_master_path)
  # グループコードをキー、グループ名を値としてマッピング
  group_map = group_data
  # FIELD_ENTITYをグループマッピングに追加
  group_map.update(field_entities)
  return group_map

def load_userValid_list(user_list_path):
  """
  user_list.yaml からユーザーコードと有効性を読み込む

  Args:
    user_list_path (str): user_list.yaml のファイルパス

  Returns:
    dict: ユーザーコードをキー、validを値とする辞書
  """
  try:
    with open(user_list_path, 'r', encoding='utf-8') as f:
      user_data = yaml.safe_load(f)
      user_map = {}
      for key, user_info in user_data.items():
        code = user_info.get('code')
        valid = user_info.get('valid', False)
        if code:
          user_map[code] = valid
      return user_map
  except Exception as e:
    logging.warning(f"警告: user_list.yaml の読み込みに失敗しました: {str(e)}")
    return {}

def load_userName_list(group_master_path):
  """
  group_user_list.yaml からユーザー名の一覧を読み込む

  Args:
    group_master_path (str): group_user_list.yaml のファイルパス

  Returns:
    dict: ユーザーIDをキー、ユーザー名を値とする辞書
  """
  try:
    with open(group_master_path, 'r', encoding='utf-8') as f:
      group_data = yaml.safe_load(f)
      user_map = {}
      
      # 全グループをループして、ユニークなユーザー情報を収集
      for group_info in group_data.values():
        for user in group_info.get('users', []):
          username = user.get('username')
          if username:
            user_map[username] = username  # usernameをそのまま表示名として使用
      
      return user_map
  except Exception as e:
    logging.warning(f"警告: group_user_list.yaml の読み込みに失敗しました: {str(e)}")
    return {}

def extract_field_conditions(data):
  """
  全ての条件フィールドとその値を抽出

  Args:
    data (dict): YAMLファイルから読み込んだデータ

  Returns:
    dict: フィールド名をキー、値のリストを値とする辞書
  """
  field_values = {}
  # 条件のパターンを定義（例: "ステータス in (値1, 値2)"）
  pattern = r'(\w+)\s+in\s+\((.*?)\)'

  for rights_block in data.get('rights', []):
    condition = rights_block.get('filterCond', '')
    matches = re.finditer(pattern, condition)

    for match in matches:
      field = match.group(1)
      values = match.group(2)

      # 値を個別に分割して整理
      cleaned_values = set()
      for value in values.split(','):
        clean_value = value.strip().strip('"').strip("'").strip()
        if clean_value:
          cleaned_values.add(clean_value)

      if field not in field_values:
        field_values[field] = set()
      field_values[field].update(cleaned_values)

  # 各フィールドの値をソートしてリスト化
  return {k: sorted(list(v)) for k, v in field_values.items()}

def parse_rights_block(rights_block, group_map, entity_type_map, user_map):
  """
  権限ブロックを解析してデータを抽出する（グループ名に置換）

  Args:
    rights_block (dict): 権限ブロックのデータ
    group_map (dict): エンティティコードからグループ名へのマッピング
    entity_type_map (dict): エンティティコードからタイプへのマッピング
    user_map (dict): ユーザーコードから有効性へのマッピング

  Returns:
    dict: 各種グループの権限と条件を含む辞書
  """
  entities = rights_block.get('entities', [])

  entity_permissions = []  # リストに変更
  viewable_groups = []
  editable_groups = []
  deletable_groups = []
  invalid_groups = set()

  for entity in entities:
    group_code = entity['entity']['code']
    entity_type = entity['entity'].get('type', 'GROUP')
    invalid_group = False
    invalid_user = False

    if entity_type == 'GROUP':
      group_name = group_map.get(group_code, '')
      invalid_group = (group_name == '')
      if invalid_group:
        logging.warning(f"無効なGROUP: {group_code}")
        group_name = group_code
    elif entity_type == 'FIELD_ENTITY':
      group_name = group_map.get(group_code, '')
      invalid_group = (group_name == '')
      if invalid_group:
        group_name = group_code
    elif entity_type == 'USER':
      group_name = group_code
      # ユーザーの有効性チェックは別の方法で行う必要がある場合は
      # ここで実装する
    else:
      group_name = group_code
      invalid_group = True
      logging.warning(f"未知のタイプ '{entity_type}' を持つエンティティ: {group_code}")

    permissions = []

    if entity.get('viewable', False):
      viewable_groups.append(group_name)
      permissions.append('閲覧')
    if entity.get('editable', False):
      editable_groups.append(group_name)
      permissions.append('編集')
    if entity.get('deletable', False):
      deletable_groups.append(group_name)
      permissions.append('削除')
    permissions.append('') # 権限なし　

    # 権限情報を辞書として追加
    entity_info = {
      'name': group_name,
      'permissions': permissions,
      'invalid': invalid_group or invalid_user
    }
    entity_permissions.append(entity_info)  # リストに追加

    if invalid_group or invalid_user:
      invalid_groups.add(group_name)

  return {
    'viewable_groups': viewable_groups if viewable_groups else ['-'],
    'editable_groups': editable_groups if editable_groups else ['-'],
    'deletable_groups': deletable_groups if deletable_groups else ['-'],
    'conditions': rights_block.get('filterCond', ''),
    'entity_permissions': entity_permissions,  # リストを返す
    'invalid_groups': invalid_groups
  }

def check_condition_match(condition, field, value):
  """
  条件に特定のフィールドと値が含まれているかチェック

  Args:
    condition (str): フィルター条件の文字列
    field (str): フィールド名
    value (str): 値

  Returns:
    bool: 含まれていればTrue、そうでなければFalse
  """
  pattern = f'{field}\\s+in\\s+\\((.*?)\\)'

  if f'{field} in' in condition:
    match = re.search(pattern, condition)
    if match:
      values = match.group(1).split(',')
      return any(value == v.strip().strip('"').strip("'").strip() for v in values)
  return False

def create_header_cell(ws, row, col, value, rotation=False, merge_cells=None,
        invalid_status=False, invalid_group=False, invalid_user=False, tate_center=False, background_color='D9D9D9'):
  """
  ヘッダーセルを作成する（無効なステータス、グループ、またはユーザーのスタイリングを追加）

  Args:
    ws (Worksheet): Excelのワークシート
    row (int): 行番号
    col (int): 列番号
    value (str): セルに設定する値
    rotation (bool, optional): 縦書きにするかどうか
    merge_cells (tuple, optional): セルのマージ範囲（start_row, end_row, start_col, end_col）
    invalid_status (bool, optional): 無効なステータスかどうか
    invalid_group (bool, optional): グループマスタに存在しないグループかどうか
    invalid_user (bool, optional): ユーザーマスタに存在しない、または無効なユーザーかどうか
    tate_center (bool, optional): 縦方向の中央揃え

  Returns:
    Cell: 作成したセル
  """
  cell = ws.cell(row=row, column=col, value=value)

  # フォント設定
  if invalid_user:
    # 無効なユーザーは赤色の太字
    cell.font = Font(bold=True, color='FF0000')
  elif invalid_group:
    # グループマスタに存在しないグループは赤色の太字
    cell.font = Font(bold=True, color='FF0000')
  elif invalid_status:
    # 無効なステータスは白色の太字
    cell.font = Font(bold=True, color='FFFFFF')
  else:
    # 通常のフォント
    cell.font = Font(bold=True, color='000000')

  # 塗りつぶし設定
  if invalid_status:
    # 無効なステータスのヘッダ背景は赤色
    cell.fill = PatternFill(start_color='FF0000', end_color='FF0000', fill_type='solid')
  else:
    # 通常はグレー
    cell.fill = PatternFill(start_color=background_color, end_color=background_color, fill_type='solid')

  # アライメント設定
  vertical_value = 'center' if tate_center else 'top'
  align_params = {'wrap_text': True, 'vertical': vertical_value}
  if rotation:
    align_params['textRotation'] = 90 # 縦書き
    align_params['horizontal'] = 'center' # 縦書き時の中央揃え
  else:
    align_params['horizontal'] = 'center' # その他のセルは中央揃え
  cell.alignment = Alignment(**align_params)

  # セルのマージ
  if merge_cells:
    start_row, end_row, start_col, end_col = merge_cells
    ws.merge_cells(start_row=start_row, end_row=end_row,
           start_column=start_col, end_column=end_col)

  # ボーダー設定（全方向に細線）
  thin_border = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
  )
  cell.border = thin_border

  return cell

def convert_yaml_to_excel(wb, header_name, base_dir, group_map, entity_type_map, user_map, acl_type, sheet_name, userName_map, group_members, permission_target_user_names):
  """
  YAMLファイルをエクセルファイルの指定シートに変換する
  
  Args:
    wb (Workbook): openpyxl Workbookオブジェクト
    header_name (str): ヘッダー名（例: '52'）
    base_dir (str): ディレクトリのパス
    group_map (dict): エンティティコードからグループ名へのマッピング
    entity_type_map (dict): エンティティコードからタイプへのマッピング
    user_map (dict): ユーザーコードから有効性へのマッピング
    acl_type (str): ACLの種類 ('record' または 'app')
    sheet_name (str): Excelシート名 ('レコード' または 'アプリ')
    userName_map (dict): ユーザーコードからユーザー名へのマッピング
    group_members (dict): グループコードからメンバー情報へのマッピング
  """
  input_file = os.path.join(base_dir, f"{header_name}_{acl_type}_acl.yaml")

  # マスタステータスの読み込み（必要に応じて）
  master_statuses = load_master_status(header_name, base_dir)

  # YAMLファイルの読み込み
  try:
    with open(input_file, 'r', encoding='utf-8') as f:
      data = yaml.safe_load(f)
  except Exception as e:
    logging.error(f"エラー: {acl_type}_aclファイル {input_file} の読み込みに失敗しました: {str(e)}")
    return

  # シートの取得または作成
  if sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
  else:
    ws = wb.create_sheet(title=sheet_name)

  # ウインドウ枠を固定
  ws.freeze_panes = 'E4'

  # D列の幅を32pxに設定
  d_width_in_pixels = 32
  excel_width = (d_width_in_pixels - 5) / 7 + 1
  ws.column_dimensions['D'].width = excel_width

  # 1行目を空ける
  current_row = 2

  if acl_type == 'record':
    # 基本ヘッダーの設定（A-E列）
    headers = ['No', '条件', 'グループ', '権限', '備考']
    for col_num, header in enumerate(headers, 1):
      create_header_cell(ws, current_row, col_num, header,
                        merge_cells=None, tate_center=True, background_color='00CCCC')

    # 全エンティティの収集と分類
    all_entities = set()
    entity_types = {'USER': [], 'GROUP': [], 'FIELD_ENTITY': []}
    
    for rights_block in data.get('rights', []):
      for entity in rights_block.get('entities', []):
        code = entity['entity']['code']
        entity_type = entity['entity'].get('type', 'GROUP')
        all_entities.add(code)
        if code not in entity_types[entity_type]:
          entity_types[entity_type].append(code)

    # エンティティ列の作成（F列以降）
    current_col = 6
    
    # ユーザー/グループ列のヘッダー作成
    for entity_type, codes in entity_types.items():
      if codes:
        type_label = {
          'USER': 'ユーザ',
          'GROUP': 'グループ',
          'FIELD_ENTITY': 'フィールド'
        }.get(entity_type, entity_type)
        
        # タイプヘッダー
        create_header_cell(ws, current_row, current_col, type_label,
                         merge_cells=(current_row, current_row, current_col, current_col + len(codes) - 1))
        
        # 個別エンティティ名
        for code in codes:
          display_name = group_map.get(code, code)  # 'name' フィールドを取得せず、直接グループ名を取得
          create_header_cell(ws, current_row + 1, current_col, display_name, rotation=True)
          current_col += 1

    # 重複を除いた単純なユーザ名一覧の取得
    unique_user_names = set()
    for rights_block in data.get('rights', []):
      for entity in rights_block.get('entities', []):
        code = entity['entity']['code']
        entity_type = entity['entity'].get('type', 'GROUP')
        if entity_type == 'GROUP':
          group_info = group_members.get(code, {})
          users = group_info.get('users', [])
          for user in users:
            user_name = user.get('username', '不明')
            unique_user_names.add(user_name)
        elif entity_type == 'USER':
          user_name = userName_map.get(code, code)
          unique_user_names.add(user_name)
    
    # 全対象でのユーザー名列のヘッダー作成
    if permission_target_user_names:
      # ユーザー名ヘッダー
      create_header_cell(ws, current_row, current_col, '個別ユーザー権限',
                       merge_cells=(current_row, current_row, current_col, current_col + len(permission_target_user_names) - 1), background_color='CC7777')
      
      # 個別ユーザー名
      for user_name in sorted(permission_target_user_names):
        create_header_cell(ws, current_row + 1, current_col, user_name, rotation=True, background_color='CC7777')
        current_col += 1

    # データ行の書き込み
    data_row = current_row + 2
    for i, rights_block in enumerate(data.get('rights', []), 1):
      parsed_data = parse_rights_block(rights_block, group_map, entity_type_map, user_map)
      
      # 基本情報の書き込み（A-E列）
      start_row = data_row
      
      # 権限を持つエンティティの数を取得
      entities_with_rights = [
        entity_info for entity_info in parsed_data['entity_permissions']
        if entity_info['permissions']
      ]
      num_entities = len(entities_with_rights)
      
      # 太い罫線の設定
      thick_border = Border(
        left=Side(style='thick'),
        right=Side(style='thick'),
        top=Side(style='thick'),
        bottom=Side(style='thick')
      )
      
      # 最右列の列番号を取得
      max_col = ws.max_column
      
      # 枠の罫線を引く（4行目以降）
      if start_row >= 4:
        # 枠の上端
        for col in range(1, max_col + 1):
          cell = ws.cell(row=start_row, column=col)
          current_border = cell.border
          cell.border = Border(
            left=current_border.left,
            right=current_border.right,
            top=Side(style='thick'),
            bottom=current_border.bottom
          )
        
        # 枠の左端と右端
        for row in range(start_row, start_row + num_entities):
          # 左端
          cell = ws.cell(row=row, column=1)
          current_border = cell.border
          cell.border = Border(
            left=Side(style='thick'),
            right=current_border.right,
            top=current_border.top,
            bottom=current_border.bottom
          )
          # 右端
          cell = ws.cell(row=row, column=max_col)
          current_border = cell.border
          cell.border = Border(
            left=current_border.left,
            right=Side(style='thick'),
            top=current_border.top,
            bottom=current_border.bottom
          )
        
        # 枠の下端
        for col in range(1, max_col + 1):
          cell = ws.cell(row=start_row + num_entities - 1, column=col)
          current_border = cell.border
          cell.border = Border(
            left=current_border.left,
            right=current_border.right,
            top=current_border.top,
            bottom=Side(style='thick')
          )

      # 以降の既存のコード（No列とCondition列のマージなど）
      if num_entities > 1:
        ws.merge_cells(start_row=start_row, start_column=1, end_row=start_row + num_entities - 1, end_column=1)
        ws.merge_cells(start_row=start_row, start_column=2, end_row=start_row + num_entities - 1, end_column=2)
      
      # No列とCondition列の値を設定
      ws.cell(row=start_row, column=1, value=i)
      ws.cell(row=start_row, column=2, value=parsed_data['conditions'])
      
      # C列に各エンティティを別々のセルで表示し、F列以降の権限も対応する行に表示
      current_entity_row = start_row
      for entity_info in entities_with_rights:
        if entity_info['permissions']:  # 権限がある場合のみ追加
          # C列にエンティティ名のみを表示
          cell = ws.cell(row=current_entity_row, column=3, value=entity_info['name'])
          
          # D列に権限を表示（括弧なし）
          permissions_str = '/'.join(entity_info['permissions'])
          ws.cell(row=current_entity_row, column=4, value=permissions_str)
          
          # E列に計算式を設定（4行目以降）- D列を参照するように変更
          if current_entity_row >= 4:
            formula = f'=IF(A{current_entity_row}=0,IF(LEN(D{current_entity_row-1})<LEN(D{current_entity_row}),"権限順不整合",""),"")'
            ws.cell(row=current_entity_row, column=5, value=formula)
          
          # F列以降の権限マッピング
          col = 6
          for entity_type, codes in entity_types.items():
            for code in codes:
                # current_entity_row が data_row よりも大きい場合のみループを実行
                if current_entity_row > 3:
                    current_row = current_entity_row
                    entity_name = ws.cell(row=current_row, column=3).value  # C列の値を取得
                    permissions = ws.cell(row=current_row, column=4).value  # D列の権限を取得
                    
                    display_name = group_map.get(code, code)  # エンティティの表示名を取得
                    
                    # C列の値とエンティティの表示名が一致する場合、権限を転記
                    if entity_name == display_name:
                        ws.cell(row=current_row, column=col, value=permissions)
                col += 1
          
          # 無効なエンティティの場合のみ赤字で表示
          if entity_info['invalid']:
            cell.font = Font(bold=True, color='FF0000')
          
          current_entity_row += 1
      
      # 次の権限ブロックの開始行を設定
      data_row = current_entity_row

    # ユーザー権限列の処理
    logging.debug(f"unique_user_names: {unique_user_names}")  
    logging.debug(f"permission_target_user_names: {permission_target_user_names}")    
    # 全出現ユーザーでのユーザー名列のヘッダーは上の方で既に作成されています。
    # なので、順序をsortedで合わせる必要があります。
    if permission_target_user_names:
        for user_name in sorted(permission_target_user_names):
            # 各ユーザーの権限ブロックごとの権限を記録する辞書
            previous_permissions = {}
            current_block_start = 4  # 最初の権限ブロックの開始行

            for current_row in range(4, data_row):
                entity_name = ws.cell(row=current_row, column=3).value  # C列の値を取得
                permissions = ws.cell(row=current_row, column=4).value  # D列の権限を取得
                
                # A列の値が変わったら新しい権限ブロックの開始
                block_number = ws.cell(row=current_row, column=1).value
                if block_number and block_number != ws.cell(row=current_row-1, column=1).value:
                    current_block_start = current_row
                    previous_permissions[block_number] = None

                # ユーザーの権限を設定
                user_permission = None
                # ユーザー名が直接マッチする場合
                if entity_name == user_name:
                    user_permission = permissions
                else:
                    # グループに所属しているか確認
                    for group_code, group_info in group_members.items():
                        if group_info.get('name') == entity_name:  # グループ名がマッチ
                            users = group_info.get('users', [])
                            for user in users:
                                if user.get('username') == user_name:  # ユーザーが所属している
                                    user_permission = permissions
                                    break
                            if user_permission:
                                break

                if user_permission:
                    cell = ws.cell(row=current_row, column=col, value=user_permission)
                    
                    # 同じ権限ブロック内で以前の権限と比較
                    block_number = ws.cell(row=current_block_start, column=1).value
                    if previous_permissions.get(block_number) is not None:
                        # 同じブロック内で2回目以降の権限は赤色で表示
                        cell.font = Font(color='FF0000')
                    
                    previous_permissions[block_number] = user_permission

            col += 1

    # 列幅の設定
    ws.column_dimensions[get_column_letter(1)].width = 5  # No
    ws.column_dimensions[get_column_letter(2)].width = 40  # 条件       240pxにしたい
    ws.column_dimensions[get_column_letter(3)].width = 40  # グループ    250pxにしたい
    ws.column_dimensions[get_column_letter(4)].width = 20  # 権限
    ws.column_dimensions[get_column_letter(5)].width = 20  # 備考
    
    # F列以降の列幅設定
    for col in range(6, current_col):
      ws.column_dimensions[get_column_letter(col)].width = 15

    ws.cell(row=1, column=2).value = "凡例：　閲覧：V　　編集：E　　削除：D"

    # B列4行目以降を「折り返して全体表示」に設定
    for row in range(4, ws.max_row + 1):
        cell = ws.cell(row=row, column=2)  # B列
        cell.alignment = Alignment(wrap_text=True)

    # A列からE列の2行目と3行目を結合
    for col in range(1, 6):  # 1から5 (A～E列)
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)

    # A列からE列の2行目と3行目を結合
    for col in range(1, 6):  # 1から5 (A～E列)
        ws.merge_cells(start_row=2, start_column=col, end_row=3, end_column=col)

    # 下線のスタイルを定義
    thin_border = Border(bottom=Side(style='thin'))  # 細い下線

    # A列から最右列まで1行目に下線を引く
    for col in range(1, ws.max_column + 1):  # 1から最右列まで
        cell = ws.cell(row=1, column=col)
        cell.border = thin_border
        
    # 行の高さを設定
    ws.row_dimensions[1].height = 30 # ヘッダー1行目
    ws.row_dimensions[3].height = 200  # 3行目の高さを200pxに設定

def convert_app_acl_to_excel(wb, header_name, base_dir, group_map, entity_type_map, userName_map, userValid_map, sheet_name):
  """
  app_acl.yaml をアプリシートに変換する

  Args:
    wb (Workbook): openpyxl Workbookオブジェクト
    header_name (str): ヘッダー名（例: '52'）
    base_dir (str): ディレクトリのパス
    group_map (dict): エンティティコードからグループ名またはフィールドラベルへのマッピング
    entity_type_map (dict): エンティティコードからタイプへのマッピング
    userValid_map (dict): ユーザーコードから有効性へのマッピング
    sheet_name (str): Excelシート名 ('アプリ')
  """
  input_file = os.path.join(base_dir, f"{header_name}_app_acl.yaml")

  # YAMLファイルの読み込み
  try:
    with open(input_file, 'r', encoding='utf-8') as f:
      data = yaml.safe_load(f)
  except Exception as e:
    logging.error(f"エラー: app_aclファイル {input_file} の読み込みに失敗しました: {str(e)}")
    return

  # 全エンティティを取得
  all_entities = get_all_entities(data)

  # シートの取得または作成
  if sheet_name in wb.sheetnames:
    ws = wb[sheet_name]
  else:
    ws = wb.create_sheet(title=sheet_name)

  # ヘッダーの設定
  headers = ['No', '種別', 'code', 'レコード閲覧', 'レコード追加', 'レコード編集', 'レコード削除', 'アプリ管理', 'ファイル読み込み', 'ファイル書き出し']
  for col_num, header in enumerate(headers, 1):
    create_header_cell(ws, 1, col_num, header)
 
  # 行の高さを設定
  ws.row_dimensions[1].height = 30 # ヘッダー1行目

  # データ行の書き込み
  current_row = 2
  for i, rights_block in enumerate(data.get('rights', []), 1):
    entity = rights_block.get('entity', {})
    code = entity.get('code', '')
    entity_type = entity.get('type', '')

    # 種別を日本語に変換
    if entity_type == 'USER':
      type_jp = 'ユーザ'
    elif entity_type == 'GROUP':
      type_jp = 'グループ'
    elif entity_type == 'FIELD_ENTITY':
      type_jp = 'フィールド'
    elif entity_type == 'CREATOR':
      type_jp = 'アプリ作成者'
    else:
      type_jp = entity_type # 未知のタイプはそのまま表示

    # 各権限を '●' または '－' に変換
    permissions = {
      'レコード閲覧': '●' if rights_block.get('recordViewable', False) else '－',
      'レコード追加': '●' if rights_block.get('recordAddable', False) else '－',
      'レコード編集': '●' if rights_block.get('recordEditable', False) else '－',
      'レコード削除': '●' if rights_block.get('recordDeletable', False) else '－',
      'アプリ管理': '●' if rights_block.get('appEditable', False) else '－',
      'ファイル読み込み': '●' if rights_block.get('recordImportable', False) else '－',
      'ファイル書き出し': '●' if rights_block.get('recordExportable', False) else '－',
    }

    # グループ名またはフィールドラベルに置換
    group_name = group_map.get(code, '')
    invalid_group = False
    invalid_user = False
    if entity_type == 'GROUP' or entity_type == 'FIELD_ENTITY':
      if not group_name:
        invalid_group = True
        group_name = code
    elif entity_type == 'USER':
      # とりあえずentityをそのまま使用
      user_name = entity
      if user_name['code'] not in userName_map:
        logging.warning(f'NOT IN: {user_name["code"]}')
    else:
      invalid_group = True # 未知のタイプは無効とする

    # データを各セルに書き込み
    for col_num, header in enumerate(headers, 1):
      if header == 'No':
        value = i
        cell = ws.cell(row=current_row, column=col_num, value=value)
      elif header == '種別':
        value = type_jp
        cell = ws.cell(row=current_row, column=col_num, value=value)
      elif header == 'code':
        value = group_name
        cell = ws.cell(row=current_row, column=col_num, value=value)
      else:
        value = permissions.get(header, '－')
        cell = ws.cell(row=current_row, column=col_num, value=value)
     
      # 権限セルのアライメント設定
      if header in ['レコード閲覧', 'レコード追加', 'レコード編集', 'レコード削除', 'アプリ管理', 'ファイル読み込み', 'ファイル書き出し']:
        cell.alignment = Alignment(horizontal='center', vertical='center')
      else:
        cell.alignment = Alignment(wrap_text=True, vertical='top', horizontal='center')

      # ボーダー設定
      cell.border = Border(left=Side(style='thin'), right=Side(style='thin'),
                top=Side(style='thin'), bottom=Side(style='thin'))

      # グループまたはユーザーが無効の場合、セルのフォントを赤色の太字に
      if (header == 'code') and (invalid_group or invalid_user):
        cell.font = Font(bold=True, color='FF0000')

    current_row += 1

  # 列幅の設定
  column_widths = {
    'A': 5,  # No
    'B': 10,  # 種別
    'C': 50,  # code
    'D': 12,  # レコード閲覧
    'E': 12,  # レコード追加
    'F': 12,  # レコード編集
    'G': 12,  # レコード削除
    'H': 12,  # アプリ管理
    'I': 16,  # ファイル読み込み
    'J': 16,  # ファイル書き出し
  }
  for col, width in column_widths.items():
    ws.column_dimensions[col].width = width

  # 行の高さの設定（データ行）
  for row in range(2, current_row):
    ws.row_dimensions[row].height = 20 # 適宜調整

def compare_permissions_and_mark(wb, group_map, group_members, header_name, base_dir):
    logging.debug(f"group_map: {group_map}")
    logging.debug(f"group_members: {group_members}")
    """
    レコードシートとアプリシートの権限を比較し、矛盾をマークする

    Args:
        wb (Workbook): openpyxl Workbookオブジェクト
        group_map (dict): グループコードをグループ名にマッピングする辞書
        group_members (dict): グループコードをメンバー情報にマッピングする辞書
        header_name (str): ヘッダー名（例: '52'）
        base_dir (str): ディレクトリのパス
    """
    # シートの取得
    try:
        record_ws = wb['レコード']
        app_ws = wb['アプリ']
    except KeyError as e:
        logging.error(f"エラー: シート {e} が見つかりません。")
        return

    # アプリシートのヘッダー行から権限列のインデックスを取得
    app_headers = {cell.value: idx for idx, cell in enumerate(app_ws[1], 1)}
    target_permissions = {
        'レコード閲覧': '閲覧',
        'レコード編集': '編集',
        'レコード削除': '削除'
    }

    # アプリシートのグループ順序と権限をリストとして保持
    app_group_order = []
    app_permissions = {}
    everyone_permissions = set()

    for row in range(2, app_ws.max_row + 1):
        group_code = app_ws.cell(row=row, column=3).value
        if not group_code:
            continue

        # グループ種別を取得（種別列）
        entity_type = app_ws.cell(row=row, column=2).value

        # フィールドエンティティはスキップ
        if entity_type == 'フィールド':
            continue

        permissions = set()
        for app_perm_key, user_perm in target_permissions.items():
            col_idx = app_headers.get(app_perm_key)
            if col_idx and app_ws.cell(row=row, column=col_idx).value == '●':
                permissions.add(user_perm)

        # グループ名を取得
        group_name = group_map.get(group_code, group_code)

        # "Everyone" の権限を保持
        if group_name.lower() == 'everyone':
            everyone_permissions = permissions
            continue

        # グループ順序リストに追加
        app_group_order.append(group_name)
        app_permissions[group_name] = permissions

    # 個人ユーザーの所属グループをマッピング
    user_to_groups = {}
    for group_code, info in group_members.items():
        group_name = group_map.get(group_code, group_code)
        for user in info.get('users', []):
            username = user.get('username')
            if username:
                user_to_groups.setdefault(username, set()).add(group_name)

    # 各ユーザーの適用権限を決定
    user_effective_permissions = {}
    for user, groups in user_to_groups.items():
        effective_perms = None

        # アプリシートのグループ順に従って権限を決定
        for group in app_group_order:
            if group == user:
                # ユーザー自身がグループとして存在する場合
                effective_perms = app_permissions.get(group, set())
                break
            if group in groups:
                effective_perms = app_permissions.get(group, set())
                break

        if effective_perms is None:
            # どのグループにも該当しない場合はEveryoneの権限
            effective_perms = everyone_permissions

        user_effective_permissions[user] = effective_perms

    def format_permissions(permissions, all_types=('閲覧', '編集', '削除'), left = '　', right = '　', non_val = '    '):
        """権限を固定位置に揃えて表示する"""
        return '　'.join(f'{left}{perm}{right}' if perm in permissions else f'　{non_val}　' for perm in all_types)

    # 警告情報を収集するための辞書とカウンターを準備
    warnings_collection = {}
    warnings_counter = Counter()

    # レコードシートのグループごとの権限をチェック
    for row in range(4, record_ws.max_row + 1):
        group_name = record_ws.cell(row=row, column=3).value
        record_perm_str = record_ws.cell(row=row, column=4).value

        if not group_name or not record_perm_str:
            continue

        if group_name.endswith('(フィールド)'):
            continue

        # グループかユーザーかを判定
        is_group = group_name in group_map.values()

        if is_group:
            # グループの場合はアプリシートの権限を使用
            app_perms = app_permissions.get(group_name, everyone_permissions)
        else:
            # ユーザーの場合は集計済みの権限を使用
            app_perms = user_effective_permissions.get(group_name, everyone_permissions)

        # レコードシートの権限をパース
        record_permissions = set(filter(None, map(str.strip, record_perm_str.strip('/').split('/'))))
        extra_permissions = record_permissions - app_perms

        if extra_permissions:
            red_font = Font(color='FF0000')
            c_cell = record_ws.cell(row=row, column=3)
            d_cell = record_ws.cell(row=row, column=4)
            c_cell.font = red_font
            d_cell.font = red_font

            # 警告情報を収集
            entity_type = 'グループ' if is_group else 'ユーザー'
            warning_key = (entity_type, group_name, tuple(sorted(record_permissions)), tuple(sorted(app_perms)))
            warnings_collection[warning_key] = tuple(sorted(extra_permissions))
            warnings_counter[warning_key] += 1

    # CSV出力用のファイルパスを構築（base_dirを使用）
    csv_path = os.path.join(base_dir, f"{header_name}_acl_problem.csv")
    
    # CSVファイルに警告情報を出力（警告がない場合も空ファイルを作成）
    with open(csv_path, 'w', encoding='utf-8', newline='') as f:
        writer = csv.writer(f, delimiter='\t')
        # ヘッダー行を書き込み
        writer.writerow(['アプリ番号', 'タイプ', '名称', '矛盾タイプ', '出現回数', '過剰な権限'])
        
        if len(warnings_collection.items()) > 0:
          # 警告情報を書き込み
          for (entity_type, group_name, record_perms, app_perms), extra_perms in warnings_collection.items():
              count = warnings_counter[(entity_type, group_name, record_perms, app_perms)]
              # 過剰な権限を文字列に変換
              extra_perms_str = '/'.join(sorted(extra_perms))
              writer.writerow([
                  header_name,
                  entity_type,
                  group_name,
                  'レコードにあるがアプリに無い',
                  count,
                  extra_perms_str
              ])
      
    logging.info(f"権限矛盾情報をCSVファイルに出力しました: {csv_path}")

    if len(warnings_collection.items()) > 0:

        # 既存の警告ログ出力
        logging.warning("=== 権限矛盾の集計結果 ===")
        for (entity_type, group_name, record_perms, app_perms), extra_perms in warnings_collection.items():
            count = warnings_counter[(entity_type, group_name, record_perms, app_perms)]
            logging.warning(f"警告: {entity_type} '{group_name}' の権限矛盾 (出現回数: {count}回)")
            logging.warning(f"  レコードシート: {format_permissions(record_perms)}")
            logging.warning(f"    アプリシート: {format_permissions(app_perms)}")
            logging.warning(f"      過剰な権限: {format_permissions(extra_perms, left='【', right='】', non_val='----')}")
            logging.warning(f"")

        # 合計の出力
        total_warnings = sum(warnings_counter.values())
        unique_warnings = len(warnings_collection)
        logging.warning(f"=== 集計サマリー ===")
        logging.warning(f"総警告数: {total_warnings}件")
        logging.warning(f"ユニークな警告数: {unique_warnings}件")

def load_group_members(group_list_path):
  """
  group_user_list.yaml からグループとメンバーの対応を読み込む

  Args:
    group_list_path (str): group_user_list.yaml のファイルパス

  Returns:
    dict: グループ情報の辞書
  """
  try:
    with open(group_list_path, 'r', encoding='utf-8') as f:
      return yaml.safe_load(f)
  except Exception as e:
    logging.warning(f"警告: group_user_list.yaml の読み込みに失敗しました: {str(e)}")
    return {}

def print_group_members(group_data):
    """
    各グループのメンバー一覧をログ出力する
    """
    logging.debug("=== グループメンバー一覧 ===")
    for group_code, group_info in group_data.items():
        logging.debug(f"\nグループ: {group_code}")
        logging.debug(f"グループ名: {group_info['name']}")
        logging.debug("メンバー:")
        for user in group_info['users']:
            logging.debug(f"  - ユーザー名: {user['username']}")
            logging.debug(f"    メールアドレス: {user['email']}")
            logging.debug(f"    ID: {user['id']}")

def main():
  """
  スクリプトのエントリーポイント
  """
  parser = argparse.ArgumentParser(description='YAMLファイルをExcelファイルに変換するスクリプト')
  parser.add_argument('header_name', type=str, help='ヘッダー名 (例: 14)')
  parser.add_argument('--group-master', '-g', type=str, 
                     default=os.path.join('..', 'kintone_get_user_group', 'group_user_list.yaml'),
                     help='グループマスタファイルのパス')
  parser.add_argument('--log-level', type=str, choices=['DEBUG', 'INFO', 'WARNING', 'ERROR', 'CRITICAL'],
                     default='INFO', help='ログレベル (デフォルト: INFO)')
  parser.add_argument('--silent', action='store_true', help='ログ出力を抑制する')

  args = parser.parse_args()

  # ロギングの設定
  setup_logging(args.log_level, args.silent)

  logging.info(f"処理を開始します: ヘッダー名 = {args.header_name}")
  
  header_name = args.header_name
  group_master_path = args.group_master

  output_dir = os.path.join(os.getcwd(), 'output')
  logging.debug(f"出力ディレクトリ: {output_dir}")

  # "{header_name}_" で始まるディレクトリを検索
  matching_dirs = [d for d in os.listdir(output_dir) if d.startswith(f"{header_name}_") and os.path.isdir(os.path.join(output_dir, d))]

  if not matching_dirs:
    logging.error(f'"{header_name}_" で始まるディレクトリが見つかりません。')
    sys.exit(1)
  elif len(matching_dirs) > 1:
    logging.error(f'複数の "{header_name}_" で始まるディレクトリが見つかりました: {matching_dirs}')
    logging.error('一意に特定できるようにディレクトリ名を修正してください。')
    sys.exit(1)

  # 見つかったディレクトリ名からヘッダー部分を除いた名前を取得
  dir_name_without_header = matching_dirs[0][len(f"{header_name}_"):]
  
  # アプリ名_YYYYMMDD_HHMMSS 形式の場合、アプリ名部分のみを抽出
  app_name = dir_name_without_header.split('_')[0] if '_20' in dir_name_without_header else dir_name_without_header
  
  logging.info(f"処理対象アプリ名 = {header_name}, 処理対象 = {app_name}")
  logging.info(f"対象ディレクトリ: {dir_name_without_header}")

  # 見つかったディレクトリのパスを設定
  base_dir = os.path.join(output_dir, matching_dirs[0])
  record_acl_file = os.path.join(base_dir, f"{header_name}_record_acl.yaml")
  app_acl_file = os.path.join(base_dir, f"{header_name}_app_acl.yaml")

  # 必要なファイルが存在するか確認
  if not os.path.exists(record_acl_file):
    logging.error(f'エラー: ファイル "{record_acl_file}" が見つかりません。')
    sys.exit(1)
  if not os.path.exists(app_acl_file):
    logging.error(f'エラー: ファイル "{app_acl_file}" が見つかりません。')
    sys.exit(1)

  # エンティティタイプマップの作成
  entity_type_map = load_entity_type_map(header_name, base_dir)

  # グループマスタとフォームフィールドの読み込み
  field_entities = load_form_fields(header_name, base_dir)
  group_map = load_group_map(header_name, base_dir, group_master_path, field_entities)
  if not group_map:
    logging.warning(f'警告: グループマスタおよびフォームフィールドファイルが正しく読み込まれていません。グループ名はグループコードのまま表示されます。')

  # userName_map の読み込み
  userName_map = load_userName_list(group_master_path)
  if not userName_map:
      logging.warning(f'警告: group_user_list.yaml が正しく読み込まれていません。USERタイプのエンティティは無効として扱われます。: {group_master_path}')

  # グループメンバー情報の読み込みと出力
  group_members = load_group_members(group_master_path)
  print_group_members(group_members)

  # record_aclとapp_aclファイルからエンティティを読み込む
  try:
    with open(record_acl_file, 'r', encoding='utf-8') as f:
      record_data = yaml.safe_load(f)
    with open(app_acl_file, 'r', encoding='utf-8') as f:
      app_data = yaml.safe_load(f)
      
    # 両方のファイルからエンティティを抽出
    record_entities = set(get_all_entities(record_data))
    app_entities = set(get_all_entities(app_data))
    
    # 全エンティティを結合してソート
    all_entities = sorted(record_entities.union(app_entities))
    
    logging.debug("\n=== 全出現ユーザ/グループ一覧（everyoneを除く） ===")
    for entity in all_entities:
      entity_type = entity_type_map.get(entity, 'USER')  # デフォルトは 'USER'
      if entity_type == 'GROUP':
        group_info = group_members.get(entity, {})
        group_name = group_info.get('name', entity)
        logging.debug(f"\nグループ: {group_name} (コード: {entity})")
        users = group_info.get('users', [])
        if users:
            logging.debug("所属ユーザー:")
            for user in users:
                username = user.get('username', '不明')
                email = user.get('email', '不明')
                user_id = user.get('id', '不明')
                logging.debug(f"  - ユーザー名: {username}, メールアドレス: {email}, ID: {user_id}")
        else:
            logging.debug("  ※ 所属ユーザーなし")
      elif entity_type == 'USER':
        # ユーザー名を表示するためにユーザー情報を取得
        user_name = userName_map.get(entity, entity)  # マッピングがなければコードを名前として使用
        logging.debug(f"ユーザー: {entity} (名前: {user_name})")
      else:
        # その他のタイプ
        logging.debug(f"その他エンティティ: {entity} (タイプ: {entity_type})")
        
    # 全出現ユーザから重複を除いた一覧の作成
    permission_target_user_names = set()

    for entity in all_entities:
        entity_type = entity_type_map.get(entity, 'USER')  # デフォルトは 'USER'
        if entity_type == 'GROUP':
            group_info = group_members.get(entity, {})
            users = group_info.get('users', [])
            for user in users:
                user_name = user.get('username', '不明')
                permission_target_user_names.add(user_name)
        elif entity_type == 'USER':
            user_name = userName_map.get(entity, entity)  # マッピングがなければコードを名前として使用
            permission_target_user_names.add(user_name)
        # その他のタイプはユーザーではないと仮定

    logging.debug("\n=== 全出現ユーザで重複を除いた単純なユーザ名一覧 ===")
    for user_name in sorted(permission_target_user_names):
        logging.debug(f"ユーザー名: {user_name}")

    # 重複を除いた単純なユーザ名一覧をCSVファイルに保存
    with open(os.path.join(base_dir, f"{header_name}permission_target_user_names.csv"), 'w', encoding='utf-8', newline='') as csvfile:
        csvwriter = csv.writer(csvfile)
        csvwriter.writerow(['ユーザー名'])
        for user_name in sorted(permission_target_user_names):
            csvwriter.writerow([user_name])

    logging.debug(f'全出現ユーザでユニークユーザー名一覧を {os.path.join(base_dir, f"{header_name}permission_target_user_names.csv")} に保存しました。')

  except Exception as e:
    logging.error(f"エラー: ACLファイルの読み込みに失敗しました: {str(e)}")


  # Excelワークブックの作成
  wb = Workbook()
  # デフォルトで作成されるシートを削除
  default_sheet = wb.active
  wb.remove(default_sheet)

  # record_acl.yaml を「レコード」シートに変換
  convert_yaml_to_excel(wb, header_name, base_dir, group_map, entity_type_map, user_map={}, 
                       acl_type='record', sheet_name='レコード',
                       userName_map=userName_map, group_members=group_members, permission_target_user_names=permission_target_user_names)

  # app_acl.yaml を「アプリ」シートに変換
  convert_app_acl_to_excel(wb, header_name, base_dir, group_map, entity_type_map, 
                          userName_map=userName_map, userValid_map={}, sheet_name='アプリ')

  # 新規追加: レコードシートとアプリシートの権限を比較してマークする
  compare_permissions_and_mark(wb, group_map, group_members, header_name, base_dir)

  # ファイルを保存
  output_file = os.path.join(base_dir, f"{header_name}_acl.xlsx")
  try:
      wb.save(output_file)
      logging.info(f'変換完了: {output_file}')
  except PermissionError:
      logging.error(f'ファイル "{output_file}" が他のプログラムで開かれています。')
      logging.error('Excelを閉じてから再度実行してください。')
      sys.exit(1)
  except Exception as e:
      logging.error(f'ファイルの保存中に予期せぬエラーが発生しました: {str(e)}')
      sys.exit(1)

if __name__ == '__main__':
  main()
