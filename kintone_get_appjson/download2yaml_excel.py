import os
import sys
import requests
import yaml
import json
import re
import subprocess
import csv
import shutil
import base64
from datetime import datetime
from pathlib import Path
from collections import defaultdict
from bs4 import BeautifulSoup
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from openpyxl.utils.cell import column_index_from_string, get_column_letter, coordinate_from_string
from typing import Union

# ─── 補助関数 ─────────────────────────────────────────────
def process_file(layout_file_path, fields_file_path, output_file):
    """レイアウトファイルとフィールドファイルを処理してTSVを生成"""
    with open(layout_file_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()

    with open(output_file, 'w', encoding='utf-8') as out_f:
        type_pattern = re.compile(r'\"type\":\s*\"([^\"]+)\"')
        label_code_pattern = re.compile(r'\"(label|code|elementId)\":\s*\"(.+)\"')
        indent_level = 0
        current_type = None
        current_group = None
        group_counter = 0
        group_indent = None
        current_subtable = None
        subtable_counter = 0
        subtable_indent = None
        current_italic = None
        italic_counter = 0
        italic_in_group = False

        for i, line in enumerate(lines):
            indent_level += line.count('{') - line.count('}')
            if current_group is not None and indent_level < group_indent:
                current_group = None
                if italic_in_group:
                    current_italic = None
            if subtable_indent is not None and indent_level < subtable_indent:
                current_subtable = None

            type_match = type_pattern.search(line)
            if type_match:
                current_type = type_match.group(1)
                if current_type == "GROUP":
                    group_counter += 1
                    current_group = group_counter
                    group_indent = indent_level
                    continue
                if current_type == "SUBTABLE":
                    subtable_counter += 1
                    current_subtable = subtable_counter
                    subtable_indent = indent_level
                    continue
                if current_type == "HR":
                    out_f.write(f"{indent_level}\t{current_italic or ''}\t{current_group or ''}\t{current_subtable or ''}\t{current_type}\n")
                    current_type = None
                    continue

            label_code_match = label_code_pattern.search(line)
            if label_code_match and current_type:
                key_type = label_code_match.group(1)
                key_value = label_code_match.group(2)
                if current_type == "SPACER" and key_type == "elementId":
                    out_f.write(f"{indent_level}\t{current_italic or ''}\t{current_group or ''}\t{current_subtable or ''}\t{current_type}\t{key_value}\t\n")
                elif key_type == "code":
                    additional_properties = grep_code_properties(fields_file_path, key_value)
                    additional_info = ', '.join([f"{k}: {v}" for k, v in additional_properties.items()])
                    out_f.write(f"{indent_level}\t{current_italic or ''}\t{current_group or ''}\t{current_subtable or ''}\t{current_type}\t{key_value}\t\t\t\t\t{additional_info}\n")
                else:
                    if key_type == 'label' and (('background-color:rgb(' in key_value and len(key_value) < 30) or ('<i>' in key_value)):
                        italic_counter += 1
                        current_italic = italic_counter
                        italic_in_group = True if current_group is not None else False
                        soup = BeautifulSoup(key_value, 'html.parser')
                        tmp_key_value = soup.get_text().strip()
                        out_f.write(f"{indent_level}\t{current_italic or ''}\t{current_group or ''}\t{current_subtable or ''}\t{current_type}\t\t{tmp_key_value}\n")
                    else:
                        if key_type == 'label':
                            soup = BeautifulSoup(key_value, 'html.parser')
                            tmp_key_value = soup.get_text().strip()
                            out_f.write(f"{indent_level}\t{current_italic or ''}\t{current_group or ''}\t{current_subtable or ''}\t{current_type}\t\t\t\t\t\t{tmp_key_value}\n")
                        else:
                            out_f.write(f"{indent_level}\t{current_italic or ''}\t{current_group or ''}\t{current_subtable or ''}\t{current_type}\t{key_value}\t\n")
                current_type = None

def grep_code_properties(fields_file_path, target_code):
    """form_fields.jsonから指定したコードのプロパティを抽出"""
    with open(fields_file_path, 'r', encoding='utf-8') as f:
        lines = f.readlines()
    code_properties = {}
    is_matching_code = False
    indent_level = 0
    for line in lines:
        indent_level += line.count('{') - line.count('}')
        if f'"code": "{target_code}"' in line:
            is_matching_code = True
            indent_code = indent_level
            continue
        if is_matching_code:
            if ':' in line:
                key_value_match = re.match(r'\"(.+?)\":\s*(.+)', line.strip())
                if key_value_match:
                    key, value = key_value_match.groups()
                    if key not in code_properties:
                        code_properties[key] = value.strip().rstrip(',')
            if indent_level < indent_code:
                break
    return code_properties

def process_raw_layout(input_file, output_file):
    """TSVファイルを処理して不要な行を削除・修正"""
    with open(input_file, 'r', encoding='utf-8') as infile:
        reader = csv.reader(infile, delimiter='\t')
        rows = list(reader)
    result = []
    skip_next = False
    label_col2_to_space = False
    for i, row in enumerate(rows):
        row[0] = ''
        if label_col2_to_space:
            if row[1] == label_col2_Number:
                row[1] = ''
            else:
                label_col2_to_space = False
        if skip_next:
            skip_next = False
            continue
        if len(row) > 10:
            label_match = re.search(r'label: "(.*?)"', row[10])
            if label_match and row[4] not in ['GROUP']:
                row[6] = label_match.group(1)
        if row[4] in ['HR']:
            continue
        if row[4] in ['GROUP'] and rows[i + 1][4] in ['LABEL'] and rows[i + 1][6] != '' and row[1] != '' and rows[i+1][1] != '':
            row[1] = ''
        if row[4] in ['GROUP'] and rows[i + 1][4] in ['LABEL'] and rows[i + 1][6] != '' and row[1] == '':
            label_col2_to_space = True
            label_col2_Number = rows[i+1][1]
        if row[4] in ['LABEL'] and row[6] == '':
            continue
        if row[4] in ['RECORD_NUMBER']:
            row[8] = '必須'
        if row[4] in ['SINGLE_LINE_TEXT', 'MULTI_LINE_TEXT', 'DATE', 'DATETIME', 'NUMBER']:
            require_true_match = re.search(r'required: true', row[10])
            if require_true_match:
                row[8] = '必須'
        if row[4] == 'GROUP' and i + 1 < len(rows) and rows[i + 1][4] == 'LABEL':
            row[6] = rows[i + 1][6]
            skip_next = True
        row[0] = 0
        row[0] += 1 if row[1] != '' else 0
        row[0] += 1 if row[2] != '' else 0
        row[0] += 1 if row[3] != '' else 0
        result.append(row)
    with open(output_file, 'w', newline='', encoding='utf-8') as outfile:
        writer = csv.writer(outfile, delimiter='\t')
        writer.writerows(result)


def flatten_record(record):
    """レコードをフラット化し、ネストされた 'value' フィールドを展開"""
    flattened = {}
    for key, value in record.items():
        extracted = None
        formatted_value = None
        sub_key = None
       
        if isinstance(value, dict):
            if 'value' in value:
                extracted = extract_value(value)
                formatted_value = format_custom_fields(flattened, key, extracted)
                if isinstance(extracted, dict): # システムフィールドのとき  作成者(type:CREATER)、更新者(TYPE:MODIFIER)
                    for sub_key, sub_value in extracted.items():
                        flattened[key] = replace_custom_format(formatted_value)
                else:
                    flattened[key] = replace_custom_format(formatted_value)
            elif 'type' in value and 'value' in value:
                if isinstance(value['value'], dict):
                    for sub_key, sub_value in value['value'].items():
                        flattened[sub_key] = clean_string(sub_value)
                else:
                    flattened[key] = clean_string(value['value'])
            else:
                for sub_key, sub_value in value.items():
                    flattened[sub_key] = extract_value(sub_value)
        else:
            flattened[key] = clean_string(value)
       
    return flattened


def extract_value(field_data):
    """フィールドデータから値を抽出"""
    if isinstance(field_data, dict):
        return field_data.get('value', field_data)
    return field_data

def clean_string(value):
    """文字列をクリーンアップ"""
    if isinstance(value, str):
        return value.strip()
    return value

def replace_custom_format(value):
    """カスタムフォーマットを置換"""
    if isinstance(value, str):
        return value.replace('\r\n', '\n').replace('\r', '\n')
    return value

def format_custom_fields(record, key, value):
    """リストの場合は結合、その他はそのまま返す"""
    if isinstance(value, list):
        return ', '.join(str(v) for v in value)
    return value

def extract_field_codes_with_lines(filepath):
    """JavaScriptファイルからフィールドコードの使用箇所を抽出"""
    patterns = [
        re.compile(r'record\[\s*["\']([\w-]+)["\']\s*\]'),
        re.compile(r'kintone\.app\.record\.\w+\(\s*["\']([\w-]+)["\']'),
        re.compile(r'event\.record\.([\w-]+)\.value'),
        re.compile(r'\["([^"]+)"\]\)\]\)},fanction\(\){'),
    ]
    result = defaultdict(list)
    try:
        with open(filepath, 'r', encoding='utf-8') as f:
            for lineno, line in enumerate(f, start=1):
                for pattern in patterns:
                    for match in pattern.findall(line):
                        result[match].append(lineno)
    except Exception as e:
        print(f"Error reading {filepath}: {e}")
    return {field: sorted(set(lines)) for field, lines in result.items()} if result else {}

def prepare_kaigyo_files(js_dir):
    """1行が1000文字を超える行があるJavaScriptファイルを処理し、._kaigyo_.jsファイルを生成"""
    for file_path in js_dir.glob('*.js'):
        with open(file_path, 'r', encoding='utf-8', errors='replace') as f:
            lines = f.readlines()

        # 1行が1000文字を超える行があるかチェック
        long_lines_exist = any(len(line) > 1000 for line in lines)

        if long_lines_exist:
            # 元のファイルを .js_moto にリネーム
            moto_file_path = file_path.with_suffix('.js_moto')
            file_path.rename(moto_file_path)

            # ._kaigyo_.js ファイルを生成
            kaigyo_file_path = file_path.with_name(file_path.stem + '._kaigyo_.js')
            with open(kaigyo_file_path, 'w', encoding='utf-8') as f:
                for line in lines:
                    if len(line) > 10:
                        parts = line.split(';')
                        for part in parts:
                            if part.strip():
                                f.write(part.strip() + ';\n')
                    else:
                        f.write(line)

def scan_directory_for_field_codes_with_lines(js_dir):
    """ディレクトリ内のJavaScriptファイルをスキャンしてフィールドコードの使用箇所をマップ化"""
    field_code_map = defaultdict(dict)
    for file_path in js_dir.glob('*.js'):
        # ._kaigyo_.js ファイルが存在する場合はそれを使用
        kaigyo_file_path = file_path.with_name(file_path.stem + '._kaigyo_.js')
        if kaigyo_file_path.exists():
            file_result = extract_field_codes_with_lines(kaigyo_file_path)
        else:
            file_result = extract_field_codes_with_lines(file_path)

        if file_result:
            for field, lines in file_result.items():
                field_code_map[field][file_path.name] = lines

    return dict(field_code_map)

# ─── ExcelFormatter クラス ─────────────────────────────────────────────
class ExcelFormatter:
    def __init__(self, workbook=None, worksheet=None, filename='output.xlsx', background_color='FF95B3D7'):
        self.wb = workbook if workbook else Workbook()
        self.ws = worksheet if worksheet else self.wb.active
        self.filename = filename
        self.background_color = background_color
        self.fill = PatternFill(start_color=self.background_color, end_color=self.background_color, fill_type="solid")
        self.thin_border = Border(left=Side(style='thin'), right=Side(style='thin'),
                                  top=Side(style='thin'), bottom=Side(style='thin'))
        self.font = Font(name='メイリオ', size=9)

    def set_row_height(self, row_count=200, height_px=20):
        row_height = height_px / 1.33
        for row in range(1, row_count + 1):
            self.ws.row_dimensions[row].height = row_height

    def set_column_width(self, start_col=1, end_col=26, width_px=25):
        column_width = width_px / 7
        for col in range(start_col, end_col + 1):
            col_letter = get_column_letter(col)
            self.ws.column_dimensions[col_letter].width = column_width

    def merge_cells_and_set_content(self, start_cell, end_cell, text,
                                    alignment="left", bottom_border=False, right_border=False,
                                    isMerge=True, isBackcolor=True):
        if isMerge:
            self.ws.merge_cells(f'{start_cell}:{end_cell}')
        cell = self.ws[start_cell]
        cell.value = text if text is not None else cell.value
        cell.font = self.font
        cell.alignment = Alignment(horizontal=alignment, vertical='center')
        if isBackcolor:
            cell.fill = self.fill

        border_sides = {
            'left': Side(style='thin'),
            'top': Side(style='thin'),
            'bottom': Side(style='thin') if bottom_border else None,
            'right': Side(style='thin') if right_border else None
        }
        # 新規 Border オブジェクトを生成
        border = Border(**{k: v for k, v in border_sides.items() if v is not None})
        cell.border = border

        cells = self.ws[f'{start_cell}:{end_cell}']
        for row in cells:
            for cell in row:
                if isBackcolor:
                    cell.fill = self.fill
                cell.border = border

    def move_cell_str(self, cell_ref, direction):
        col_str, row = coordinate_from_string(cell_ref)
        col = column_index_from_string(col_str)
        if direction == "right":
            col += 1
        elif direction == "left":
            col -= 1
        elif direction == "up":
            row -= 1
        elif direction == "down":
            row += 1
        else:
            raise ValueError("不正な方向です。right, left, up, down のいずれかを指定してください。")
        new_col_str = get_column_letter(col)
        return f"{new_col_str}{row}"

    def set_by_out02_tsv(self, tsv_filename):
        from openpyxl.utils import get_column_letter

        def set_val_font(in_cell, in_value):
            in_cell.value = in_value
            in_cell.font = self.font

        light_pink_fill = PatternFill(start_color='FFE6E6', end_color='FFE6E6', fill_type='solid')

        with open(tsv_filename, 'r', encoding='utf-8') as infile:
            reader = csv.reader(infile, delimiter='\t')
            rows = list(reader)
        for i, row in enumerate(rows):
            new_row = [''] * 14
            count_be = 1
            if row[1] != '':
                new_row[count_be] = 'L' + row[1]
                count_be += 1
            if row[2] != '':
                new_row[count_be] = 'G' + row[2]
                count_be += 1
            if row[3] != '':
                new_row[count_be] = 'S' + row[3]
                count_be += 1
            new_row[0] = row[0]
            indent_level = int(row[0])
            start_index = 2 + indent_level
            
            # フィールドタイプに応じて項目名を設定
            if row[4] == 'SPACER':
                field_name = f"スペース ({row[5]})"  # スペース型のみフィールドコードを含める
            elif row[4] == 'GROUP':
                field_name = row[5] + ' ' + row[6]  # グループは従来通り
            else:
                field_name = row[6]  # その他のフィールドは従来通り
            
            new_row[start_index] = field_name
            new_row[6] = '〇' if (row[8] if len(row) > 8 else '') == '必須' else ''
            set_val_font(self.ws[f'B{i+3}'], new_row[0])
            set_val_font(self.ws[f'C{i+3}'], new_row[1])
            set_val_font(self.ws[f'D{i+3}'], new_row[2])
            set_val_font(self.ws[f'E{i+3}'], new_row[3])
            set_val_font(self.ws[f'F{i+3}'], new_row[4])
            set_val_font(self.ws[f'G{i+3}'], new_row[5])
            set_val_font(self.ws[f'S{i+3}'], new_row[6])
            if len(row) > 5 and row[4] not in ['GROUP', 'LABEL', 'HR']:
                field_code = row[5]
                if row[4] == 'SPACER':
                    set_val_font(self.ws[f'BA{i+3}'], field_code)
                elif row[4] not in ['GROUP', 'LABEL', 'HR']:
                    set_val_font(self.ws[f'BA{i+3}'], field_code)
            if len(row) > 4:
                field_type = row[4]
                field_type_ja = {
                    'SINGLE_LINE_TEXT': '文字列（1行）',
                    'MULTI_LINE_TEXT': '文字列（複数行）',
                    'RICH_TEXT': 'リッチエディター',
                    'NUMBER': '数値',
                    'CALC': '計算',
                    'DATE': '日付',
                    'TIME': '時刻',
                    'DATETIME': '日時',
                    'DROP_DOWN': 'ドロップダウン',
                    'RADIO_BUTTON': 'ラジオボタン',
                    'CHECK_BOX': 'チェックボックス',
                    'MULTI_SELECT': '複数選択',
                    'FILE': '添付ファイル',
                    'LINK': 'リンク',
                    'USER_SELECT': 'ユーザー選択',
                    'GROUP_SELECT': 'グループ選択',
                    'ORGANIZATION_SELECT': '組織選択',
                    'STATUS': 'ステータス',
                    'ASSIGNEE': '作業者',
                    'CATEGORY': 'カテゴリー',
                    'GROUP': 'グループ',
                    'SUBTABLE': 'テーブル',
                    'REFERENCE_TABLE': '関連レコード一覧',
                    'LABEL': 'ラベル',
                    'HR': '罫線',
                    'SPACER': 'スペース'
                }.get(field_type, field_type)
                set_val_font(self.ws[f'BB{i+3}'], field_type_ja)
                if field_type == 'DROP_DOWN' and len(row) > 10:
                    options_str = row[10]
                    options = []
                    try:
                        items = options_str.split(',')
                        for item in items:
                            if ': {' in item:
                                option = item.split(': {')[0].strip()
                                if option not in ['options', 'index', 'defaultValue'] and not option.startswith('"'):
                                    options.append(option)
                        if options:
                            set_val_font(self.ws[f'BC{i+3}'], '選択肢: ' + ', '.join(options))
                    except Exception as e:
                        print(f"選択肢の解析エラー: {e}")

            # SPACERフィールドの行 A列からC列を薄いピンク色に設定
            if row[4] == 'SPACER':
                for col_letter in [get_column_letter(col) for col in range(1, 4)]:
                    cell = self.ws[f'{col_letter}{i+3}']
                    cell.fill = light_pink_fill

            field_start_col_letter = get_column_letter(start_index + 2)
            merge_range = f"{field_start_col_letter}{i+3}:R{i+3}"
            self.ws.merge_cells(merge_range)
            self.merge_cells_and_set_content(f"{field_start_col_letter}{i+3}", f"R{i+3}",
                                             new_row[start_index],
                                             alignment="left",
                                             bottom_border=True,
                                             right_border=True,
                                             isMerge=True,
                                             isBackcolor=False)
            self.merge_cells_and_set_content(f'S{i+3}', f'T{i+3}',
                                             None, alignment="center",
                                             bottom_border=True, right_border=True,
                                             isMerge=True, isBackcolor=False)
            self.merge_cells_and_set_content(f'U{i+3}', f'V{i+3}',
                                             None, alignment="center",
                                             bottom_border=True, right_border=True,
                                             isMerge=True, isBackcolor=False)
            self.merge_cells_and_set_content(f'W{i+3}', f'X{i+3}',
                                             None, alignment="center",
                                             bottom_border=True, right_border=True,
                                             isMerge=True, isBackcolor=False)
            self.merge_cells_and_set_content(f'Y{i+3}', f'AO{i+3}',
                                             None, alignment="left",
                                             bottom_border=True, right_border=True,
                                             isMerge=True, isBackcolor=False)
            set_val_font(self.ws[f'BE{i+3}'], str(row))
            if len(row) > 10:
                set_val_font(self.ws[f'BF{i+3}'], row[10])
        self.get_column_group_arrays()
        L_G = self.get_groups_by_first_char('L')
        G_G = self.get_groups_by_first_char('G')
        S_G = self.get_groups_by_first_char('S')
        shifted_L_G = [self.shift_columns(sublist) for sublist in L_G]
        shifted_G_G = [self.shift_columns(sublist) for sublist in G_G]
        shifted_S_G = [self.shift_columns(sublist) for sublist in S_G]
        self.draw_l_line(shifted_L_G)
        self.draw_l_line(shifted_G_G)
        self.draw_l_line(shifted_S_G, font_color='F2F2F2', background_color='F2F2F2')

    def get_column_group_arrays(self):
        def get_column_groups(column_letter, min_valid_b_value):
            groups = []
            current_group = None
            worksheet = self.ws
            previous_value = None
            for row_cells in worksheet.iter_rows(min_row=1, min_col=1, max_col=5):
                b_cell = row_cells[1]
                b_value = b_cell.value
                if b_value is None:
                    b_value = 0
                else:
                    try:
                        b_value = int(b_value)
                    except ValueError:
                        b_value = 0
                if b_value >= min_valid_b_value:
                    cell = worksheet[f"{column_letter}{b_cell.row}"]
                    value = cell.value
                    if value is not None:
                        if value != previous_value and previous_value is not None:
                            if current_group and current_group['cells']:
                                groups.append(current_group)
                            current_group = {'cells': [], 'first_char': value[0]}
                        elif current_group is None:
                            current_group = {'cells': [], 'first_char': value[0]}
                        current_group['cells'].append(cell.coordinate)
                        previous_value = value
                    else:
                        if current_group and current_group['cells']:
                            groups.append(current_group)
                            current_group = None
                        previous_value = None
                else:
                    if current_group and current_group['cells']:
                        groups.append(current_group)
                        current_group = None
                    previous_value = None
            if current_group and current_group['cells']:
                groups.append(current_group)
            return groups
        self.c_groups = get_column_groups('C', 1)
        self.d_groups = get_column_groups('D', 2)
        self.e_groups = get_column_groups('E', 3)
        self.add_additional_cells(self.c_groups, start_column='C')
        self.add_additional_cells(self.d_groups, start_column='D')
        self.add_additional_cells(self.e_groups, start_column='E')
        return self.c_groups, self.d_groups, self.e_groups

    def add_additional_cells(self, groups, start_column):
        def column_to_number(col):
            num = 0
            for c in col:
                if c.isalpha():
                    num = num * 26 + (ord(c.upper()) - ord('A') + 1)
            return num
        def number_to_column(num):
            col = ''
            while num > 0:
                num, remainder = divmod(num - 1, 26)
                col = chr(65 + remainder) + col
            return col
        worksheet = self.ws
        start_col_index = column_to_number(start_column.upper())
        for group in groups:
            top_row = min(int(cell[1:]) for cell in group['cells'])
            end_col_index = column_to_number('R') if group['first_char'] == 'S' else column_to_number('AO')
            for col_index in range(start_col_index, end_col_index + 1):
                col_letter = number_to_column(col_index)
                cell_coordinate = f"{col_letter}{top_row}"
                if cell_coordinate not in group['cells']:
                    group['cells'].append(cell_coordinate)

    def get_groups_by_first_char(self, char):
        filtered_groups = []
        for group_list in [self.c_groups, self.d_groups, self.e_groups]:
            for group in group_list:
                if group['first_char'] == char:
                    filtered_groups.append(group['cells'])
        return filtered_groups

    def draw_l_line(self, cols_lists, font_color='B8CCE4', background_color='B8CCE4'):
        for colA in cols_lists:
            for colB in colA:
                left = None if self.move_cell_str(colB, 'left') in colA else Side(style='thin')
                right = None if self.move_cell_str(colB, 'right') in colA else Side(style='thin')
                top = None if self.move_cell_str(colB, 'up') in colA else Side(style='thin')
                bottom = None if self.move_cell_str(colB, 'down') in colA else Side(style='thin')
                self.ws[colB].border = Border(left=left, right=right, top=top, bottom=bottom)
                if background_color is not None:
                    self.ws[colB].fill = PatternFill(start_color=background_color,
                                                     end_color=background_color, fill_type="solid")
                if top is None:
                    self.ws[colB].value = ''

    def shift_columns(self, cell_positions):
        column_map = {'B': 'C', 'C': 'D', 'D': 'E'}
        new_cell_positions = []
        for cell in cell_positions:
            match = re.match(r'^([A-Z]+)(\d+)$', cell)
            if match:
                col, row = match.groups()
                new_col = column_map.get(col, col)
                new_cell_positions.append(f"{new_col}{row}")
            else:
                new_cell_positions.append(cell)
        return new_cell_positions

    def get_field_details(self, row):
        details = {}
        if row[4] == 'DROP_DOWN' and len(row) > 10:
            try:
                options_str = row[10]
                items = options_str.split(',')
                options = []
                for item in items:
                    if ': {' in item:
                        option = item.split(': {')[0].strip()
                        if option not in ['options', 'index', 'defaultValue'] and not option.startswith('"'):
                            options.append(option)
                if options:
                    details['BC'] = '選択肢: ' + ', '.join(options)
            except Exception as e:
                print(f"選択肢の解析エラー: {e}")
        return details

    # ★ save メソッドを追加 ★
    def save(self):
        """Excelファイルを保存"""
        self.wb.save(self.filename)
        print(f"Excelファイル '{self.filename}' が作成されました。")


import json
import os

class PropertyInfo:
    def __init__(self, key, code, is_subtable=False, subtable_key=None):
        self.key = key
        self.code = code
        self.is_subtable = is_subtable
        self.subtable_key = subtable_key

    def __repr__(self):
        return f"PropertyInfo(key='{self.key}', code='{self.code}', is_subtable={self.is_subtable}, subtable_key='{self.subtable_key}')"


# ─── PropertyFieldMapper クラス ─────────────────────────────────────────────

class PropertyFieldMapper:
    def __init__(self, properties: dict):
        self.key_to_info = {}
        self.code_to_info = {}
        self._parse_properties(properties)

    @classmethod
    def from_json_file(cls, path: str):
        """
        JSONファイルパスからPropertyFieldMapperを作成。
        ファイルが存在しない・形式不正の場合は例外を投げる。
        """
        if not os.path.isfile(path):
            raise FileNotFoundError(f"ファイルが存在しません: {path}")

        try:
            with open(path, 'r', encoding='utf-8') as f:
                data = json.load(f)
        except json.JSONDecodeError as e:
            raise ValueError(f"JSONの読み込みに失敗しました: {e}")

        if "properties" not in data or not isinstance(data["properties"], dict):
            raise ValueError("JSONの形式が不正です: 'properties' フィールドが見つかりません")

        return cls(data["properties"])


    def export_debug_info(self, filename: str):
        """
        全フィールド情報を指定ファイル名でCSV出力（実行ディレクトリ）。
        出力項目：display_key, display_code, is_subtable, subtable_key
        """
        try:
            with open(filename, 'w', newline='', encoding='utf-8') as f:
                writer = csv.writer(f)
                writer.writerow(["display_key", "display_code", "is_subtable", "subtable_key"])
                for info in self.code_to_info.values():
                    display_key = self.get_display_key_by_code(info.code)
                    display_code = self.get_display_code_by_code(info.code)
                    writer.writerow([
                        info.code,
                        display_key,
                        display_code,
                        str(info.is_subtable),
                        info.subtable_key or ""
                    ])
            print(f"[OK] フィールド情報を '{filename}' に出力しました。")
        except Exception as e:
            print(f"[ERROR] 出力失敗: {e}")
            
    def _parse_properties(self, properties: dict):
        for key, value in properties.items():
            code = value.get("code")
            prop_type = value.get("type")

            if prop_type == "SUBTABLE":
                fields = value.get("fields", {})
                for sub_key, sub_value in fields.items():
                    sub_code = sub_value.get("code")
                    info = PropertyInfo(
                        key=sub_key,
                        code=sub_code,
                        is_subtable=True,
                        subtable_key=key
                    )
                    self.key_to_info[sub_key] = info
                    self.code_to_info[sub_code] = info

                table_info = PropertyInfo(
                    key=key,
                    code=code,
                    is_subtable=False,
                    subtable_key=None
                )
                self.key_to_info[key] = table_info
                self.code_to_info[code] = table_info
            else:
                info = PropertyInfo(
                    key=key,
                    code=code,
                    is_subtable=False,
                    subtable_key=None
                )
                self.key_to_info[key] = info
                self.code_to_info[code] = info

    def get_by_key(self, key):
        return self.key_to_info.get(key)

    def get_by_code(self, code):
        if code not in self.code_to_info:
            return code
        return self.code_to_info.get(code)

    def get_display_key_by_code(self, code: str) -> Union[str, None]:
        try:
            info = self.get_by_code(code)
            if not info:
                return None
            if isinstance(info, str):
                return info
            if info.is_subtable:
                subtable_info = self.get_by_key(info.subtable_key)
                return f"{subtable_info.key}[{info.key}]"
            return info.key
        except Exception as e:
            import traceback
            error_msg = f"{str(e)}\n{traceback.format_exc()}"
            print(f"エラーが発生しました: {error_msg}")  # デバッグ用の出力
            return error_msg

    def get_display_code_by_code(self, code: str) -> Union[str, None]:
        info = self.get_by_code(code)
        if not info:
            return None
        if info.is_subtable:
            subtable_info = self.get_by_key(info.subtable_key)
            return f"{subtable_info.code}[{info.code}]"
        return info.code

# ─── KintoneApp クラス ─────────────────────────────────────────────
class KintoneApp:
    def __init__(self, appid, api_token=None, subdomain=None, username=None, password=None, config_path='config_UserAccount.yaml'):
        self.appid = appid
        config = self.load_config(config_path)
        self.subdomain = subdomain or config.get('subdomain')
        self.username = username or config.get('username')
        self.password = password or config.get('password')
        self.api_token = api_token or config.get('api_token')
        if not all([self.subdomain, self.username, self.password]):
            print("Error: 認証情報が不足しています。コマンドライン引数または設定ファイルで指定してください。")
            sys.exit(1)
        self.app_name = self.get_app_name_by_settings()
        self.base_dir, self.js_dir, self.json_dir = self.create_directory_structure()

    def load_config(self, config_path):
        try:
            with open(config_path, 'r', encoding='utf-8') as f:
                return yaml.safe_load(f)
        except FileNotFoundError:
            return {}

    @staticmethod
    def convert_to_utf8_if_sjis(content):
        try:
            return content.decode('utf-8')
        except UnicodeDecodeError:
            return content.decode('shift_jis').encode('utf-8').decode('utf-8')

    def fetch_data(self, url, headers):
        try:
            # レコード通知設定の場合、POSTメソッドとリクエストボディが必要
            if "perRecord.json" in url:
                data = {"app": self.appid}
                response = requests.get(url, headers=headers, json=data)
            else:
                response = requests.get(url, headers=headers)
            response.raise_for_status()
            content = self.convert_to_utf8_if_sjis(response.content)
            return json.loads(content)
        except requests.exceptions.RequestException as e:
            print(f"Error fetching data from {url}: {e}")
            sys.exit(1)

    @staticmethod
    def sanitize_app_name(app_name):
        return re.sub(r'[\\/:*?"<>|]+', '', app_name)

    def get_app_name_by_settings(self):
        url = f"https://{self.subdomain}.cybozu.com/k/v1/app/settings.json?app={self.appid}"
        headers = {"X-Cybozu-API-Token": self.api_token}
        print(f'url: {url}   headers: {headers}')
        data = self.fetch_data(url, headers)
        raw_app_name = data.get("name", "")
        return self.sanitize_app_name(raw_app_name)

    def create_directory_structure(self):
        output_dir = Path('./output')
        backup_dir = Path('./backup')
        output_dir.mkdir(exist_ok=True)
        backup_dir.mkdir(exist_ok=True)
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        base_dir = output_dir / f'{self.appid}_{self.app_name}_{timestamp}'
        js_dir = base_dir / 'javascript'
        json_dir = base_dir / 'json'
        base_dir.mkdir(parents=True, exist_ok=True)
        js_dir.mkdir(parents=True, exist_ok=True)
        json_dir.mkdir(parents=True, exist_ok=True)
        print(f"新しいディレクトリ構造を作成しました: {base_dir}")
        return base_dir, js_dir, json_dir

    def save_json_file(self, data, filename):
        file_path = self.json_dir / f"{self.appid}_{filename}.json"
        with open(file_path, 'w', encoding='utf-8') as f:
            json.dump(data, f, ensure_ascii=False, indent=4)
        return file_path

    def save_yaml_file(self, data, filename):
        file_path = self.base_dir / f"{self.appid}_{filename}.yaml"
        with open(file_path, 'w', encoding='utf-8') as f:
            yaml.dump(data, f, allow_unicode=True)
        return file_path

    def download_file(self, file_key, file_name, js_info):
        url = f"https://{self.subdomain}.cybozu.com/k/v1/file.json?fileKey={file_key}"
        headers = {"X-Cybozu-API-Token": self.api_token}
        try:
            response = requests.get(url, headers=headers, stream=True, allow_redirects=True)
            response.raise_for_status()
            safe_filename = file_name
            file_path = self.js_dir / safe_filename
            with open(file_path, 'wb') as f:
                for chunk in response.iter_content(chunk_size=8192):
                    f.write(chunk)
            js_info.append({'file_id': file_key, 'file_name': safe_filename, 'type': 'file'})
        except requests.exceptions.RequestException as e:
            print(f"Error downloading file {file_name}: {e}")
            sys.exit(1)

    def download_url_content(self, url, js_info):
        safe_filename = f"{self.appid}_url_" + re.sub(r'[\\/*?:"<>|]', '_', url) + ".js"
        file_path = self.js_dir / safe_filename
        try:
            subprocess.run(['curl', '-L', '-o', str(file_path), url], check=True)
            js_info.append({'url': url, 'file_name': safe_filename, 'type': 'url'})
        except subprocess.CalledProcessError as e:
            print(f"Error downloading URL content {url} with curl: {e}")
            sys.exit(1)

    def get_customize_info(self):
        url = f"https://{self.subdomain}.cybozu.com/k/v1/app/customize.json?app={self.appid}"
        auth_string = f"{self.username}:{self.password}"
        encoded_auth = base64.b64encode(auth_string.encode()).decode()
        headers = {"X-Cybozu-Authorization": encoded_auth}
        try:
            response = requests.get(url, headers=headers)
            response.raise_for_status()
            content = self.convert_to_utf8_if_sjis(response.content)
            return json.loads(content)
        except requests.exceptions.RequestException as e:
            print(f"Error fetching customize info: {e}")
            return {"desktop": {"js": []}}

    def download_app_data(self):
        endpoints = {
            "form": {"url": f"https://{self.subdomain}.cybozu.com/k/v1/form.json?app={self.appid}", "auth_type": "X-Cybozu-API-Token"},
            "record_acl": {"url": f"https://{self.subdomain}.cybozu.com/k/v1/record/acl.json?app={self.appid}", "auth_type": "X-Cybozu-API-Token"},
            "field_acl": {"url": f"https://{self.subdomain}.cybozu.com/k/v1/field/acl.json?app={self.appid}", "auth_type": "X-Cybozu-API-Token"},
            "form_fields": {"url": f"https://{self.subdomain}.cybozu.com/k/v1/app/form/fields.json?app={self.appid}", "auth_type": "X-Cybozu-API-Token"},
            "form_layout": {"url": f"https://{self.subdomain}.cybozu.com/k/v1/app/form/layout.json?app={self.appid}", "auth_type": "X-Cybozu-API-Token"},
            "views": {"url": f"https://{self.subdomain}.cybozu.com/k/v1/app/views.json?app={self.appid}", "auth_type": "X-Cybozu-API-Token"},
            "settings": {"url": f"https://{self.subdomain}.cybozu.com/k/v1/app/settings.json?app={self.appid}", "auth_type": "X-Cybozu-API-Token"},
            "process_management": {"url": f"https://{self.subdomain}.cybozu.com/k/v1/app/status.json?app={self.appid}", "auth_type": "X-Cybozu-API-Token"},
            "plugins": {"url": f"https://{self.subdomain}.cybozu.com/k/v1/app/plugins.json?app={self.appid}", "auth_type": "X-Cybozu-API-Token"},
            "app_notifications": {"url": f"https://{self.subdomain}.cybozu.com/k/v1/app/notifications/general.json?app={self.appid}", "auth_type": "X-Cybozu-API-Token"},
            "record_notifications": {"url": f"https://{self.subdomain}.cybozu.com/k/v1/app/notifications/perRecord.json", "auth_type": "X-Cybozu-API-Token"},
            "reminder_notifications": {"url": f"https://{self.subdomain}.cybozu.com/k/v1/app/notifications/reminder.json?app={self.appid}", "auth_type": "X-Cybozu-API-Token"},
            "app_acl": {"url": f"https://{self.subdomain}.cybozu.com/k/v1/app/acl.json?app={self.appid}", "auth_type": "X-Cybozu-API-Token"},
            "actions": {"url": f"https://{self.subdomain}.cybozu.com/k/v1/app/actions.json?app={self.appid}", "auth_type": "X-Cybozu-API-Token"},
            "graphs": {"url": f"https://{self.subdomain}.cybozu.com/k/v1/app/reports.json?app={self.appid}", "auth_type": "X-Cybozu-API-Token"},
            "general_notifications": {"url": f"https://{self.subdomain}.cybozu.com/k/v1/app/notifications/general.json?app={self.appid}", "auth_type": "X-Cybozu-API-Token"},
        }
        js_info = []
        for name, endpoint in endpoints.items():
            url = endpoint["url"]
            auth_type = endpoint["auth_type"]
            headers = {auth_type: self.api_token if auth_type == "X-Cybozu-API-Token" else None}
            data = self.fetch_data(url, headers)
            self.save_json_file(data, name)
            self.save_yaml_file(data, name)
        customize_data = self.get_customize_info()
        self.save_json_file(customize_data, "customize")
        self.save_yaml_file(customize_data, "customize")
        files = customize_data.get('desktop', {}).get('js', [])
        for file_info in files:
            if file_info.get('type') == 'URL':
                self.download_url_content(file_info['url'], js_info)
            else:
                file_data = file_info.get('file', {})
                if file_data.get('fileKey') and file_data.get('name'):
                    self.download_file(file_data['fileKey'], file_data['name'], js_info)
        self.save_json_file(js_info, "javascript_info")
        self.save_yaml_file(js_info, "javascript_info")

    def process_layout_and_fields(self):
        layout_file = self.json_dir / f"{self.appid}_form_layout.json"
        fields_file = self.json_dir / f"{self.appid}_form_fields.json"
        output_file = self.base_dir / f"{self.appid}_layout_raw.tsv"
        if layout_file.exists() and fields_file.exists():
            process_file(layout_file, fields_file, output_file)
            print(f"レイアウト情報を {output_file} に出力しました。")
        else:
            print(f"必要なファイルが見つかりません: {layout_file} または {fields_file}")

    def process_layout_to_structured(self):
        input_file = self.base_dir / f"{self.appid}_layout_raw.tsv"
        output_file = self.base_dir / f"{self.appid}_layout_structured.tsv"
        process_raw_layout(input_file, output_file)
        print(f"構造化されたレイアウト情報を {output_file} に出力しました。")

    # Excelレポート作成処理をサブメソッドに分割
    def create_excel_report(self):
        tsv_filename = self.base_dir / f"{self.appid}_layout_structured.tsv"
        excel_filename = self.base_dir / f"{self.appid}_layout_report.xlsx"
        workbook = Workbook()
        worksheet = workbook.active
        formatter = ExcelFormatter(workbook=workbook, worksheet=worksheet, filename=excel_filename)
        self._setup_excel_format(formatter)
        self._write_excel_headers(formatter)
        self._apply_group_formatting(formatter)
        self._write_js_field_code_usage(formatter)
        formatter.save()  # save メソッドを呼び出して Excel ファイルを保存
        print(f"Excelレポートを作成しました: {excel_filename}")

    def _setup_excel_format(self, formatter):
        formatter.set_row_height(200, 20)
        formatter.set_column_width(1, 26*5, 22)
        ws = formatter.ws
        ws.column_dimensions['BA'].width = 25
        ws.column_dimensions['BB'].width = 25
        ws.column_dimensions['BC'].width = 30
        ws.column_dimensions['BD'].width = 25
        ws.column_dimensions['BE'].width = 50
        ws.column_dimensions['BF'].width = 50
        white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type="solid")
        for row in range(1, 201):
            for col in range(1, 54):
                cell = ws.cell(row=row, column=col)
                cell.fill = white_fill

    def _write_excel_headers(self, formatter):
        formatter.merge_cells_and_set_content('D2', 'R2', '項目名', alignment="center", bottom_border=True)
        formatter.merge_cells_and_set_content('S2', 'T2', '必須', alignment="center", bottom_border=True, right_border=True)
        formatter.merge_cells_and_set_content('U2', 'V2', 'JS', alignment="center", bottom_border=True, right_border=True)
        formatter.merge_cells_and_set_content('W2', 'X2', 'plugin', alignment="center", bottom_border=True, right_border=True)
        formatter.merge_cells_and_set_content('Y2', 'AO2', '備考', alignment="center", bottom_border=True, right_border=True)
        formatter.merge_cells_and_set_content('BA2', 'BA2', 'フィールドコード', alignment="center", bottom_border=True, right_border=True)
        formatter.merge_cells_and_set_content('BB2', 'BB2', 'フィールド種別', alignment="center", bottom_border=True, right_border=True)
        formatter.merge_cells_and_set_content('BC2', 'BC2', 'ドロップダウン選択肢', alignment="center", bottom_border=True, right_border=True)
        formatter.merge_cells_and_set_content('BD2', 'BD2', 'JS使用箇所', alignment="center", bottom_border=True, right_border=True)
        formatter.merge_cells_and_set_content('BE2', 'BE2', '行データ（全体）', alignment="center", bottom_border=True, right_border=True)
        formatter.merge_cells_and_set_content('BF2', 'BF2', 'JSON文字列', alignment="center", bottom_border=True, right_border=True)

    def _apply_group_formatting(self, formatter):
        formatter.get_column_group_arrays()
        g_groups = formatter.get_groups_by_first_char('G')
        s_groups = formatter.get_groups_by_first_char('S')
        if g_groups:
            formatter.draw_l_line(g_groups, background_color='E6F0F9')
        if s_groups:
            formatter.draw_l_line(s_groups, background_color='D4E4F4')


    def _write_js_field_code_usage(self, formatter):
        js_dir = self.base_dir / 'javascript'
        # まず、.js_kaigyo.jsファイルを準備
        prepare_kaigyo_files(js_dir)
        field_codes_by_js_line_map = scan_directory_for_field_codes_with_lines(js_dir)
        field_codes_yaml_path = self.base_dir / f"{self.appid}_field_codes_usage_at_javascript.yaml"
        with open(field_codes_yaml_path, 'w', encoding='utf-8') as f:
            yaml.dump( field_codes_by_js_line_map, f, allow_unicode=True, sort_keys=False)
        print(f"フィールドコードのjs内での使用行番号情報を {field_codes_yaml_path} に保存しました。")
        formatter.set_by_out02_tsv(self.base_dir / f"{self.appid}_layout_structured.tsv")
        ws = formatter.ws
        for row in range(3, ws.max_row + 1):
            field_code_cell = ws.cell(row=row, column=column_index_from_string('BA'))
            field_code = field_code_cell.value
            if field_code and field_code in  field_codes_by_js_line_map:
                usage_info =  field_codes_by_js_line_map[field_code]
                usage_text = ""
                for js_file, line_numbers in usage_info.items():
                    usage_text += f"{js_file}: {', '.join(map(str, line_numbers))}\n"
                bd_cell = ws.cell(row=row, column=column_index_from_string('BD'))
                bd_cell.value = usage_text.strip()
                bd_cell.font = formatter.font
       
        # JSファイル別にシートを作成して内容を表示
        self._create_js_code_sheets(formatter.wb,  field_codes_by_js_line_map)

    def _create_js_code_sheets(self, workbook, field_codes_by_js_line_map):
        """各JSファイルの内容を別シートに表示し、フィールドコードの使用箇所を強調表示する"""
        js_dir = self.base_dir / 'javascript'

        # PropertyFieldMapperを使用してフィールドコードとその表示名の対応を取得
        try:
            property_mapper = PropertyFieldMapper.from_json_file(self.json_dir / f"{self.appid}_form_fields.json")
        except Exception as e:
            print(f"フィールド情報の読み込みエラー: {e}")
            return

        property_mapper.export_debug_info(self.base_dir / f"{self.appid}_field_info_debug.csv")

        # 背景色の設定
        light_blue_fill = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")
        light_green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
        dark_green_fill = PatternFill(start_color="C6E0B4", end_color="C6E0B4", fill_type="solid")

        # 各JSファイルに対してシートを作成
        for js_file in js_dir.glob('*.js'):
            try:
                # ファイルの内容を読み込む
                with open(js_file, 'r', encoding='utf-8', errors='replace') as f:
                    lines = f.readlines()

                # シート名はファイル名の右端から31文字以内に設定
                sheet_name = js_file.name.replace('._kaigyo_.js', '.js')[-31:]

                # シートが既に存在する場合は削除
                if sheet_name in workbook.sheetnames:
                    ws = workbook[sheet_name]
                    workbook.remove(ws)

                # 新しいシートを作成
                ws = workbook.create_sheet(sheet_name)

                # ヘッダー行の設定
                ws['A1'] = 'ファイル名:'
                ws['B1'] = js_file.name.replace('._kaigyo_.js', '.js')
                if '._kaigyo_.' in js_file.name:
                    ws.merge_cells('C1:D1')
                    ws['C1'] = f'※ 1行が1000文字を超えている為、適宜改行した {js_file.name} を使用しています。'
                    ws['C1'].fill = PatternFill(start_color="FFFF0000", end_color="FFFF0000", fill_type="solid")
                    ws['C1'].font = Font(color="FFFFFF", bold=True)
                    ws['C1'].alignment = Alignment(vertical='center')

                # A1, B1に淡い水色の背景色を設定
                ws['A1'].fill = light_blue_fill
                ws['B1'].fill = light_blue_fill

                # テーブルヘッダーの設定
                ws['A3'] = '行番号'
                ws['A3'].alignment = Alignment(horizontal='center', vertical='center')
                ws['B3'] = 'フィールド名'
                ws['B3'].alignment = Alignment(horizontal='center', vertical='center')
                ws['C3'] = 'フィールドコード'
                ws['C3'].alignment = Alignment(horizontal='center', vertical='center')
                ws['D3'] = 'ソースコード'
                ws['D3'].alignment = Alignment(horizontal='center', vertical='center')

                # A3, B3, C3, D3に淡い緑色の背景色を設定
                ws['A3'].fill = light_green_fill
                ws['B3'].fill = light_green_fill
                ws['C3'].fill = light_green_fill
                ws['D3'].fill = dark_green_fill

                # 列幅の設定
                ws.column_dimensions['A'].width = 10
                ws.column_dimensions['B'].width = 34
                ws.column_dimensions['C'].width = 34
                ws.column_dimensions['D'].width = 140

                # 使用されているフィールドコードとその行番号を特定
                field_usage = {}
                for field_code, usage_info in field_codes_by_js_line_map.items():
                    if js_file.name in usage_info:
                        line_numbers = usage_info[js_file.name]
                        for line_num in line_numbers:
                            if line_num <= len(lines):
                                if line_num not in field_usage:
                                    field_usage[line_num] = []
                                try:
                                    field_name = property_mapper.get_display_key_by_code(field_code)    
                                except Exception as e:
                                    field_name = "ERROR"
                                field_usage[line_num].append((field_name, field_code))

                # コードをセルに表示（500行を超える場合は対象行のみ表示）
                if len(lines) > 500:
                    # フィールドコードを含む行とその前後10行を特定
                    target_lines = set()
                    for field_code, usage_info in field_codes_by_js_line_map.items():
                        if js_file.name in usage_info:
                            for line_num in usage_info[js_file.name]:
                                # 前後10行を含める
                                for i in range(max(1, line_num - 10), min(len(lines) + 1, line_num + 11)):
                                    target_lines.add(i)

                    # ソートして順序を保持
                    target_lines = sorted(target_lines)
                else:
                    target_lines = range(1, len(lines) + 1)

                # 対象行を表示
                for i, line_num in enumerate(target_lines, 1):
                    row_num = i + 4  # 5行目から開始
                    ws[f'A{row_num}'] = line_num

                    if line_num in field_usage:
                        field_names = []
                        field_codes = []
                        for name, code in field_usage[line_num]:
                            field_names.append(name)
                            field_codes.append(code)

                        ws[f'B{row_num}'] = '\n'.join(field_names)
                        ws[f'C{row_num}'] = '\n'.join(field_codes)

                    ws[f'D{row_num}'] = lines[line_num-1].rstrip('\n\r')
                    for col in ['A', 'B', 'C', 'D']:
                        if ws[f'{col}{row_num}'].value is not None:
                            ws[f'{col}{row_num}'].font = Font(name='メイリオ', size=9)
                        ws[f'B{row_num}'].alignment = Alignment(wrap_text=True, vertical='top')
                        ws[f'C{row_num}'].alignment = Alignment(wrap_text=True, vertical='top')
                        ws[f'D{row_num}'].alignment = Alignment(wrap_text=False, horizontal='center', vertical='center')

                print(f"JSファイル {js_file.name} のシートを作成しました。")
            except Exception as e:
                print(f"シート {sheet_name} の作成中にエラーが発生しました: {e}")

    def export_all_records(self, get_all=False):
        url = f"https://{self.subdomain}.cybozu.com/k/v1/records.json"
        headers = {"X-Cybozu-API-Token": self.api_token}
        all_records = []
        offset = 0
        limit = 100
        max_records = float('inf') if get_all else 500
        while True:
            params = {"app": self.appid, "query": f"limit {limit} offset {offset}"}
            try:
                response = requests.get(url, headers=headers, params=params)
                response.raise_for_status()
                data = response.json()
                records = data.get("records", [])
                if not records:
                    break
                all_records.extend(records)
                if len(all_records) >= max_records:
                    all_records = all_records[:max_records]
                    break
                offset += limit
            except requests.exceptions.RequestException as e:
                print(f"Error fetching records: {e}")
                sys.exit(1)
        if all_records:
            self._export_records_json(all_records)
            self._export_records_tsv_excel(all_records)
        else:
            print("エクスポートするレコードが見つかりませんでした。")

    def _export_records_json(self, all_records):
        json_file = self.base_dir / f"{self.appid}_records.json"
        try:
            with open(json_file, "w", encoding="utf-8") as f_json:
                json.dump(all_records, f_json, ensure_ascii=False, indent=4)
            print(f"全レコードをJSON形式で {json_file} にエクスポートしました。")
        except IOError as e:
            print(f"JSONファイルの保存中にエラーが発生しました: {e}")
            sys.exit(1)

    def _export_records_tsv_excel(self, all_records):
        flattened_records = [flatten_record(record) for record in all_records]
        field_names = sorted({key for record in flattened_records for key in record.keys()})
        field_max_lengths = {field: max(len(str(record.get(field, ""))) for record in flattened_records) for field in field_names}
        normal_fields = [field for field in field_names if field_max_lengths[field] < 50]
        long_fields = sorted([field for field in field_names if field_max_lengths[field] >= 50], key=lambda x: field_max_lengths[x])
        new_field_order = normal_fields + long_fields

        def remove_img_tag(value):
            pattern = r'<img\s+src=["\']?data:image/png[^>]*>'
            return re.sub(pattern, '', value)

        tsv_file = self.base_dir / f"{self.appid}_records.tsv"
        try:
            with open(tsv_file, "w", encoding="utf-8", newline="") as f_tsv:
                writer = csv.DictWriter(f_tsv, fieldnames=new_field_order, delimiter="\t")
                writer.writeheader()
                for record in flattened_records:
                    row = {field: remove_img_tag(str(record.get(field, ""))) if '<img src=data:image/png' in str(record.get(field, "")) else str(record.get(field, "")) for field in new_field_order}
                    writer.writerow(row)
            print(f"全レコードをTSV形式で {tsv_file} にエクスポートしました。")
            self._export_records_excel(tsv_file)
        except IOError as e:
            print(f"ファイルの保存中にエラーが発生しました: {e}")
            sys.exit(1)

    def _export_records_excel(self, tsv_file):
        excel_file = self.base_dir / f"{self.appid}_records.xlsx"
        wb = Workbook()
        ws = wb.active
        with open(tsv_file, 'r', encoding='utf-8') as f:
            tsv_reader = csv.reader(f, delimiter='\t')
            header = next(tsv_reader)
            for col, value in enumerate(header, 1):
                cell = ws.cell(row=1, column=col, value=value)
                cell.fill = PatternFill(start_color='B8CCE4', end_color='B8CCE4', fill_type='solid')
                cell.alignment = Alignment(horizontal='center', vertical='center')
                cell.font = Font(bold=True)
            for row_idx, row in enumerate(tsv_reader, 2):
                for col_idx, value in enumerate(row, 1):
                    cell_value = ' '.join(value.replace('\n', ' ').replace('\r', ' ').replace('\t', ' ').split())
                    ws.cell(row=row_idx, column=col_idx, value=cell_value).number_format = '@'
        for column in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            ws.column_dimensions[column_letter].width = min(max_length + 2, 50)
        wb.save(excel_file)
        print(f"全レコードをExcel形式で {excel_file} にエクスポートしました。")

    def run(self):
        self.download_app_data()
        self.process_layout_and_fields()
        self.process_layout_to_structured()
        self.create_excel_report()
        self.export_all_records()

# ─── エントリーポイント ─────────────────────────────────────────────
if __name__ == "__main__":
    if len(sys.argv) == 2:
        appid = sys.argv[1]
        app = KintoneApp(appid)
        app.run()
    elif len(sys.argv) == 6:
        appid = sys.argv[1]
        api_token = sys.argv[2]
        subdomain = sys.argv[3]
        username = sys.argv[4]
        password = sys.argv[5]
        app = KintoneApp(appid, api_token, subdomain, username, password)
        app.run()
    else:
        print("Usage: python script.py <appid> [<api_token> <subdomain> <username> <password>]")
        print("Note: 認証情報は config_UserAccount.yaml からも読み込めます")
        sys.exit(1)
