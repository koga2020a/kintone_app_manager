#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
kintoneアプリのプロセス管理ワークフローをExcelに出力するスクリプト
"""

import os
import sys
import yaml
import json
import argparse
import logging
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font

# 定数定義
SCRIPT_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = SCRIPT_DIR.parent / "output"

def setup_logging():
    """ロギングの設定"""
    log_dir = SCRIPT_DIR.parent / "logs"
    log_dir.mkdir(exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = log_dir / f"process_workflow_{timestamp}.log"
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler()
        ]
    )
    
    return logging.getLogger("process_workflow")

def create_workflow_graph(process_states):
    """プロセス管理の状態遷移図を作成"""
    G = nx.DiGraph()
    
    # ノード（状態）の追加
    for state_name, state_info in process_states.items():
        G.add_node(state_name)
    
    # エッジ（遷移）の追加
    for state_name, state_info in process_states.items():
        for action_name, action_info in state_info.get("actions", {}).items():
            next_state = action_info.get("nextState", "")
            if next_state:
                G.add_edge(state_name, next_state, action=action_name)
    
    return G

def find_all_paths(states, actions):
    from collections import defaultdict
    action_map = defaultdict(list)
    for action in actions:
        action_map[action['from']].append(action)

    all_states = set(states.keys())
    to_states = set(a['to'] for a in actions)
    terminal_states = all_states - set(action_map.keys())

    paths = []
    def dfs(current, path, visited):
        if current in terminal_states or not action_map[current]:
            paths.append(path[:])
            return
        for action in action_map[current]:
            # (状態, アクション名, 遷移先) のタプルで循環検出
            visit_key = (current, action['name'], action['to'])
            if visit_key in visited:
                continue
            visited.add(visit_key)
            path.append((action['name'], action['to']))
            dfs(action['to'], path, visited)
            path.pop()
            visited.remove(visit_key)

    for state_name, state_info in states.items():
        if state_info.get('index') == '0':
            dfs(state_name, [(None, state_name)], set())
    return paths

def create_workflow_excel(app_id, process_data, output_file=None):
    """ワークフロー情報をExcelに出力"""
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = OUTPUT_DIR / f"{app_id}_workflow_{timestamp}.xlsx"
    
    # Excelワークブックを作成
    wb = Workbook()
    
    # デフォルトのSheet1を削除
    if "Sheet" in wb.sheetnames:
        wb.remove(wb["Sheet"])
    
    # スタイル定義
    header_font = Font(bold=True, size=11, name='Arial')
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    green_fill = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
    green_light_fill = PatternFill(start_color="EBFAEB", end_color="EBFAEB", fill_type="solid")
    blue_fill = PatternFill(start_color="DEEBF7", end_color="DEEBF7", fill_type="solid")  
    blue_light_fill = PatternFill(start_color="F2F9FF", end_color="F2F9FF", fill_type="solid")
    
    # 1. 基本情報シート
    ws_basic = wb.create_sheet(title="基本情報")
    
    # ヘッダー行の設定
    headers = ["項目", "値"]
    for col, header in enumerate(headers, 1):
        cell = ws_basic.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 基本情報の書き込み
    basic_info = [
        ("有効化", "はい" if process_data.get('enable') else "いいえ"),
        ("リビジョン", process_data.get('revision', ''))
    ]
    
    for row, (label, value) in enumerate(basic_info, 2):
        ws_basic.cell(row=row, column=1, value=label)
        ws_basic.cell(row=row, column=2, value=value)
        for col in range(1, 3):
            cell = ws_basic.cell(row=row, column=col)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = thin_border
    
    # 2. 状態一覧シート
    ws_states = wb.create_sheet(title="状態一覧")
    
    # ヘッダー行の設定
    headers = ["状態名", "インデックス", "作業者タイプ", "作業者"]
    for col, header in enumerate(headers, 1):
        cell = ws_states.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # 状態一覧の書き込み
    row = 2
    for state_name, state_info in process_data.get('states', {}).items():
        ws_states.cell(row=row, column=1, value=state_name)
        ws_states.cell(row=row, column=2, value=state_info.get('index', ''))
        assignee = state_info.get('assignee', {})
        ws_states.cell(row=row, column=3, value=assignee.get('type', ''))
        
        # 作業者の処理を修正
        entities = assignee.get('entities', [])
        entity_info = []
        for entity in entities:
            if isinstance(entity, dict) and 'entity' in entity:
                entity_type = entity['entity'].get('type', '')
                entity_code = entity['entity'].get('code', '')
                include_subs = entity.get('includeSubs', False)
                entity_info.append(f"{entity_type}: {entity_code}" + (" (サブ組織を含む)" if include_subs else ""))
        
        ws_states.cell(row=row, column=4, value='\n'.join(entity_info))
        
        for col in range(1, 5):
            cell = ws_states.cell(row=row, column=col)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = thin_border
        
        row += 1
    
    # 3. アクション一覧シート
    ws_actions = wb.create_sheet(title="アクション一覧")
    
    # ヘッダー行の設定
    headers = ["アクション名", "遷移元", "遷移先", "条件"]
    for col, header in enumerate(headers, 1):
        cell = ws_actions.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # アクション一覧の書き込み
    row = 2
    for action in process_data.get('actions', []):
        ws_actions.cell(row=row, column=1, value=action.get('name', ''))
        ws_actions.cell(row=row, column=2, value=action.get('from', ''))
        ws_actions.cell(row=row, column=3, value=action.get('to', ''))
        ws_actions.cell(row=row, column=4, value=action.get('filterCond', ''))
        
        for col in range(1, 5):
            cell = ws_actions.cell(row=row, column=col)
            cell.alignment = Alignment(vertical='center', wrap_text=True)
            cell.border = thin_border
        
        row += 1
    
    # 4. ワークフロー遷移マトリクスシート
    ws_matrix = wb.create_sheet(title="ワークフロー遷移マトリクス")
    
    # ステータス名リスト（indexの昇順）
    states_dict = process_data.get('states', {})
    status_names = sorted(states_dict.keys(), key=lambda k: int(states_dict[k].get('index', 0)))

    """
    凡例：遷移マトリクスの構造
    ------------------------------------------------------------
    |          |  作業者  | 状態1 | 状態2 | 状態3 | ...
    |----------|------------|--------|--------|--------|--------
    | 状態1    | 担当者     | アクション情報 | ... |
    | 状態2    | 担当者     |        |      |
    | 状態3    | 担当者     |        |      |
    
    1. 作業者の形式：
       - FIELD_ENTITY: 作成者
       - USER: ユーザー名@domain
       - GROUP: グループ名
       - サブ組織を含む場合は「(サブ組織を含む)」が付加
    
    2. アクション情報の形式：
       →アクション名→
       【条件】
       条件文
       
       例：
       →処理開始→
       【条件】
       レコード番号 = "100"
    """
    
    # ヘッダー
    cell1 = ws_matrix.cell(row=1, column=1, value="ステータス名")
    cell1.alignment = Alignment(horizontal='center', vertical='center')
    cell1.font = Font(bold=True)
    ws_matrix.row_dimensions[1].height = 30
    
    cell2 = ws_matrix.cell(row=1, column=2, value="作業者") 
    cell2.alignment = Alignment(horizontal='center', vertical='center')
    cell2.font = Font(bold=True)
    for col, status in enumerate(status_names, 1):
        cell = ws_matrix.cell(row=1, column=col+2, value=status)
        cell.fill = PatternFill(start_color="006400", end_color="006400", fill_type="solid")
        cell.alignment = header_alignment
        cell.border = thin_border
        cell.font = Font(color="FFFFFF", bold=True)
    ws_matrix.cell(row=1, column=1).fill = header_fill
    ws_matrix.cell(row=1, column=2).fill = header_fill
    for row, status in enumerate(status_names, 1):
        # ステータス名
        cell = ws_matrix.cell(row=row+1, column=1, value=status)
        cell.font = header_font
        cell.fill = green_fill
        cell.alignment = header_alignment
        cell.border = thin_border
        # 作業者
        assignee = states_dict[status].get('assignee', {})
        entities = assignee.get('entities', [])
        entity_info = []
        for entity in entities:
            if isinstance(entity, dict) and 'entity' in entity:
                entity_type = entity['entity'].get('type', '')
                entity_code = entity['entity'].get('code', '')
                include_subs = entity.get('includeSubs', False)
                entity_info.append(f"{entity_type}: {entity_code}" + (" (サブ組織を含む)" if include_subs else ""))
        ws_matrix.cell(row=row+1, column=2, value='\n'.join(entity_info))
        ws_matrix.cell(row=row+1, column=2).alignment = Alignment(vertical='center', wrap_text=True)
        ws_matrix.cell(row=row+1, column=2).border = thin_border
        ws_matrix.cell(row=row+1, column=2).fill = green_fill
    # 各アクションをマトリクスに記載（→方向のみ）
    for action in process_data.get('actions', []):
        from_idx = status_names.index(action['from']) + 2
        to_idx = status_names.index(action['to']) + 2
        # →方向のみ
        cell = ws_matrix.cell(row=from_idx, column=to_idx+1)
        val = cell.value or ''
        cond = action.get('filterCond', '')
        cell.value = (
            val + ("\n" if val else "") +
            f"→{action['name']} ↑\n【条件】\n{cond}"
        )
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border
        cell.fill = green_light_fill
    # 列幅を指定
    ws_matrix.column_dimensions['A'].width = 28.57  # 200px
    ws_matrix.column_dimensions['B'].width = 42.86  # 300px
    for i in range(3, 3 + len(status_names)):
        col_letter = ws_matrix.cell(row=1, column=i).column_letter
        ws_matrix.column_dimensions[col_letter].width = 28.57  # 200px
    
    # 最大列数と最大行数を取得
    max_col = ws_matrix.max_column
    max_row = ws_matrix.max_row
    
    # 
    # 空白行を挿入するため、下から上に処理
    for row in range(max_row, 0, -1):
        # 各行の後に3行の空白行を挿入
        ws_matrix.insert_rows(row + 1, 1)
        # 元の行のスタイルを設定
        for col in range(1, max_col + 1):
            cell = ws_matrix.cell(row=row, column=col)
            # 2行目から奇数行の背景色を青色系に変更
            if row >= 2 and row % 2 == 0 and cell.value:
                if col <= 2:  # A列とB列
                    cell.fill = blue_fill
                else:  # C列以降
                    cell.fill = blue_light_fill

    black_fill = PatternFill(start_color="000000", end_color="000000", fill_type="solid")
    max_col = ws_matrix.max_column
    max_row = ws_matrix.max_row
    for row in range(max_row, 0, -1):
        # 1行おきに高さを設定
        if row % 2 == 0:
            ws_matrix.row_dimensions[row].height = 6
            for col in range(1, max_col + 1):
                cell = ws_matrix.cell(row=row, column=col)
                cell.fill = black_fill
    
    
    # ファイルを保存
    wb.save(output_file)
    return output_file

def main():
    """メイン関数"""
    # コマンドライン引数の解析
    parser = argparse.ArgumentParser(description='kintoneアプリのプロセス管理ワークフローをExcelに出力するスクリプト')
    parser.add_argument('app_id', help='アプリID')
    parser.add_argument('api_token', help='APIトークン', nargs='?')
    parser.add_argument('subdomain', help='サブドメイン', nargs='?')
    parser.add_argument('username', help='ユーザー名', nargs='?')
    parser.add_argument('password', help='パスワード', nargs='?')
    parser.add_argument('--output', help='出力ファイル名（省略時は自動生成）')
    args = parser.parse_args()
    
    # ロギングの設定
    logger = setup_logging()
    logger.info("プロセスワークフローのExcel出力を開始します")
    
    # デバッグ情報：OUTPUT_DIRの内容を表示
    logger.info(f"OUTPUT_DIRのパス: {OUTPUT_DIR.resolve()}")
    logger.info("OUTPUT_DIR内のディレクトリ一覧:")
    for item in OUTPUT_DIR.iterdir():
        if item.is_dir():
            logger.info(f"  {item.name}")
    
    # アプリIDに対応するディレクトリを探す
    app_dir = None
    for item in OUTPUT_DIR.iterdir():
        if item.is_dir() and item.name.startswith(f"{args.app_id}_"):
            app_dir = item
            break
    
    if not app_dir:
        logger.error(f"アプリID {args.app_id} に対応するディレクトリが見つかりません")
        sys.exit(1)
    
    # デバッグ情報：アプリディレクトリの内容を表示
    logger.info(f"アプリディレクトリのパス: {app_dir.resolve()}")
    logger.info("アプリディレクトリ内のファイル一覧:")
    for item in app_dir.iterdir():
        logger.info(f"  {item.name}")
    
    # プロセス管理のJSONファイルを探す
    process_file = None
    possible_patterns = [
        f"{args.app_id}_process_management.json",
        f"{args.app_id}_process.json",
        f"{args.app_id}_workflow.json",
        "process_management.json",
        "process.json",
        "workflow.json"
    ]
    
    # まず、jsonディレクトリ内を検索
    json_dir = app_dir / "json"
    if json_dir.exists():
        logger.info(f"jsonディレクトリのパス: {json_dir.resolve()}")
        logger.info("jsonディレクトリ内のファイル一覧:")
        for item in json_dir.iterdir():
            logger.info(f"  {item.name}")
        
        for pattern in possible_patterns:
            logger.info(f"パターン '{pattern}' で検索中...")
            for file in json_dir.glob(pattern):
                process_file = file
                logger.info(f"ファイルを見つけました: {file}")
                break
            if process_file:
                break
    
    # jsonディレクトリ内で見つからない場合、アプリディレクトリ直下を検索
    if not process_file:
        logger.info("jsonディレクトリ内でファイルが見つからないため、アプリディレクトリ直下を検索します")
        for pattern in possible_patterns:
            logger.info(f"パターン '{pattern}' で検索中...")
            for file in app_dir.glob(pattern):
                process_file = file
                logger.info(f"ファイルを見つけました: {file}")
                break
            if process_file:
                break
    
    if not process_file:
        logger.error(f"アプリID {args.app_id} のプロセス管理ファイルが見つかりません")
        logger.error("以下のパターンで検索しました:")
        for pattern in possible_patterns:
            logger.error(f"- {pattern}")
        sys.exit(1)
    
    # JSONファイルを読み込む
    try:
        with open(process_file, 'r', encoding='utf-8') as f:
            process_data = json.load(f)
    except Exception as e:
        logger.error(f"プロセス管理ファイルの読み込み中にエラーが発生しました: {e}")
        sys.exit(1)
    
    # 出力ファイル名の設定
    output_file = None
    if args.output:
        output_file = Path(args.output)
        if not output_file.is_absolute():
            output_file = app_dir / output_file
    
    # Excelファイルの生成
    try:
        result_file = create_workflow_excel(args.app_id, process_data, output_file)
        logger.info(f"プロセスワークフローを {result_file} に出力しました")
        print(result_file)
    except Exception as e:
        logger.error(f"Excelファイルの生成中にエラーが発生しました: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main() 