import os
import csv
import argparse
import logging
import re
import pandas as pd
from typing import List, Optional
from pathlib import Path
import yaml  # 追加

class ACLProblemCollector:
    """
    数字で始まるヘッダー名を持つディレクトリからACL問題を収集・集約するクラス
    
    対象となるディレクトリの例：
    - 10_【検証】km送信履歴_20250112_204810/  -> 対象
    - 11_【検証】案件Step2_20250112_204821/  -> 対象
    - test_【検証】ACL_TEST/ -> 対象外（数字で始まらない）
    """
    
    HEADERS = ["ヘッダ名", "メッセージ", "タイプ", "名称", "矛盾タイプ", "出現回数", "詳細", "ディレクトリ名"]
    NUMERIC_PATTERN = re.compile(r'^\d+_.+$')  # 数字で始まり、アンダースコアが含まれる形式
    
    def __init__(self, output_dir: str = "./output", output_csv: str = "all_acl_problem_report.csv"):
        """
        Args:
            output_dir (str): 探索する基準ディレクトリ（デフォルト: ./output）
            output_csv (str): 集約結果を保存するCSVファイルのパス
        """
        self.output_dir = Path(output_dir)
        self.output_csv = Path(output_csv)
        self._setup_logging()

    def _setup_logging(self) -> None:
        """ロギング設定を初期化"""
        logging.basicConfig(
            level=logging.INFO,
            format='%(asctime)s - %(levelname)s - %(message)s',
            handlers=[
                logging.StreamHandler()
            ]
        )

    def is_valid_header_dir(self, dir_name: str) -> bool:
        """
        ディレクトリ名が有効なヘッダー名（数字で始まり、アンダースコアが含まれる）かどうかを判定
        
        Args:
            dir_name (str): 判定するディレクトリ名
            
        Returns:
            bool: 数字で始まり、アンダースコアが含まれる場合はTrue
        """
        return bool(self.NUMERIC_PATTERN.match(dir_name))

    def find_header_dirs(self) -> List[Path]:
        """
        数字で始まり、アンダースコアが含まれるヘッダー名を持つディレクトリを探索
        
        Returns:
            List[Path]: 有効なヘッダーディレクトリのパスのリスト
        """
        try:
            return [
                d for d in self.output_dir.iterdir()
                if d.is_dir() and self.is_valid_header_dir(d.name)
            ]
        except Exception as e:
            logging.error(f"ディレクトリ探索中にエラーが発生: {e}")
            return []

    def read_acl_problem_file(self, file_path: Path) -> List[List[str]]:
        """
        ACL問題ファイルを読み込み
        
        Args:
            file_path (Path): 読み込むファイルのパス
            
        Returns:
            List[List[str]]: 読み込んだデータのリスト
        """
        try:
            with file_path.open("r", encoding="utf-8") as f:
                reader = csv.reader(f, delimiter="\t")
                headers = next(reader)  # ヘッダーをスキップ
                
                data = []
                for row in reader:
                    if not row:  # 空行をスキップ
                        continue
                        
                    # 必要な列数まで'-'で埋める
                    padded_row = row + ["-"] * (len(self.HEADERS) - 1 - len(row))
                    
                    # 問題がある場合（タイプ列に値がある場合）
                    if len(padded_row) >= 3 and padded_row[1] not in ["-", "問題未検出"]:
                        # 元のデータを保持
                        original_data = padded_row[1:]  # 2列目以降のデータを全て保持
                        
                        # メッセージを設定
                        padded_row[1] = "問題あり"
                        
                        # 2列目以降のデータを1列ずつずらして設定
                        for i in range(len(original_data)):
                            if i + 2 < len(padded_row):  # インデックスの範囲チェック
                                padded_row[i + 2] = original_data[i]
                    
                    data.append(padded_row)
                return data
                
        except Exception as e:
            logging.error(f"ファイル読み込み中にエラーが発生 {file_path}: {e}")
            return []

    def collect_problems(self) -> Optional[List[List[str]]]:
        """
        数字で始まり、アンダースコアが含まれるヘッダー名を持つディレクトリから問題データを収集
        
        Returns:
            Optional[List[List[str]]]: 収集したデータのリスト。エラー時はNone
        """
        collected_data = []
        header_dirs = self.find_header_dirs()
        
        if not header_dirs:
            logging.warning(f"数字で始まり、アンダースコアが含まれるヘッダー名を持つディレクトリが見つかりません: {self.output_dir}")
            return None

        for header_dir in header_dirs:
            # ディレクトリ名から最初の数字部分を抽出
            header_num_match = re.match(r'^(\d+)_', header_dir.name)
            if not header_num_match:
                logging.warning(f"ディレクトリ名からヘッダ番号を抽出できません: {header_dir.name}")
                continue
            header_num = header_num_match.group(1)
            
            problem_file = header_dir / f"{header_num}_acl_problem.csv"
            
            if problem_file.exists():
                problem_data = self.read_acl_problem_file(problem_file)
                if problem_data:
                    for row in problem_data:
                        if len(row) < len(self.HEADERS) - 1:  # ディレクトリ名列を除いた長さでチェック
                            # 必要なカラム数に満たない場合、デフォルト値で補完
                            row += ["-"] * (len(self.HEADERS) - 1 - len(row))
                        # ディレクトリ名を追加
                        row.append(header_dir.name)
                        collected_data.append(row)
                    logging.info(f"読み込み完了: {problem_file}")
                else:
                    # ファイルが空またはヘッダーのみの場合
                    collected_data.append([header_num, "No problems detected", "-", "-", "-", "-", "-", header_dir.name])
                    logging.info(f"問題なし（データなし）: {header_num}")
            else:
                collected_data.append([header_num, "No problems detected", "-", "-", "-", "-", "-", header_dir.name])
                logging.info(f"問題なし（ファイル未存在）: {header_num}")

        return collected_data

    def collect_user_acl_settings(self, header_dir: Path) -> List[List[str]]:
        """
        指定されたディレクトリから個人設定のACL情報を収集
        
        Args:
            header_dir (Path): ヘッダーディレクトリのパス
            
        Returns:
            List[List[str]]: [ヘッダー番号, ユーザーコード, ACL種別, 権限設定, ディレクトリ名] の形式のリスト
        """
        user_settings = []
        header_num = header_dir.name.split('_')[0]
        
        yaml_files = {
            'app': f"{header_num}_app_acl.yaml",
            'record': f"{header_num}_record_acl.yaml",
            'field': f"{header_num}_field_acl.yaml"
        }
        
        for acl_type, filename in yaml_files.items():
            file_path = header_dir / filename
            if not file_path.exists():
                continue
                
            try:
                with file_path.open('r', encoding='utf-8') as f:
                    data = yaml.safe_load(f)
                    
                if not data or 'rights' not in data:
                    continue
                    
                # app_aclの場合
                if acl_type == 'app':
                    for right in data['rights']:
                        if 'entity' in right and right['entity'].get('type') == 'USER':
                            permission = "編集可" if right.get('appEditable') else "閲覧のみ"
                            user_settings.append([
                                header_num,
                                right['entity']['code'],
                                'アプリ',
                                permission,
                                header_dir.name
                            ])
                
                # record_aclの場合
                elif acl_type == 'record':
                    for right in data['rights']:
                        if 'entities' in right:
                            for entity in right['entities']:
                                if 'entity' in entity and entity['entity'].get('type') == 'USER':
                                    permissions = []
                                    if entity.get('viewable'): permissions.append('閲覧')
                                    if entity.get('editable'): permissions.append('編集')
                                    if entity.get('deletable'): permissions.append('削除')
                                    permission = '・'.join(permissions) if permissions else '権限なし'
                                    user_settings.append([
                                        header_num,
                                        entity['entity']['code'],
                                        'レコード',
                                        permission,
                                        header_dir.name
                                    ])
                
                # field_aclの場合
                elif acl_type == 'field':
                    for right in data['rights']:
                        if 'entities' in right:
                            for entity in right['entities']:
                                if 'entity' in entity and entity['entity'].get('type') == 'USER':
                                    permission = entity.get('accessibility', 'NONE')
                                    permission_map = {
                                        'READ': '閲覧のみ',
                                        'WRITE': '編集可',
                                        'NONE': '権限なし'
                                    }
                                    permission_str = permission_map.get(permission, permission)
                                    user_settings.append([
                                        header_num,
                                        entity['entity']['code'],
                                        'フィールド',
                                        permission_str,
                                        header_dir.name
                                    ])
                        
            except Exception as e:
                logging.error(f"YAMLファイル読み込み中にエラーが発生 {file_path}: {e}")
                
        return user_settings

    def save_results(self, collected_data: List[List[str]]) -> bool:
        """
        収集結果をCSVとExcelに保存
        """
        try:
            # ACL問題のDataFrame作成
            df = pd.DataFrame(collected_data, columns=self.HEADERS)
            
            # ヘッダ名を数値として扱うために一時的な列を作成
            df['sort_key'] = df['ヘッダ名'].str.extract('(\d+)').astype(int)
            
            # 問題ありとなしで分割してソート
            no_problems_df = df[df['メッセージ'] == 'No problems detected'].copy()
            problems_df = df[df['メッセージ'] != 'No problems detected'].copy()
            no_problems_df['メッセージ'] = '問題未検出'
            
            # 数値でソート
            no_problems_df = no_problems_df.sort_values('sort_key', ascending=True)
            problems_df = problems_df.sort_values('sort_key', ascending=True)
            
            # 一時的なソート用の列を削除
            no_problems_df = no_problems_df.drop('sort_key', axis=1)
            problems_df = problems_df.drop('sort_key', axis=1)
            
            # 結合
            result_df = pd.concat([problems_df, no_problems_df])
            
            # 個人設定の収集
            user_settings = []
            for header_dir in self.find_header_dirs():
                user_settings.extend(self.collect_user_acl_settings(header_dir))
            
            # 個人設定のDataFrame作成（同様に数値でソート）
            user_settings_df = pd.DataFrame(
                user_settings,
                columns=['ヘッダー番号', 'ユーザーコード', 'ACL種別', '権限設定', 'ディレクトリ名']
            )
            user_settings_df['sort_key'] = user_settings_df['ヘッダー番号'].astype(int)
            user_settings_df = user_settings_df.sort_values(['sort_key', 'ACL種別', 'ユーザーコード'])
            user_settings_df = user_settings_df.drop('sort_key', axis=1)
            
            # CSVとして保存
            result_df.to_csv(self.output_csv, index=False, encoding='utf-8', sep='\t')
            logging.info(f"集約結果をCSVに保存: {self.output_csv}")
            
            # Excelファイルとして保存（2つのシート）
            excel_path = self.output_csv.with_suffix('.xlsx')
            with pd.ExcelWriter(excel_path, engine='openpyxl') as writer:
                # ヘッダ名を数値として扱うために一時的な列を作成
                result_df['sort_key'] = result_df['ヘッダ名'].str.extract('(\d+)').astype(int)
                
                # 問題ありとなしで分割してソート
                monndai_mikenshutu_text = '問題未検出'
                no_problems_df = result_df[result_df['メッセージ'] == monndai_mikenshutu_text].copy()
                problems_df = result_df[result_df['メッセージ'] != monndai_mikenshutu_text].copy()
                
                # 数値でソート
                no_problems_df = no_problems_df.sort_values('sort_key', ascending=True)
                problems_df = problems_df.sort_values('sort_key', ascending=True)
                
                # 一時的なソート用の列を削除
                no_problems_df = no_problems_df.drop('sort_key', axis=1)
                problems_df = problems_df.drop('sort_key', axis=1)
                
                # 結合（問題ありが上、問題なしが下）
                result_df = pd.concat([problems_df, no_problems_df])
                
                result_df.to_excel(writer, sheet_name='acl問題一覧', index=False)
                user_settings_df.to_excel(writer, sheet_name='個人名設定一覧', index=False)
                
                # スタイルの設定
                from openpyxl.styles import Alignment, PatternFill
                center_align = Alignment(horizontal='center')
                header_fill = PatternFill(start_color='E6F3FF', end_color='E6F3FF', fill_type='solid')
                
                # acl問題一覧シートの設定
                worksheet = writer.sheets['acl問題一覧']
                
                # 列幅の設定（acl問題一覧シート）
                column_widths = [72, 88, 70, 250, 210, 78, 100, 410]
                for i, width in enumerate(column_widths):
                    excel_width = width / 7
                    worksheet.column_dimensions[chr(65 + i)].width = excel_width
                
                # ヘッダー行の背景色設定（acl問題一覧シート）
                for cell in worksheet[1]:
                    cell.fill = header_fill
                
                # A列(1), C列(3), F列(6)を中央寄せに設定（acl問題一覧シート）
                center_columns = [1, 3, 6]
                for col in center_columns:
                    for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=col, max_col=col):
                        for cell in row:
                            cell.alignment = center_align
                
                # 個人名設定シートの設定
                user_settings_df = user_settings_df.sort_values('ヘッダー番号', key=lambda x: x.astype(int))
                user_settings_df.to_excel(writer, sheet_name='個人名設定一覧', index=False)
                user_settings_worksheet = writer.sheets['個人名設定一覧']
                
                # 列幅の設定（個人名設定シート）
                user_settings_widths = [105, 220, 76, 112, 300]
                for i, width in enumerate(user_settings_widths):
                    excel_width = width / 7
                    user_settings_worksheet.column_dimensions[chr(65 + i)].width = excel_width
                
                # ヘッダー行の背景色設定（個人名設定シート）
                for cell in user_settings_worksheet[1]:
                    cell.fill = header_fill
                
                # A列を中央寄せに設定（個人名設定シート）
                for row in user_settings_worksheet.iter_rows(min_row=1, max_row=user_settings_worksheet.max_row, min_col=1, max_col=1):
                    for cell in row:
                        cell.alignment = center_align

            logging.info(f"集約結果をExcelに保存: {excel_path}")
            
            return True
            
        except Exception as e:
            logging.error(f"結果保存中にエラーが発生: {e}")
            return False

    def run(self) -> bool:
        """
        メイン実行処理
        
        Returns:
            bool: 処理成功時はTrue、失敗時はFalse
        """
        if not self.output_dir.exists():
            logging.error(f"指定されたディレクトリが存在しません: {self.output_dir}")
            return False

        collected_data = self.collect_problems()
        if collected_data is None:
            return False

        return self.save_results(collected_data)

def main():
    parser = argparse.ArgumentParser(
        description="""
数字で始まり、アンダースコアが含まれるヘッダー名を持つディレクトリからACL問題を集約するスクリプト

対象となるディレクトリ構造の例：
  baseディレクトリ/
    123_/
      123_acl_problem.csv  -> 処理対象
    456_/
      456_acl_problem.csv  -> 処理対象
    abc_/
      abc_acl_problem.csv  -> 処理対象外
    123abc_/
      123abc_acl_problem.csv  -> 処理対象外
""",
        formatter_class=argparse.RawDescriptionHelpFormatter
    )
    parser.add_argument(
        "-d", "--directory",
        type=str,
        default="./output",
        help="探索する基準ディレクトリ (デフォルト: ./output)"
    )
    parser.add_argument(
        "-o", "--output",
        type=str,
        default="all_acl_problem_report.csv",
        help="集約結果を保存するCSVファイル名 (デフォルト: all_acl_problem_report.csv)"
    )
    parser.add_argument(
        "-v", "--verbose",
        action="store_true",
        help="詳細なログ出力を有効にする"
    )

    args = parser.parse_args()

    # 詳細ログ出力の設定
    if args.verbose:
        logging.getLogger().setLevel(logging.DEBUG)

    collector = ACLProblemCollector(args.directory, args.output)
    success = collector.run()
    
    if not success:
        logging.error("処理が失敗しました")
        return 1
    return 0

if __name__ == "__main__":
    exit(main())
