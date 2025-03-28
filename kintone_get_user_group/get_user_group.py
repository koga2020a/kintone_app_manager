import argparse
import sys
import base64
from getpass import getpass
from typing import List, Dict, Any

import requests

import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Alignment, PatternFill, Font, Border, Side
from openpyxl.utils import column_index_from_string, get_column_letter
import logging
import yaml
import glob
import zipfile
import tempfile
import os
import random
import colorsys

class ArgumentParser:
  @staticmethod
  def parse_arguments():
    parser = argparse.ArgumentParser(
      description='Kintoneの全ユーザーと各ユーザーの所属グループをExcelに出力します。\n\n引数を省略した場合、config_UserAccount.yaml を参照して認証情報を取得し、デフォルトの出力ファイル名を使用します。',
      formatter_class=argparse.RawTextHelpFormatter
    )
    parser.add_argument('--subdomain', help='Kintoneのサブドメイン (例: sample)')
    parser.add_argument('--username', help='管理者ユーザーのログイン名 (例: user@example.com)')
    parser.add_argument('--password', help='管理者ユーザーのパスワード (指定しない場合、プロンプトで入力)')
    parser.add_argument('--output', default='kintone_users_groups.xlsx', help='出力するExcelファイルの名前 (デフォルト: kintone_users_groups.xlsx)')
    parser.add_argument('--silent', action='store_true', help='サイレントモードを有効にします。詳細なログを表示しません。')
    
    return parser.parse_args()

class KintoneClient:
  def __init__(self, subdomain: str, username: str, password: str, logger: logging.Logger):
    self.subdomain = subdomain
    self.headers = self._get_auth_header(username, password)
    self.logger = logger

  @staticmethod
  def _get_auth_header(username: str, password: str) -> Dict[str, str]:
    credentials = f"{username}:{password}"
    base64_credentials = base64.b64encode(credentials.encode('utf-8')).decode('utf-8')
    return {
      'X-Cybozu-Authorization': base64_credentials
    }

  def _fetch_data(self, endpoint: str, params: Dict[str, Any], key: str) -> List[Dict[str, Any]]:
    url = f"https://{self.subdomain}.cybozu.com/v1/{endpoint}.json"
    data = []
    size = 100
    offset = 0

    while True:
      current_params = params.copy()
      current_params.update({'size': size, 'offset': offset})
      response = requests.get(url, headers=self.headers, params=current_params)
      if response.status_code != 200:
        self.logger.error(f"{endpoint.capitalize()}の取得に失敗しました: {response.status_code} {response.text}")
        sys.exit(1)
      batch = response.json().get(key, [])
      if not batch:
        break
      data.extend(batch)
      if len(batch) < size:
        break
      offset += size
      self.logger.debug(f"Fetched {len(batch)} items from {endpoint} (offset: {offset})")
    self.logger.info(f"全{endpoint}を取得しました。総数: {len(data)}")
    return data

  def get_all_users(self) -> List[Dict[str, Any]]:
    return self._fetch_data('users', {}, 'users')

  def get_all_groups(self) -> List[Dict[str, Any]]:
    return self._fetch_data('groups', {}, 'groups')

  def get_users_in_group(self, group_code: str) -> List[Dict[str, Any]]:
    params = {'code': group_code}
    return self._fetch_data('group/users', params, 'users')

class DataProcessor:
  def __init__(self, users: List[Dict[str, Any]], groups: List[Dict[str, Any]], client: KintoneClient, logger: logging.Logger):
    self.users = users
    self.groups = groups
    self.client = client
    self.user_id_to_info: Dict[str, Dict[str, Any]] = {}
    self.logger = logger

  def filter_groups(self) -> List[Dict[str, Any]]:
    filtered = [group for group in self.groups if group.get('name') != 'Everyone' and group.get('code')]
    self.logger.info(f"「Everyone」を除外し、codeが存在するグループ数: {len(filtered)}")
    return filtered

  def organize_groups(self, filtered_groups: List[Dict[str, Any]]) -> List[str]:
    group_names = [group['name'] for group in filtered_groups]
    self.logger.info(f"グループ名一覧: {group_names}")

    if 'Administrators' in group_names:
      group_names.remove('Administrators')
      group_names.insert(0, 'Administrators')
      self.logger.info("「Administrators」グループをグループ名一覧の最初に配置しました。")
    else:
      self.logger.info("「Administrators」グループがグループ名一覧に存在しませんでした。")
    self.logger.info(f"最終的なグループ名一覧: {group_names}")
    return group_names

  def map_users(self):
    for user in self.users:
      user_id = str(user.get('id'))
      self.user_id_to_info[user_id] = {
        'ユーザーID': user_id,
        'ステータス': '停止中' if not user.get('valid', True) else '',
        'ログイン名': user.get('code'),
        '氏名': user.get('name'),
        'メールアドレス': user.get('email'),
        '所属グループ一覧': [],
        '最終アクセス日': '',
        '経過日数': ''
      }
    self.logger.info(f"ユーザー情報をマッピングしました。総ユーザー数: {len(self.user_id_to_info)}")

  def populate_group_memberships(self, filtered_groups: List[Dict[str, Any]]):
    self.logger.info("各グループの所属ユーザーを取得中...")
    for group in filtered_groups:
      group_code = group.get('code')
      group_name = group.get('name')
      self.logger.info(f"グループ '{group_name}' ({group_code}) のユーザーを取得中...")
      users_in_group = self.client.get_users_in_group(group_code)
      self.logger.info(f"グループ '{group_name}' に所属するユーザー数: {len(users_in_group)}")
      for user in users_in_group:
        user_id = str(user.get('id'))
        if user_id in self.user_id_to_info:
          self.user_id_to_info[user_id]['所属グループ一覧'].append(group_name)
          self.logger.debug(f"ユーザーID {user_id} はグループ '{group_name}' に所属しています。")
        else:
          # ユーザーが全ユーザーリストに存在しない場合（稀）
          self.user_id_to_info[user_id] = {
            'ユーザーID': user_id,
            'ステータス': '',
            'ログイン名': user.get('code'),
            '氏名': user.get('name'),
            'メールアドレス': user.get('email'),
            '所属グループ一覧': [group_name],
            '最終アクセス日': '',
            '経過日数': ''
          }
          self.logger.debug(f"ユーザーID {user_id} はグループ '{group_name}' に所属しています（新規追加）。")
    self.logger.info("グループの所属ユーザー情報を更新しました。")

  def generate_dataframes(self, group_names: List[str]) -> Dict[str, pd.DataFrame]:
    self.logger.info("データフレームを作成中...")
    user_data_active = []
    user_data_stopped = []

    for user_info in self.user_id_to_info.values():
      login_name = user_info['ログイン名'] or ''  # Noneの場合は空文字に
      email = user_info['メールアドレス'] or ''   # Noneの場合は空文字に
      
      # 条件に基づいて「相違」列を設定
      if login_name and email:  # 両方とも値が存在する場合のみ比較
        if login_name != email:
          if login_name.lower() == email.lower():
            discrepancy = "大小相違"
          else:
            discrepancy = "相違"
        else:
          discrepancy = ""
      else:
        discrepancy = ""  # どちらかが空の場合は相違なしとする
      
      user_info['相違'] = discrepancy
      user_info['所属グループ一覧'] = ', '.join(user_info['所属グループ一覧'])
      if user_info['ステータス'] == '停止中':
        user_data_stopped.append(user_info)
      else:
        user_data_active.append(user_info)

    df_active = pd.DataFrame(user_data_active)
    df_stopped = pd.DataFrame(user_data_stopped)

    # グループごとの「●」をマークする列を追加
    for group in group_names:
      df_active[group] = df_active['所属グループ一覧'].apply(
        lambda x: '●' if group in [g.strip() for g in x.split(',')] else ''
      )
      df_stopped[group] = df_stopped['所属グループ一覧'].apply(
        lambda x: '●' if group in [g.strip() for g in x.split(',')] else ''
      )

    # 列の順序を設定（「相違」列をB列に挿入し、GとHを初期から含める）
    basic_columns = ['ユーザーID', '相違', 'ステータス', 'ログイン名', '氏名', 'メールアドレス', '最終アクセス日', '経過日数', '所属グループ一覧']
    group_columns = group_names
    columns_order = basic_columns + group_columns

    df_active = df_active[columns_order].sort_values(by=['所属グループ一覧'], ascending=False)
    df_stopped = df_stopped[columns_order].sort_values(by=['所属グループ一覧'], ascending=False)

    self.logger.info("データフレームの作成が完了しました。")
    return {'アクティブ': df_active, '停止中': df_stopped}

  def export_group_user_list(self, filtered_groups: List[Dict[str, Any]]):
    """グループとユーザーの関連をYAMLファイルとして出力"""
    self.logger.info("group_user_list.yaml と group_user_list_NoUse.yaml を生成中...")
    
    active_group_data = {}
    inactive_group_data = {}
    
    # まず全グループの基本情報を設定
    for group in filtered_groups:
      group_code = group.get('code')
      active_group_data[group_code] = {
        'name': group.get('name'),
        'users': []
      }
      inactive_group_data[group_code] = {
        'name': group.get('name'),
        'users': []
      }
      
      # グループ内のユーザー情報を取得
      users_in_group = self.client.get_users_in_group(group_code)
      for user in users_in_group:
        user_info = {
          'username': user.get('code'),
          'email': user.get('email'),
          'id': str(user.get('id'))
        }
        # ユーザーの状態を確認
        if user.get('valid', True):
          active_group_data[group_code]['users'].append(user_info)
        else:
          inactive_group_data[group_code]['users'].append(user_info)
    
    # Everyoneグループを追加
    active_everyone_users = []
    inactive_everyone_users = []
    for user in self.users:
      user_info = {
        'username': user.get('code'),
        'email': user.get('email'),
        'id': str(user.get('id'))
      }
      if user.get('valid', True):
        active_everyone_users.append(user_info)
      else:
        inactive_everyone_users.append(user_info)
    
    active_group_data['everyone'] = {
      'name': 'Everyone',
      'users': active_everyone_users
    }
    
    inactive_group_data['everyone'] = {
      'name': 'Everyone',
      'users': inactive_everyone_users
    }
    
    # 空のグループを削除
    active_group_data = {k: v for k, v in active_group_data.items() if v['users']}
    inactive_group_data = {k: v for k, v in inactive_group_data.items() if v['users']}
    
    # YAMLファイルに出力
    try:
      # アクティブユーザー用のファイル
      with open('group_user_list.yaml', 'w', encoding='utf-8') as f:
        yaml.dump(active_group_data, f, allow_unicode=True, sort_keys=False)
      
      # 停止中ユーザー用のファイル
      with open('group_user_list_NoUse.yaml', 'w', encoding='utf-8') as f:
        yaml.dump(inactive_group_data, f, allow_unicode=True, sort_keys=False)
      
      self.logger.info("group_user_list.yaml と group_user_list_NoUse.yaml の生成が完了しました。")
    except Exception as e:
      self.logger.error(f"YAMLファイルの生成中にエラーが発生しました: {e}")

class ExcelExporter:
  def __init__(self, dataframes: Dict[str, pd.DataFrame], group_names: List[str], output_file: str, logger: logging.Logger):
    self.dataframes = dataframes
    self.group_names = group_names
    self.output_file = output_file
    self.logger = logger
    self.group_data = {}  # グループ情報を保持する辞書を追加
    self.domain_list = []  # 全ユーザーのドメイン一覧を収集
    self.user_groups = {}  # ユーザーIDと所属グループ一覧のマッピングを追加

  def prepare_group_data(self, client: KintoneClient):
    """グループごとのユーザー情報を準備"""
    self.logger.info("グループ情報シート用のデータを準備中...")
    
    # ユーザーIDと所属グループ一覧のマッピングを作成
    for df in self.dataframes.values():
        for _, row in df.iterrows():
            user_id = row['ユーザーID']
            groups = row['所属グループ一覧']
            self.user_groups[user_id] = groups

    # グループ情報シートに出現する全メールアドレスからドメイン一覧を収集
    all_domains = set()
    for group in self.group_names:
        for df in self.dataframes.values():
            mask = df[group] == '●'
            users = df[mask]
            # メールアドレスからドメイン部分を抽出して一覧に追加
            domains = users['メールアドレス'].dropna().apply(lambda x: x.split('@')[-1] if '@' in x else '').unique()
            all_domains.update([d for d in domains if d])
    
    # kirin.co.jpを先頭に、残りをアルファベット順にソート
    ordered_domains = ['kirin.co.jp'] if 'kirin.co.jp' in all_domains else []
    other_domains = sorted([d for d in all_domains if d != 'kirin.co.jp'])
    ordered_domains.extend(other_domains)
    
    self.logger.info(f"検出されたドメイン一覧: {ordered_domains}")
    self.domain_list = ordered_domains
    
    for group in self.group_names:
        group_users = []
        for df in self.dataframes.values():
            # グループに所属するユーザーを抽出
            mask = df[group] == '●'
            users = df[mask][['ユーザーID', 'ログイン名', '氏名', 'メールアドレス', 'ステータス']].copy()
            
            if not users.empty:
                # 所属グループ一覧を追加
                users['所属グループ一覧'] = users['ユーザーID'].map(self.user_groups)
                
                # メールアドレスを分解
                users['domain'] = users['メールアドレス'].str.split('@').str[1]
                users['localpart'] = users['メールアドレス'].str.split('@').str[0]
                
                # kirin.co.jpとその他でデータを分割
                kirin_users = users[users['domain'] == 'kirin.co.jp'].copy()
                other_users = users[users['domain'] != 'kirin.co.jp'].copy()
                
                # それぞれをソート
                kirin_users = kirin_users.sort_values('localpart')
                other_users = other_users.sort_values(['domain', 'localpart'])
                
                # 結合（kirinが上、その他が下）
                users = pd.concat([kirin_users, other_users], ignore_index=True)
                
                # 一時的なソート用列を削除
                users = users.drop(['domain', 'localpart'], axis=1)
                
                # 停止中のユーザーに「●」を追加
                users['停止中'] = users['ステータス'].apply(lambda x: '●' if x == '停止中' else '')
                group_users.append(users)
        
        if group_users:
            self.group_data[group] = pd.concat(group_users, ignore_index=True)
        else:
            self.group_data[group] = pd.DataFrame(columns=['ユーザーID', 'ログイン名', '氏名', 'メールアドレス', '停止中', '所属グループ一覧'])

  def export_to_excel(self):
    self.logger.info("Excelファイルに出力中...")
    
    with pd.ExcelWriter(self.output_file, engine='openpyxl') as writer:
      # 既存シート（アクティブ、停止中）の出力
      for sheet_name, df in self.dataframes.items():
        df.to_excel(writer, sheet_name=sheet_name, index=False)
      
      # グループ情報シートを新規作成
      if self.group_data:
        sheet_name = 'グループ情報'
        # シートを追加
        workbook = writer.book
        ws = workbook.create_sheet(title=sheet_name)
        
        # ドメイン情報を最初に追加
        cell1 = ws.cell(row=1, column=1, value="ドメイン一覧")
        cell1.alignment = Alignment(horizontal='center')
        cell2 = ws.cell(row=2, column=1, value="ドメイン") 
        cell2.alignment = Alignment(horizontal='center')
        cell3 = ws.cell(row=2, column=2, value="背景色")
        cell3.alignment = Alignment(horizontal='center')
        
        # ドメインとその色情報を設定
        domain_colors = {}
        start_row = 3
        for i, domain in enumerate(self.domain_list):
            cell = ws.cell(row=start_row+i, column=1, value='@'+domain)
            cell.alignment = Alignment(horizontal='right')
            row = start_row+i
            domain_colors[domain] = row
        
        # グループデータの出力開始位置を設定
        start_row = start_row + len(self.domain_list) + 2  # ドメイン一覧の後に2行空ける
        
        for group_name, df in self.group_data.items():
          # --- 1. グループ名行 ---
          ws.cell(row=start_row, column=1, value="グループ: " + group_name)
          start_row += 1
          
          # --- 2. ヘッダー行 ---
          headers = ["ユーザーID", "ログイン名", "氏名", "メールアドレス", "停止中", "所属グループ一覧"]
          for col, header in enumerate(headers, 1):
            ws.cell(row=start_row, column=col, value=header)
          start_row += 1
          
          # --- 3. データ行 ---
          if not df.empty:
            for r_idx, row in df.iterrows():
              ws.cell(row=start_row, column=1, value=row['ユーザーID'])
              ws.cell(row=start_row, column=2, value=row['ログイン名'])
              ws.cell(row=start_row, column=3, value=row['氏名'])
              ws.cell(row=start_row, column=4, value=row['メールアドレス'])
              ws.cell(row=start_row, column=5, value=row['停止中'])
              ws.cell(row=start_row, column=6, value=row['所属グループ一覧'])
              start_row += 1
          else:
            # データがない場合は空行を出力
            ws.cell(row=start_row, column=1, value="(データなし)")
            start_row += 1
          
          # --- 4. セット間に空行を追加 ---
          start_row += 1
          ws.row_dimensions[start_row-1].height = 30
        
        writer.sheets[sheet_name] = ws
    
    self.logger.info(f"Excelファイル '{self.output_file}' を作成しました。")

  def format_excel(self):
    self.logger.info("Excelファイルのフォーマットを設定中...")
    
    # 定数として列名を定義
    COLUMN_USER_ID = 'A'       # ユーザーID
    COLUMN_DISCREPANCY = 'B'    # 相違
    COLUMN_STATUS = 'C'         # ステータス
    COLUMN_LOGIN_NAME = 'D'     # ログイン名
    COLUMN_NAME = 'E'           # 氏名
    COLUMN_EMAIL = 'F'          # メールアドレス
    COLUMN_LAST_ACCESS = 'G'    # 最終アクセス日
    COLUMN_DAYS_SINCE = 'H'     # 経過日数
    COLUMN_GROUPS = 'I'         # 所属グループ一覧

    # 単位変換：px → 文字数（openpyxlでは幅は文字数）
    def px_to_char(px):
        return px / 7

    # ヘッダー行の基本スタイル
    header_fill = PatternFill(start_color='243C5C', end_color='243C5C', fill_type='solid')
    header_font = Font(bold=True, color='FFFFFF')
    
    # 列F～Iのヘッダー背景色
    fg_fill = PatternFill(start_color='4C5D3C', end_color='4C5D3C', fill_type='solid')

    wb = load_workbook(self.output_file)
    sheets = ['アクティブ', '停止中']

    # 監査ログの処理を最適化
    last_access_dates = {}
    try:
        # 監査ログファイルの一括処理
        audit_files = glob.glob('./audit/*.csv') + [
            f for f in glob.glob('./audit/*.zip')
        ]
        
        # audit.csvが存在する場合は追加
        if os.path.exists('audit.csv'):
            audit_files.append('audit.csv')
            self.logger.info("カレントディレクトリのaudit.csvを追加しました。")
        
        if audit_files:
            # データフレームのリストを作成
            audit_df_list = []
            for file in audit_files:
                if file.endswith('.zip'):
                    with zipfile.ZipFile(file, 'r') as zip_ref:
                        for csv_file in zip_ref.namelist():
                            if csv_file.endswith('.csv'):
                                with zip_ref.open(csv_file) as f:
                                    df = pd.read_csv(f)
                                    audit_df_list.append(df)
                                    self.logger.debug(f"zipファイル内のCSVを読み込みました: {csv_file}")
                else:
                    df = pd.read_csv(file)
                    audit_df_list.append(df)
                    self.logger.debug(f"CSVファイルを読み込みました: {file}")

            if audit_df_list:
                # 全データを結合して処理
                audit_df = pd.concat(audit_df_list, ignore_index=True)
                audit_df['Date'] = pd.to_datetime(audit_df['Date'])
                
                # 監査ログ全体の最終日時を取得
                latest_log_date = audit_df['Date'].max()
                self.logger.info(f"監査ログの最終日時: {latest_log_date}")
                
                # ユーザー情報の抽出を効率化
                mask = audit_df['User Name (account/uid)'].str.contains('/', na=False)
                valid_records = audit_df[mask].copy()
                
                # UIDの抽出を一括処理
                valid_records['uid'] = valid_records['User Name (account/uid)'].str.extract(r'/([^)]+)')
                
                # グループ化して最新の日付を取得
                latest_access = valid_records.groupby('uid')['Date'].max()
                
                # 経過日数の計算（最終ログ日時基準）
                days_since = (latest_log_date - latest_access).dt.days
                
                # 結果を辞書に格納
                for uid, date in latest_access.items():
                    last_access_dates[uid] = {
                        'date': date,
                        'days_since': days_since[uid]
                    }
                
    except Exception as e:
        self.logger.error(f"監査ログの読み込みに失敗しました: {e}")

    # 進捗表示を追加
    self.logger.info(f"監査ログの処理が完了しました。{len(last_access_dates)}件のアクセス記録を取得。")

    for sheet in sheets:
      self.logger.info(f"{sheet}シートのフォーマットを設定中...")
      ws = wb[sheet]

      # ヘッダー行（1行目）の各セルに背景色とフォントを設定（A～I列）
      for col in [COLUMN_USER_ID, COLUMN_DISCREPANCY, COLUMN_STATUS, COLUMN_LOGIN_NAME,
                  COLUMN_NAME, COLUMN_EMAIL, COLUMN_LAST_ACCESS, COLUMN_DAYS_SINCE, COLUMN_GROUPS]:
          cell = ws[f'{col}1']
          cell.fill = header_fill
          cell.font = header_font

      # 列幅の設定（ピクセル値を文字数に変換）
      column_widths_px = {
          COLUMN_USER_ID: 180,     # ユーザーID
          COLUMN_DISCREPANCY: 80,   # 相違
          COLUMN_STATUS: 80,       # ステータス
          COLUMN_LOGIN_NAME: 270,  # ログイン名
          COLUMN_NAME: 270,        # 氏名
          COLUMN_EMAIL: 334,       # メールアドレス
          COLUMN_LAST_ACCESS: 160, # 最終アクセス日
          COLUMN_DAYS_SINCE: 60,   # 経過日数
          COLUMN_GROUPS: 1195      # 所属グループ一覧
      }
      for col, px in column_widths_px.items():
          ws.column_dimensions[col].width = px_to_char(px)

      # グループごとの列をJ列以降に設定（幅は15）
      start_col_letter = 'J'
      start_col_num = column_index_from_string(start_col_letter)
      for i, group in enumerate(self.group_names, start=start_col_num):
          col_letter = get_column_letter(i)
          ws.column_dimensions[col_letter].width = 15

      # 列F～I（メールアドレス、最終アクセス日、経過日数、所属グループ一覧）のヘッダーに別背景色を設定
      for col_letter in [COLUMN_EMAIL, COLUMN_LAST_ACCESS, COLUMN_DAYS_SINCE, COLUMN_GROUPS]:
          cell = ws[f'{col_letter}1']
          cell.fill = fg_fill

      # データ行（2行目以降）のセル配置を設定
      exclude_columns = [COLUMN_LOGIN_NAME, COLUMN_NAME, COLUMN_EMAIL]
      for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
          for cell in row:
              if cell.column_letter in [COLUMN_USER_ID, COLUMN_DISCREPANCY, COLUMN_STATUS,
                                        COLUMN_LAST_ACCESS, COLUMN_DAYS_SINCE]:
                  cell.alignment = Alignment(horizontal='center', vertical='center')
              elif cell.column_letter in exclude_columns:
                  cell.alignment = Alignment(horizontal='left', vertical='center')
              else:
                  cell.alignment = Alignment(horizontal='center', vertical='center')

      # 「Administrators」グループに所属している場合は、氏名（E列）を太字にする
      for row in ws.iter_rows(min_row=2, max_row=ws.max_row, min_col=1, max_col=ws.max_column):
          group_cell = row[column_index_from_string(COLUMN_GROUPS) - 1]
          if group_cell.value:
              groups = [g.strip() for g in group_cell.value.split(',')]
              if 'Administrators' in groups:
                  name_cell = row[column_index_from_string(COLUMN_NAME) - 1]
                  name_cell.font = Font(bold=True)

      # 所属グループ一覧内に「Administrators」が含まれている場合は除去
      for row in ws.iter_rows(min_row=2, min_col=column_index_from_string(COLUMN_GROUPS),
                              max_col=column_index_from_string(COLUMN_GROUPS), max_row=ws.max_row):
          for cell in row:
              if cell.value and 'Administrators' in cell.value:
                  cell.value = cell.value.replace('Administrators', '').strip()
                  if cell.value.endswith(','):
                      cell.value = cell.value[:-1].strip()

      # 最終アクセス日と経過日数を設定
      for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        user_id = str(row[0].value)  # A列のユーザーID
        if user_id in last_access_dates:
          row[6].value = last_access_dates[user_id]['date'].strftime('%Y-%m-%d %H:%M:%S')
          row[7].value = last_access_dates[user_id]['days_since']
          
          # セルの配置を中央に
          row[6].alignment = Alignment(horizontal='center', vertical='center')
          row[7].alignment = Alignment(horizontal='center', vertical='center')

      self.logger.info(f"{sheet}シートのフォーマット設定が完了しました。")

    # グループ情報シートのフォーマット
    if 'グループ情報' in wb.sheetnames:
        ws = wb['グループ情報']
        self.logger.info("グループ情報シートのフォーマットを設定中...")
        
        # カラム幅の設定（A～F列）
        column_widths = {
            'A': 25,  # ユーザーID/ドメイン
            'B': 30,  # ログイン名/背景色
            'C': 20,  # 氏名
            'D': 35,  # メールアドレス
            'E': 10,  # 停止中
            'F': 80   # 所属グループ一覧
        }
        for col, width in column_widths.items():
            ws.column_dimensions[col].width = width
        
        # ドメイン一覧のフォーマット
        domain_title_fill = PatternFill(start_color='4472C4', end_color='4472C4', fill_type='solid')
        domain_title_font = Font(bold=True, color='FFFFFF')
        
        # ドメイン一覧ヘッダー
        ws.cell(row=1, column=1).fill = domain_title_fill
        ws.cell(row=1, column=1).font = domain_title_font
        
        # ドメイン一覧のヘッダー行
        ws.cell(row=2, column=1).fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws.cell(row=2, column=2).fill = PatternFill(start_color='D9E1F2', end_color='D9E1F2', fill_type='solid')
        ws.cell(row=2, column=1).font = Font(bold=True)
        ws.cell(row=2, column=2).font = Font(bold=True)
        
        # ドメインごとの色を設定
        domain_to_color = {}
        generated_colors = generate_similar_colors(len(self.domain_list))
        for i, domain in enumerate(self.domain_list):
            cell = ws.cell(row=3+i, column=1)
            color_cell = ws.cell(row=3+i, column=2)
            
            color = generated_colors[i]
            domain_to_color[domain] = color
            if domain != 'kirin.co.jp':
                color_sample = PatternFill(start_color=color, end_color=color, fill_type='solid')
                color_cell.fill = color_sample
            
            cell.font = Font(bold=True)
            
            # 罫線を追加
            border = Border(left=Side(style='thin'), right=Side(style='thin'), top=Side(style='thin'), bottom=Side(style='thin'))
            cell.border = border
            color_cell.border = border
        
        # 背景・フォント設定（グループ情報用）
        group_title_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
        group_title_font = Font(bold=True, color='FFFFFF')
        
        header_fill = PatternFill(start_color='243C5C', end_color='243C5C', fill_type='solid')
        header_font = Font(bold=True, color='FFFFFF')
        
        # 枠線（太線）の設定
        thick_side = Side(border_style='thick', color="000000")
        
        # シート全体を走査して、各セットごとにフォーマットを適用する
        row = len(self.domain_list) + 5  # ドメイン一覧の後に開始行を設定
        while row <= ws.max_row:
            cell_val = ws.cell(row=row, column=1).value
            if isinstance(cell_val, str) and cell_val.startswith("グループ:"):
                block_start = row
                # グループ名行の背景設定（E列まで）
                for col in range(1, 6):  # A～E列のみ
                    cell = ws.cell(row=row, column=col)
                    cell.fill = group_title_fill
                    cell.font = group_title_font
                row += 1
                
                # ヘッダー行の背景設定
                if row <= ws.max_row and ws.cell(row=row, column=1).value == "ユーザーID":
                    # A～E列は通常の青系の背景色
                    for col in range(1, 6):  # A～E列
                        cell = ws.cell(row=row, column=col)
                        cell.fill = header_fill
                        cell.font = header_font
                        cell.alignment = Alignment(horizontal='center')
                    
                    # F列（所属グループ一覧）は緑系の背景色
                    group_list_fill = PatternFill(start_color='4C5D3C', end_color='4C5D3C', fill_type='solid')
                    cell = ws.cell(row=row, column=6)  # F列
                    cell.fill = group_list_fill
                    cell.font = header_font
                    cell.alignment = Alignment(horizontal='center')
                    row += 1
                else:
                    continue
                
                # セットのデータ行の最終行を検出
                data_start = row
                while row <= ws.max_row and ws.cell(row=row, column=1).value not in [None, ""]:
                    # D列（メールアドレス）を右寄せに
                    email_cell = ws.cell(row=row, column=4)
                    email_cell.alignment = Alignment(horizontal='right')
                    
                    # E列が「●」のとき、B列の背景色を薄いグレーに設定
                    if ws.cell(row=row, column=5).value == '●':
                        b_cell = ws.cell(row=row, column=2)
                        b_cell.fill = PatternFill(start_color='D3D3D3', end_color='D3D3D3', fill_type='solid')
                    
                    # メールアドレスに基づいてセルの背景色を設定
                    if email_cell.value:
                        email_value = email_cell.value
                        domain = email_value.split('@')[-1] if '@' in email_value else ''
                        
                        # kirin.co.jp以外のドメインの場合のみ背景色を設定
                        if domain in domain_to_color and domain != 'kirin.co.jp':
                            color = domain_to_color[domain]
                            email_cell.fill = PatternFill(start_color=color, end_color=color, fill_type='solid')
                    
                    # 停止中列を中央揃えに
                    ws.cell(row=row, column=5).alignment = Alignment(horizontal='center')
                    row += 1
                block_end = row - 1
                
                # ブロック全体に太線の枠線を設定（E列まで）
                for r in range(block_start, block_end + 1):
                    for c in range(1, 6):  # A～E列のみ
                        cell = ws.cell(row=r, column=c)
                        new_border = Border(
                            left=thick_side if c == 1 else cell.border.left,
                            right=thick_side if c == 5 else cell.border.right,
                            top=thick_side if r == block_start else cell.border.top,
                            bottom=thick_side if r == block_end else cell.border.bottom
                        )
                        cell.border = new_border
                row += 1
            else:
                row += 1
        
        self.logger.info("グループ情報シートのフォーマット設定が完了しました。")
    
    wb.save(self.output_file)
    self.logger.info(f"Excelファイル '{self.output_file}' のフォーマットを設定しました。")

def generate_similar_colors(num_colors, seed=None):
    """
    元のカラーパレットに近い色合いの淡い色を生成する関数
    """
    if seed is not None:
        random.seed(seed)
    
    # 元のカラーパレットを基準とした色相（HSV）
    base_hues = [
        0.58,   # 青
        0.15,   # 黄色
        0.33,   # 緑
        0.08,   # オレンジ
        0.75,   # 紫
        0.55,   # 別の青
        0.17,   # 別の黄色
    ]
    
    # 新しい色を生成するための色相のリスト
    hues = []
    
    # 必要な数の色相を用意
    if num_colors <= len(base_hues):
        hues = base_hues[:num_colors]
    else:
        hues = base_hues.copy()
        remaining = num_colors - len(base_hues)
        step = 1.0 / remaining
        
        for i in range(remaining):
            new_hue = (i * step) % 1.0
            min_distance = min(abs(new_hue - h) % 1.0 for h in hues)
            if min_distance > 0.05:
                hues.append(new_hue)
            else:
                hues.append(random.random())
    
    colors = {}
    
    # 各色相から色を生成
    for i in range(num_colors):
        hue_idx = i % len(hues)
        hue = hues[hue_idx]
        
        # 彩度と明度を調整して淡い色に
        saturation = random.uniform(0.15, 0.25)  # 彩度を下げる（0.4-0.6 → 0.15-0.25）
        value = random.uniform(0.90, 0.95)       # 明度を上げる（0.7-0.8 → 0.90-0.95）
        
        rgb = colorsys.hsv_to_rgb(hue, saturation, value)
        rgb_int = tuple(int(255 * x) for x in rgb)
        hex_color = '{:02X}{:02X}{:02X}'.format(*rgb_int)
        
        colors[i] = hex_color
        
    return colors

def setup_logging(silent: bool, debug: bool) -> logging.Logger:
  logger = logging.getLogger("KintoneExporter")
  # 既存のハンドラーをクリア
  logger.handlers.clear()
  
  if debug:
    log_level = logging.DEBUG
  elif silent:
    log_level = logging.WARNING
  else:
    log_level = logging.INFO
    
  logger.setLevel(log_level)
  handler = logging.StreamHandler(sys.stdout)
  handler.setLevel(log_level)  # ハンドラーのレベルも設定
  formatter = logging.Formatter('%(levelname)s: %(message)s')
  handler.setFormatter(formatter)
  logger.addHandler(handler)
  
  return logger

def load_config(config_path: str) -> Dict[str, Any]:
    with open(config_path, 'r', encoding='utf-8') as file:
        return yaml.safe_load(file)

def main():
  args = ArgumentParser.parse_arguments()
  logger = setup_logging(args.silent, False)

  # auditディレクトリの作成（カレントディレクトリ直下）
  os.makedirs('./audit', exist_ok=True)
  
  # 認証情報の初期化
  subdomain = args.subdomain
  username = args.username
  password = args.password

  # 引数が指定されていない場合、デフォルトのconfig_UserAccount.yamlを使用
  if not (subdomain and username and password):
    default_config = 'config_UserAccount.yaml'
    try:
      config = load_config(default_config)
      if not subdomain:
        subdomain = config.get('subdomain')
      if not username:
        username = config.get('username')
      if not password:
        password = config.get('password')
      logger.info(f"デフォルト設定ファイル '{default_config}' から認証情報を読み込みました。")
    except Exception as e:
      logger.error(f"デフォルト設定ファイルの読み込みに失敗しました: {e}")
      sys.exit(1)

  # Kintoneクライアントの初期化
  logger.info("認証情報を設定中...")
  client = KintoneClient(subdomain, username, password, logger)

  # データの取得
  logger.info("全ユーザーを取得中...")
  all_users = client.get_all_users()

  logger.info("全グループを取得中...")
  all_groups = client.get_all_groups()

  # データの処理
  processor = DataProcessor(all_users, all_groups, client, logger)
  processor.map_users()
  filtered_groups = processor.filter_groups()
  group_names = processor.organize_groups(filtered_groups)
  processor.populate_group_memberships(filtered_groups)
  dataframes = processor.generate_dataframes(group_names)
  
  # group_user_list.yamlの生成を追加
  processor.export_group_user_list(filtered_groups)

  # Excelへのエクスポートとフォーマット
  exporter = ExcelExporter(dataframes, group_names, args.output, logger)
  exporter.prepare_group_data(client)
  exporter.export_to_excel()
  exporter.format_excel()

  logger.info(f"Excelファイル '{args.output}' の作成とフォーマット設定が完了しました。")

if __name__ == "__main__":
  main()