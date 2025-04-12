#!/usr/bin/env python
# -*- coding: utf-8 -*-

"""
Kintoneアプリの全体設定一覧表をエクセルで出力するスクリプト

このスクリプトは、kintone_runnerで取得したアプリ設定情報を元に、
アプリの全体設定一覧表をエクセルで出力します。
"""

import os
import sys
import yaml
import json
import argparse
import logging
from pathlib import Path
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Alignment, PatternFill, Border, Side, Font
from openpyxl.utils.cell import get_column_letter

# 定数定義
SCRIPT_DIR = Path(__file__).resolve().parent
OUTPUT_DIR = SCRIPT_DIR / "output"

# ログ設定
def setup_logging():
    """ロギングの設定"""
    log_dir = SCRIPT_DIR / "logs"
    log_dir.mkdir(exist_ok=True)
    
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    log_file = log_dir / f"app_settings_summary_{timestamp}.log"
    
    logging.basicConfig(
        level=logging.INFO,
        format='%(asctime)s - %(name)s - %(levelname)s - %(message)s',
        handlers=[
            logging.FileHandler(log_file),
            logging.StreamHandler()
        ]
    )
    
    return logging.getLogger("app_settings_summary")

def find_app_directories(output_dir=None):
    """
    outputディレクトリ内のアプリディレクトリを検索
    
    Returns:
        list: アプリディレクトリのリスト [(app_id, app_name, directory_path), ...]
    """
    if output_dir is None:
        output_dir = OUTPUT_DIR
    
    app_dirs = []
    
    if not output_dir.exists():
        return app_dirs
    
    for item in output_dir.iterdir():
        if item.is_dir():
            # ディレクトリ名が "数字_" で始まる場合はアプリディレクトリと判断
            dir_name = item.name
            if "_" in dir_name and dir_name.split("_")[0].isdigit():
                app_id = dir_name.split("_")[0]
                # アプリ名は2番目の要素（存在する場合）
                app_name_parts = dir_name.split("_")[1:-1]  # 最後の日時部分を除く
                app_name = "_".join(app_name_parts) if app_name_parts else "不明"
                app_dirs.append((app_id, app_name, item))
    
    return app_dirs

def load_app_settings(app_dir, app_id):
    """
    アプリディレクトリから設定ファイルを読み込む
    
    Args:
        app_dir (Path): アプリディレクトリのパス
        app_id (str): アプリID
    
    Returns:
        dict: アプリ設定情報
    """
    settings = {}
    
    # 読み込むファイルのリスト
    files_to_load = [
        f"{app_id}_settings.yaml",
        f"{app_id}_form_fields.yaml",
        f"{app_id}_form_layout.yaml",
        f"{app_id}_views.yaml",
        f"{app_id}_process_management.yaml",
        f"{app_id}_app_acl.yaml",
        f"{app_id}_field_acl.yaml",
        f"{app_id}_record_acl.yaml",
        f"{app_id}_actions.yaml",
        f"{app_id}_app_notifications.yaml",
        f"{app_id}_customize.yaml",
        f"{app_id}_plugins.yaml",
        f"{app_id}_graphs.yaml",
        f"{app_id}_general_notifications.yaml",
        f"{app_id}_record_notifications.yaml"
    ]
    
    for file_name in files_to_load:
        file_path = app_dir / file_name
        if file_path.exists():
            try:
                with open(file_path, 'r', encoding='utf-8') as f:
                    data = yaml.safe_load(f)
                    key = file_name.replace(f"{app_id}_", "").replace(".yaml", "")
                    settings[key] = data
            except Exception as e:
                logging.warning(f"ファイル {file_name} の読み込み中にエラーが発生しました: {e}")
    
    return settings

def extract_app_summary(app_id, app_name, settings):
    """
    アプリ設定から概要情報を抽出
    
    Args:
        app_id (str): アプリID
        app_name (str): アプリ名
        settings (dict): アプリ設定情報
    
    Returns:
        dict: アプリ概要情報
    """
    summary = {
        "app_id": app_id,
        "app_name": app_name,
        "description": settings.get("settings", {}).get("description", ""),
        "creator": settings.get("settings", {}).get("creator", {}).get("name", ""),
        "created_at": settings.get("settings", {}).get("createdAt", ""),
        "modifier": settings.get("settings", {}).get("modifier", {}).get("name", ""),
        "modified_at": settings.get("settings", {}).get("modifiedAt", ""),
        "theme": settings.get("settings", {}).get("theme", ""),
        "revision": settings.get("settings", {}).get("revision", ""),
        "field_count": len(settings.get("form_fields", {}).get("properties", {})),
        "view_count": len(settings.get("views", {}).get("views", {})),
        "has_process": bool(settings.get("process_management", {}).get("states", {})),
        "has_js_customize": bool(settings.get("customize", {}).get("desktop", {}).get("js", [])),
        "has_css_customize": bool(settings.get("customize", {}).get("desktop", {}).get("css", [])),
        "has_mobile_customize": bool(settings.get("customize", {}).get("mobile", {}).get("js", [])),
        "plugin_count": len(settings.get("plugins", {}).get("desktop", {}).get("plugins", [])),
        "has_actions": bool(settings.get("actions", {}).get("actions", [])),
        "has_notifications": bool(settings.get("record_notifications", {}).get("notifications", [])) or bool(settings.get("general_notifications", {}).get("notifications", [])),
        "has_graphs": bool(settings.get("graphs", {}).get("reports", [])),
    }
    
    # アプリ管理者権限を持つユーザー/グループを抽出
    app_admins = []
    for entity in settings.get("app_acl", {}).get("rights", []):
        if entity.get("appEditable", False):
            entity_type = entity.get("entity", {}).get("type", "")
            entity_code = entity.get("entity", {}).get("code", "")
            app_admins.append(f"{entity_type}:{entity_code}")
    
    summary["app_admins"] = ", ".join(app_admins)
    
    # 権限設定の数を抽出
    summary["app_acl_count"] = len(settings.get("app_acl", {}).get("rights", []))
    summary["record_acl_count"] = len(settings.get("record_acl", {}).get("rights", []))
    summary["field_acl_count"] = len(settings.get("field_acl", {}).get("rights", []))
    
    # プロセス管理の情報を抽出
    process_states = settings.get("process_management", {}).get("states", {})
    if process_states:
        # ステータス一覧を抽出
        status_list = []
        status_actions = {}
        
        for status_name, status_info in process_states.items():
            status_list.append(status_name)
            
            # ステータス毎のアクションを抽出
            actions = []
            for action_name, action_info in status_info.get("actions", {}).items():
                actions.append(action_name)
            
            status_actions[status_name] = actions
        
        summary["process_status_list"] = status_list
        summary["process_status_actions"] = status_actions
    else:
        summary["process_status_list"] = []
        summary["process_status_actions"] = {}
    
    return summary

def create_excel_summary(app_summaries, output_file=None):
    """
    アプリ概要情報からエクセルファイルを作成
    
    Args:
        app_summaries (list): アプリ概要情報のリスト
        output_file (str, optional): 出力ファイル名
    
    Returns:
        str: 作成したエクセルファイルのパス
    """
    if output_file is None:
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_file = OUTPUT_DIR / f"kintone_app_settings_summary_{timestamp}.xlsx"
    
    # ディレクトリが存在しない場合は作成
    output_file.parent.mkdir(exist_ok=True, parents=True)
    
    # Excelワークブックを作成
    wb = Workbook()
    ws = wb.active
    ws.title = "アプリ設定一覧"
    
    # ヘッダー行の設定
    headers = [
        "アプリID", "アプリ名", "説明", "作成者", "作成日時", "更新者", "更新日時",
        "テーマ", "リビジョン", "フィールド数", "一覧数", "プロセス管理",
        "JS/CSS", "モバイルJS", "プラグイン数", "アクション", "通知", "グラフ",
        "アプリ権限数", "レコード権限数", "フィールド権限数",
        "ステータス一覧", "ステータス毎のアクション",
        "アプリ管理者"
    ]
    
    # ヘッダー行のスタイル
    header_fill = PatternFill(start_color="E6F3FF", end_color="E6F3FF", fill_type="solid")
    header_font = Font(bold=True)
    header_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # ヘッダー行を設定
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = header_alignment
        cell.border = header_border
    
    # データ行のスタイル
    data_alignment = Alignment(vertical="center", wrap_text=True)
    data_border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # データ行を設定
    for row_idx, summary in enumerate(app_summaries, 2):
        # 各列のデータを設定
        ws.cell(row=row_idx, column=1, value=summary["app_id"])
        ws.cell(row=row_idx, column=2, value=summary["app_name"])
        ws.cell(row=row_idx, column=3, value=summary["description"])
        ws.cell(row=row_idx, column=4, value=summary["creator"])
        ws.cell(row=row_idx, column=5, value=summary["created_at"])
        ws.cell(row=row_idx, column=6, value=summary["modifier"])
        ws.cell(row=row_idx, column=7, value=summary["modified_at"])
        ws.cell(row=row_idx, column=8, value=summary["theme"])
        ws.cell(row=row_idx, column=9, value=summary["revision"])
        ws.cell(row=row_idx, column=10, value=summary["field_count"])
        ws.cell(row=row_idx, column=11, value=summary["view_count"])
        ws.cell(row=row_idx, column=12, value="あり" if summary["has_process"] else "なし")
        
        # JS/CSSカスタマイズ
        js_css_value = ""
        if summary["has_js_customize"] and summary["has_css_customize"]:
            js_css_value = "JS+CSS"
        elif summary["has_js_customize"]:
            js_css_value = "JSのみ"
        elif summary["has_css_customize"]:
            js_css_value = "CSSのみ"
        else:
            js_css_value = "なし"
        ws.cell(row=row_idx, column=13, value=js_css_value)
        
        ws.cell(row=row_idx, column=14, value="あり" if summary["has_mobile_customize"] else "なし")
        ws.cell(row=row_idx, column=15, value=summary["plugin_count"])
        ws.cell(row=row_idx, column=16, value="あり" if summary["has_actions"] else "なし")
        ws.cell(row=row_idx, column=17, value="あり" if summary["has_notifications"] else "なし")
        ws.cell(row=row_idx, column=18, value="あり" if summary["has_graphs"] else "なし")
        
        # 権限設定の数
        ws.cell(row=row_idx, column=19, value=summary["app_acl_count"])
        ws.cell(row=row_idx, column=20, value=summary["record_acl_count"])
        ws.cell(row=row_idx, column=21, value=summary["field_acl_count"])
        
        # プロセス管理のステータス情報
        status_list = summary.get("process_status_list", [])
        status_actions = summary.get("process_status_actions", {})
        
        # ステータス一覧
        status_text = ", ".join(status_list) if status_list else "なし"
        ws.cell(row=row_idx, column=22, value=status_text)
        
        # ステータス毎のアクション
        actions_text = ""
        for status, actions in status_actions.items():
            if actions:
                actions_text += f"{status}: {', '.join(actions)}\n"
        
        ws.cell(row=row_idx, column=23, value=actions_text)
        
        # アプリ管理者
        ws.cell(row=row_idx, column=24, value=summary["app_admins"])
        
        # 各セルにスタイルを適用
        for col_idx in range(1, len(headers) + 1):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.alignment = data_alignment
            cell.border = data_border
    
    # 列幅の設定
    column_widths = {
        1: 10,   # アプリID
        2: 30,   # アプリ名
        3: 40,   # 説明
        4: 15,   # 作成者
        5: 20,   # 作成日時
        6: 15,   # 更新者
        7: 20,   # 更新日時
        8: 15,   # テーマ
        9: 10,   # リビジョン
        10: 12,  # フィールド数
        11: 10,  # 一覧数
        12: 12,  # プロセス管理
        13: 12,  # JS/CSS
        14: 12,  # モバイルJS
        15: 12,  # プラグイン数
        16: 12,  # アクション
        17: 10,  # 通知
        18: 10,  # グラフ
        19: 12,  # アプリ権限数
        20: 12,  # レコード権限数
        21: 12,  # フィールド権限数
        22: 30,  # ステータス一覧
        23: 50,  # ステータス毎のアクション
        24: 40,  # アプリ管理者
    }
    
    for col_idx, width in column_widths.items():
        ws.column_dimensions[get_column_letter(col_idx)].width = width
    
    # 行の高さを設定
    ws.row_dimensions[1].height = 30  # ヘッダー行
    for row_idx in range(2, len(app_summaries) + 2):
        ws.row_dimensions[row_idx].height = 30  # データ行（高さを増やして複数行のテキストを表示しやすくする）
    
    # フィルターを設定
    ws.auto_filter.ref = f"A1:X{len(app_summaries) + 1}"
    
    # ファイルを保存
    wb.save(output_file)
    
    return str(output_file)

def run_app_settings_summary(output_file: str = None) -> bool:
    """
    app_settings_summary.pyのメイン処理を実行する関数
    
    Args:
        output_file (str, optional): 出力するExcelファイルの名前
        
    Returns:
        bool: 処理が成功したかどうか
    """
    # ログ設定
    logger = setup_logging()
    
    try:
        # 引数を設定
        sys.argv = ['app_settings_summary.py']
        if output_file:
            sys.argv.extend(['--output', output_file])
            
        # main関数を実行
        main()
        return True
    except Exception as e:
        logger.error(f"エラーが発生しました: {e}")
        return False

def main():
    """メイン関数"""
    # コマンドライン引数の解析
    parser = argparse.ArgumentParser(description='Kintoneアプリの全体設定一覧表をエクセルで出力するスクリプト')
    parser.add_argument('--output', type=str, help='出力ファイル名')
    args = parser.parse_args()
    
    # ロギングの設定
    logger = setup_logging()
    logger.info("アプリ設定一覧表の作成を開始します")
    
    # アプリディレクトリを検索
    app_dirs = find_app_directories()
    if not app_dirs:
        logger.error("アプリディレクトリが見つかりませんでした")
        print("エラー: アプリディレクトリが見つかりませんでした。")
        print("kintone_runner.py app コマンドを実行して、アプリ設定を取得してください。")
        sys.exit(1)
    
    logger.info(f"{len(app_dirs)}個のアプリディレクトリを検出しました")
    
    # アプリ設定を読み込み
    app_summaries = []
    for app_id, app_name, app_dir in app_dirs:
        logger.info(f"アプリID {app_id} ({app_name}) の設定を読み込みます")
        settings = load_app_settings(app_dir, app_id)
        if settings:
            summary = extract_app_summary(app_id, app_name, settings)
            app_summaries.append(summary)
            logger.info(f"アプリID {app_id} の設定情報を抽出しました")
        else:
            logger.warning(f"アプリID {app_id} の設定情報が見つかりませんでした")
    
    if not app_summaries:
        logger.error("有効なアプリ設定が見つかりませんでした")
        print("エラー: 有効なアプリ設定が見つかりませんでした。")
        sys.exit(1)
    
    # 出力ファイル名の設定
    output_file = args.output
    if output_file:
        output_file = Path(output_file)
    
    # エクセルファイルを作成
    try:
        excel_file = create_excel_summary(app_summaries, output_file)
        logger.info(f"アプリ設定一覧表を {excel_file} に出力しました")
        print(f"アプリ設定一覧表を {excel_file} に出力しました")
    except Exception as e:
        logger.error(f"エクセルファイルの作成中にエラーが発生しました: {e}")
        print(f"エラー: エクセルファイルの作成中にエラーが発生しました: {e}")
        sys.exit(1)

if __name__ == "__main__":
    main() 